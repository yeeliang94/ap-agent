"""Survey + peek: everything the map AI is allowed to see, gathered by code.

The survey lists the batch: per subfolder, every file with its type, size,
page count and the ER(...) code in its name. The peek looks INSIDE each
file cheaply — a workbook's tab names and the first rows of each tab as
text, a PDF's or image's first page as a small thumbnail — so the agent
can find things by itself rather than being told where to look. No model
call happens here; the AI sees this survey next (mapping.py).
"""
from __future__ import annotations

import io
import re
from pathlib import Path

from . import source as batch_source

ER_CODE = re.compile(r"ER\s*\(([^)]*)\)", re.IGNORECASE)

# How much of a workbook tab the peek shows: enough to recognise a header
# block and the first expense lines, not the whole sheet.
PEEK_ROWS = 15
PEEK_COLS = 12
PEEK_CELL_CHARS = 60

# Thumbnail size (longest edge). Small on purpose: this is "what does page
# 1 look like", not a read. The workers read pages properly later.
THUMB_EDGE = 640

WORKBOOK = {".xlsx", ".xlsm"}
PDF = {".pdf"}
IMAGE = {".png", ".jpg", ".jpeg", ".webp"}


def file_type(name: str) -> str:
    suffix = Path(name).suffix.lower()
    if suffix in WORKBOOK:
        return "workbook"
    if suffix in PDF:
        return "pdf"
    if suffix in IMAGE:
        return "image"
    return "other"


def er_code_of(name: str) -> str:
    """'Aegene Ong_ER(01JUL26-21JUL26).xlsx' → 'ER(01JUL26-21JUL26)'."""
    m = ER_CODE.search(name)
    return f"ER({m.group(1).strip()})" if m else ""


def survey_batch(dest: Path, files: list[dict]) -> dict:
    """The survey + peeks for every file copied into dest.

    Enforces the pages-per-employee quota once page counts are known
    (raises QuotaExceeded). Thumbnails are written under dest/../peeks/.
    """
    peeks_dir = dest.parent / "peeks"
    peeks_dir.mkdir(parents=True, exist_ok=True)
    out_files: list[dict] = []
    folders: dict[str, dict] = {}
    for i, entry in enumerate(sorted(files, key=lambda f: f["path"])):
        rel = entry["path"]
        path = dest / rel
        top = rel.split("/", 1)[0] if "/" in rel else ""
        kind = file_type(rel)
        record = {
            "path": rel, "name": Path(rel).name, "folder": top, "type": kind,
            "size": path.stat().st_size if path.is_file() else entry.get("size"),
            "pages": batch_source.page_count(path) if path.is_file() else None,
            "er_code": er_code_of(Path(rel).name), "depth": entry.get("depth", 1),
            "peek": None,
        }
        try:
            if kind == "workbook":
                record["peek"] = {"tabs": peek_workbook(path)}
            elif kind in ("pdf", "image"):
                thumb = peeks_dir / f"{i:03d}.png"
                thumb.write_bytes(thumbnail(path))
                record["peek"] = {"thumbnail": f"peeks/{thumb.name}"}
        except Exception as exc:  # a broken file is surveyed, not fatal
            record["peek_error"] = f"{type(exc).__name__}: {str(exc)[:200]}"
        out_files.append(record)
        if top:
            folders.setdefault(top, {"path": top, "name": top, "files": []})
            folders[top]["files"].append(rel)
    batch_source.check_page_quotas(out_files)
    return {"folders": [folders[k] for k in sorted(folders)], "files": out_files,
            "root_files": [f["path"] for f in out_files if not f["folder"]]}


def peek_workbook(path: Path) -> dict[str, list[str]]:
    """Tab name → the first PEEK_ROWS rows of that tab as text lines
    ("A1: Name: | B1: Aegene Ong")."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    tabs: dict[str, list[str]] = {}
    try:
        for ws in wb.worksheets:
            lines = []
            for r, row in enumerate(ws.iter_rows(min_row=1, max_row=PEEK_ROWS,
                                                 max_col=PEEK_COLS, values_only=False), 1):
                cells = []
                for cell in row:
                    v = cell.value
                    if v is None or str(v).strip() == "":
                        continue
                    text = str(v).strip().replace("\n", " ")[:PEEK_CELL_CHARS]
                    cells.append(f"{cell.coordinate}: {text}")
                if cells:
                    lines.append(" | ".join(cells))
            tabs[ws.title] = lines
    finally:
        wb.close()
    return tabs


def thumbnail(path: Path) -> bytes:
    """Page 1 as a small PNG."""
    from PIL import Image

    suffix = path.suffix.lower()
    if suffix == ".pdf":
        import pymupdf

        with pymupdf.open(path) as pdf:
            if pdf.page_count == 0:
                raise ValueError("empty PDF")
            pix = pdf[0].get_pixmap(dpi=72)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
    else:
        img = Image.open(path)
    img.thumbnail((THUMB_EDGE, THUMB_EDGE))
    buf = io.BytesIO()
    img.convert("RGB").save(buf, format="PNG")
    return buf.getvalue()


def survey_text(survey: dict, workspace: Path | None = None) -> str:
    """The survey as the map AI reads it: one block per folder, each file
    with its facts and its peek (tab rows as text; thumbnails are attached
    as images by the caller)."""
    lines = []
    files_by_folder: dict[str, list[dict]] = {}
    for f in survey["files"]:
        files_by_folder.setdefault(f["folder"], []).append(f)
    for folder in [fo["path"] for fo in survey["folders"]] + ([""] if files_by_folder.get("") else []):
        lines.append(f"\n## Folder: {folder or '(batch root — files not inside any subfolder)'}")
        for f in files_by_folder.get(folder, []):
            facts = [f["type"], f"{(f['size'] or 0) / 1024:.0f} KB"]
            if f.get("pages") is not None:
                facts.append(f"{f['pages']} page(s)")
            if f.get("er_code"):
                facts.append(f"ER code in name: {f['er_code']}")
            lines.append(f"- File `{f['path']}` ({', '.join(facts)})")
            peek = f.get("peek") or {}
            if "tabs" in peek:
                for tab, rows in peek["tabs"].items():
                    lines.append(f"    tab {tab!r}: first rows:")
                    for row in rows[:PEEK_ROWS]:
                        lines.append(f"      {row[:300]}")
            elif "thumbnail" in peek:
                lines.append("    (page-1 thumbnail attached below, labelled with this path)")
            if f.get("peek_error"):
                lines.append(f"    (could not look inside: {f['peek_error']})")
    return "\n".join(lines)
