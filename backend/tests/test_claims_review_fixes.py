"""Regression tests for the 2026-08-19 review — the domain / business-logic
slice (docs/REVIEW-2026-08-19.md, items 1–6 and the domain suggestions).

Each test pins one finding: a receipt claimed by rows that was also
"unclaimed"; foreign-currency derived lines summed as MYR; the prefix
elimination that hid a three-name conflict; document titles read as
names; the second read's extra receipt dropped while the note said it was
added; the unbounded row span; the Reported Total standing in for the
lines total; the bulk-confirm count; Decimal money in the read models;
foreign unclaimed receipts; cumulative zip bytes; the blank name cell;
and the nits (serial numbers, the mirrored cite, input rows untouched,
MILEAGE_NO_MAP behind its switch).
"""
from __future__ import annotations

import asyncio
import copy
import io
import zipfile
from decimal import Decimal

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.claims import cases as cases_mod
from app.claims import checks, evidence, grouping, listing, report_reader, source, worker
from app.claims.models import (ClaimCase, ClaimEmployee, ClaimEvidence, ClaimFlag, ClaimRow,
                               ClaimSourceArtifact, ClaimsRun)
from app.claims.report_reader import ReportColumns, ReportReading
from app.db import Base

PROFILE = {"mileage_rates": {"Car": "0.64"}, "km_tolerance": "0", "receipt_date_window_days": 0,
           "receipt_optional_items": [], "mileage_item_pattern": "mileage", "categories": [],
           "category_rule": "", "file_role_patterns": [], "checks": {}, "set_by": {}}


@pytest.fixture()
def Session(tmp_path, monkeypatch):
    engine = create_engine(f"sqlite:///{tmp_path / 't.sqlite3'}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    S = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    monkeypatch.setattr(worker, "SessionLocal", S)
    return S


def _row(rid, date_, amount, item="Taxi", kind="expense", **over):
    v = {"date": date_, "item": item, "item_name": item, "reason": "x", "receipt_included": "Y",
         "amount": amount, "currency": "MYR", "rate": "1", "total": amount}
    v.update(over)
    return {"id": rid, "kind": kind, "sheet": "S", "row": int(rid[1:]) + 6, "values": v}


def _receipt(eid, date_, amount, page=1, vendor="Grab", position="left", currency="MYR"):
    return {"id": eid, "kind": "receipt", "file": "r.pdf", "page": page, "position": position,
            "values": {"vendor": vendor, "date": date_, "amount": amount, "currency": currency}, "confidence": {}}


def _checks(rows, ev, profile=PROFILE, tie=None):
    async def unsure(row, cs):
        return ""
    return asyncio.run(checks.run_checks(rows, ev, profile, {}, (1, 1), tie_break=tie or unsure))


def _run_with_case(s, *, claimant="A", reported_total="", status="verified", listing_headers=None):
    run = ClaimsRun(client="c", folder_url="", listing_url="", received_date="2026-08-19",
                    instructions="", snapshot={}, status="ready",
                    listing_headers=listing_headers or {"state": "missing"})
    s.add(run)
    s.commit()
    e = ClaimEmployee(run_id=run.id, folder=f"{claimant}_1", name=claimant, er_code="ER(1)", roles={},
                      status=status, report_total=reported_total, category="Travel", gl="6100")
    s.add(e)
    s.commit()
    case = cases_mod.sync_case_from_employee(s, e)
    s.commit()
    return run, e, case


# ---- 1. a receipt claimed by rows is never also "unclaimed" -------------------------

def test_duplicate_receipt_is_not_also_unclaimed():
    rows = [_row("r1", "2026-07-01", "150.00"), _row("r2", "2026-07-01", "150.00")]
    ev = [_receipt("e1", "2026-07-01", "150.00")]
    res = _checks(rows, ev)
    assert {f["code"] for f in res["flags"] if f["row_id"] in ("r1", "r2")} == {"DUPLICATE_RECEIPT"}
    assert res["verdicts"]["r1"] == ("duplicate", "e1") and res["verdicts"]["r2"] == ("duplicate", "e1")
    assert not [f for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"]
    assert set(res["contested"]) == {"e1"} and set(res["contested"]["e1"]) == {"r1", "r2"}
    # exactly at the threshold: still a duplicate, still not unclaimed
    rows = [_row("r1", "2026-07-01", "100.00"), _row("r2", "2026-07-01", "100.00")]
    res = _checks(rows, [_receipt("e1", "2026-07-01", "100.00")])
    assert not [f for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"]


def test_ambiguous_candidates_are_not_also_unclaimed():
    rows = [_row("r1", "2026-07-01", "150.00")]
    ev = [_receipt("e1", "2026-07-01", "150.00", vendor="Grab"),
          _receipt("e2", "2026-07-01", "150.00", vendor="AirAsia Ride", position="right"),
          _receipt("e3", "2026-07-02", "150.00", vendor="Hilton", page=2)]
    res = _checks(rows, ev)
    codes = {f["code"]: f for f in res["flags"]}
    assert "RECEIPT_AMBIGUOUS" in codes
    unclaimed = [f for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"]
    assert [f["evidence_id"] for f in unclaimed] == ["e3"]  # the third is nobody's: still unclaimed
    assert set(res["contested"]) == {"e1", "e2"}
    assert set(codes["RECEIPT_AMBIGUOUS"]["cite"]["candidates"]) == {"e1", "e2"}


def test_contested_evidence_is_not_listed_as_unused(Session):
    s = Session()
    run, e, case = _run_with_case(s, reported_total="300.00")
    r1 = ClaimRow(run_id=run.id, employee_id=e.id, case_id=case.id, kind="expense", sheet="S", row=7,
                  values={"amount": "150.00", "total": "150.00"}, verdict="duplicate", matched_evidence_id="")
    r2 = ClaimRow(run_id=run.id, employee_id=e.id, case_id=case.id, kind="expense", sheet="S", row=8,
                  values={"amount": "150.00", "total": "150.00"}, verdict="duplicate", matched_evidence_id="")
    r3 = ClaimRow(run_id=run.id, employee_id=e.id, case_id=case.id, kind="expense", sheet="S", row=9,
                  values={"amount": "20.00", "total": "20.00"}, verdict="ambiguous", matched_evidence_id="")
    s.add_all([r1, r2, r3])
    s.commit()
    dup = ClaimEvidence(run_id=run.id, employee_id=e.id, case_id=case.id, kind="receipt", file="r.pdf", page=1,
                        position="left", values={"vendor": "V", "date": "2026-07-01", "amount": "150.00",
                                                  "currency": "MYR"}, matched_row_id="")
    c1 = ClaimEvidence(run_id=run.id, employee_id=e.id, case_id=case.id, kind="receipt", file="r.pdf", page=2,
                       position="left", values={"vendor": "A", "date": "2026-07-02", "amount": "20.00",
                                                 "currency": "MYR"}, matched_row_id="")
    c2 = ClaimEvidence(run_id=run.id, employee_id=e.id, case_id=case.id, kind="receipt", file="r.pdf", page=2,
                       position="right", values={"vendor": "B", "date": "2026-07-02", "amount": "20.00",
                                                  "currency": "MYR"}, matched_row_id="")
    lone = ClaimEvidence(run_id=run.id, employee_id=e.id, case_id=case.id, kind="receipt", file="r.pdf", page=3,
                         position="middle", values={"vendor": "Hilton", "date": "2026-07-03", "amount": "240.00",
                                                    "currency": "MYR"}, matched_row_id="")
    s.add_all([dup, c1, c2, lone])
    s.commit()
    r1.matched_evidence_id = r2.matched_evidence_id = dup.id  # the duplicate verdict names its receipt
    s.add(ClaimFlag(run_id=run.id, employee_id=e.id, case_id=case.id, code="DUPLICATE_RECEIPT", reason="", basis="",
                    cite={"file": "r.pdf", "page": 1, "position": "left"}, row_id=r1.id, evidence_id=dup.id))
    s.add(ClaimFlag(run_id=run.id, employee_id=e.id, case_id=case.id, code="RECEIPT_AMBIGUOUS", reason="", basis="",
                    cite={"file": "r.pdf", "page": 2, "position": "left", "candidates": [c1.id, c2.id]},
                    row_id=r3.id, evidence_id=c1.id))
    s.add(ClaimFlag(run_id=run.id, employee_id=e.id, case_id=case.id, code="UNCLAIMED_RECEIPT", reason="", basis="",
                    cite={"file": "r.pdf", "page": 3}, evidence_id=lone.id, status="open"))
    s.commit()
    out = listing.build_outputs(s, run)
    assert [u["what"] for u in out["unused_evidence"]] == ["receipt from Hilton (2026-07-03)"]


# ---- 2. foreign-currency derived lines are not summed as MYR -----------------------

def test_foreign_derived_line_without_a_total_is_held_out_of_the_sum(Session):
    s = Session()
    run, e, case = _run_with_case(s)
    s.add(ClaimRow(run_id=run.id, employee_id=e.id, case_id=case.id, kind="derived", sheet="", row=1,
                   values={"amount": "10.00", "currency": "MYR", "total": "10.00"}, verdict="matched"))
    usd = ClaimRow(run_id=run.id, employee_id=e.id, case_id=case.id, kind="derived", sheet="", row=2,
                   values={"amount": "50.00", "currency": "USD", "total": None}, verdict="matched")
    s.add(usd)
    s.commit()
    out = listing.build_outputs(s, run)
    assert out["totals"]["lines_total"] == "10.00" and out["totals"]["total_myr"] == "10.00"
    inc = out["included"][0]
    assert inc["amount"] == "10.00" and inc["lines_total"] == "10.00"
    held = [n for n in out["not_included"] if n["case_id"] == case.id]
    assert len(held) == 1 and "USD 50.00" in held[0]["why"] and "rate" in held[0]["why"]
    assert out["totals"]["held_lines"] == 1
    # a MYR total on the line (a rate applied) brings it back into the sum
    usd.values = {**usd.values, "rate": "4.20", "total": "210.00"}
    s.commit()
    out = listing.build_outputs(s, run)
    assert out["totals"]["lines_total"] == "220.00" and out["totals"]["held_lines"] == 0
    assert not [n for n in out["not_included"] if n["case_id"] == case.id]


def test_a_case_whose_only_lines_are_foreign_without_totals_pays_nothing(Session):
    s = Session()
    run, e, case = _run_with_case(s)
    s.add(ClaimRow(run_id=run.id, employee_id=e.id, case_id=case.id, kind="derived", sheet="", row=1,
                   values={"amount": "50.00", "currency": "USD", "total": None}, verdict="matched"))
    s.commit()
    out = listing.build_outputs(s, run)
    assert out["rows"] == [] and out["included"] == []
    assert any(n["case_id"] == case.id and "no lines to pay" in n["why"] for n in out["not_included"])


# ---- suggestion: a verified case with no lines never pays its Reported Total -----------

def test_verified_case_with_no_lines_is_not_paid_from_the_reported_total(Session):
    s = Session()
    run, e, case = _run_with_case(s, reported_total="258.70")
    s.add(ClaimRow(run_id=run.id, employee_id=e.id, case_id=case.id, kind="mileage", sheet="KM", row=5,
                   values={"km": "10", "rate": "0.64", "amount": "6.40"}, verdict="matched"))
    s.commit()
    out = listing.build_outputs(s, run)
    assert out["rows"] == [] and out["included"] == []
    assert out["totals"]["total_myr"] == "0.00" and out["totals"]["lines_total"] == "0.00"
    ni = next(n for n in out["not_included"] if n["case_id"] == case.id)
    assert "no lines to pay" in ni["why"]


# ---- nit: serial numbers count only the included cases ------------------------------

def test_serial_numbers_count_only_included_cases(Session):
    s = Session()
    run, e, case = _run_with_case(s, claimant="Bob")
    s.add(ClaimRow(run_id=run.id, employee_id=e.id, case_id=case.id, kind="expense", sheet="S", row=7,
                   values={"amount": "10.00", "total": "10.00"}, verdict="matched"))
    e2 = ClaimEmployee(run_id=run.id, folder="Alice_1", name="Alice", er_code="ER(2)", roles={},
                       status="failed", error="boom")
    s.add(e2)
    s.commit()
    cases_mod.sync_case_from_employee(s, e2)
    s.commit()
    out = listing.build_outputs(s, run)
    assert len(out["rows"]) == 1
    serial_idx = listing.FALLBACK_ROLES.index("vendor_name")  # fallback header has no S/N: use the row order
    assert out["rows"][0][serial_idx] == "Bob"
    run.listing_headers = {"state": "ok", "tab": "JUL", "header": ["S/N", "Name", "Amount"],
                           "roles": {"serial": 0, "vendor_name": 1, "amount": 2}, "past_examples": []}
    s.commit()
    out = listing.build_outputs(s, run)
    assert out["rows"][0][0] == "1"  # Alice (failed) is sorted first but takes no number


# ---- 3 + 4. ownership conflicts ----------------------------------------------------

def _art(run_id, aid, path, case_id="", role="receipts", disp="used", media="pdf"):
    return ClaimSourceArtifact(run_id=run_id, artifact_id=aid, path=path, case_id=case_id, proposed_role=role,
                               disposition=disp, media_type=media)


def test_three_names_with_a_prefix_pair_still_conflict():
    run = ClaimsRun(id="r", client="c", survey={"files": [
        {"path": "A/Aegene Ong_ER(01JUL26-21JUL26).xlsx", "peek": {"tabs": {"Expense Report": [
            "A1: Name: | B1: Aegene Ong Li"]}}},
        {"path": "B/Nick Goh_ER(02JUL26-22JUL26).xlsx", "peek": {"tabs": {"Expense Report": [
            "A1: Name: | B1: Nick Goh"]}}}]})
    arts = [_art("r", "a1", "A/Aegene Ong_ER(01JUL26-21JUL26).xlsx", "c1", "report", media="workbook"),
            _art("r", "a2", "A/Aegene Ong_receipts.pdf", "c1"),
            _art("r", "a3", "B/Nick Goh_receipts.pdf", "c1")]
    sig = grouping.signals_for(run, arts)
    names = {x["value"] for a in ("a1", "a2", "a3") for x in sig[a] if x["kind"] == "name" and x["strength"] == "strong"}
    assert {"Aegene Ong", "Aegene Ong Li", "Nick Goh"} <= names
    why = grouping.conflict_in(ClaimCase(id="c1", run_id="r", label="A"), arts, sig)
    assert "different names" in why and "Nick Goh" in why and "Aegene Ong" in why
    # the same three minus Nick: one person, no conflict
    arts2 = [a for a in arts if a.artifact_id != "a3"]
    assert grouping.conflict_in(ClaimCase(id="c1", run_id="r", label="A"), arts2, grouping.signals_for(run, arts2)) == ""


def test_document_titles_are_not_names():
    run = ClaimsRun(id="r", client="c", survey={"files": [
        {"path": "Aegene Ong_ER(01JUL26-21JUL26).xlsx", "peek": {"tabs": {"Expense Report": [
            "A1: Name: | B1: Aegene Ong"]}}}]})
    arts = [_art("r", "a1", "Aegene Ong_ER(01JUL26-21JUL26).xlsx", "c1", "report", media="workbook"),
            _art("r", "a2", "Aegene Ong_receipts.pdf", "c1"),
            _art("r", "a3", "Expense Report_July.pdf", "c1"),
            _art("r", "a4", "Claim Form_ER(01JUL26-21JUL26).xlsx", "c1", media="workbook"),
            _art("r", "a5", "Mileage Summary_July.pdf", "c1")]
    sig = grouping.signals_for(run, arts)
    for aid in ("a3", "a4", "a5"):
        assert not [x for x in sig[aid] if x["kind"] == "name"], sig[aid]
    assert grouping.conflict_in(ClaimCase(id="c1", run_id="r", label="A"), arts, sig) == ""
    # a file-name prefix that no header cell corroborates is a hint, not ownership
    arts2 = [_art("r", "b1", "Travel Claim_July.pdf", "c2"), _art("r", "b2", "Aegene Ong_receipts.pdf", "c2")]
    sig2 = grouping.signals_for(run, arts2)
    assert [x["strength"] for x in sig2["b1"] if x["kind"] == "name"] in ([], ["weak"])
    assert grouping.conflict_in(ClaimCase(id="c2", run_id="r", label="B"), arts2, sig2) == ""


# ---- 5. the second read's extra receipt is kept -----------------------------------

def _pr(kind, receipts):
    return evidence.PageRead(kind=kind, receipts=receipts, trips=[], why="x")


def _rr(vendor, amount, position, date="2026-07-01"):
    return evidence.ReceiptRead(vendor=vendor, date=date, amount=amount, currency="MYR", position=position)


def test_extra_receipt_seen_by_the_second_read_is_kept():
    a = _pr("receipts", [_rr("Grab", 20, "left")])
    b = _pr("receipts", [_rr("Grab", 20, "left"), _rr("KFC", 15.5, "left")])
    out, notes = evidence._merge_receipt_reads(a, b)
    assert len(out) == 2, out
    extra = out[1]
    assert extra["vendor"] == "KFC" and extra["amount"] == "15.50"
    assert extra["confidence"]["receipt"] == "only one of two reads saw this receipt"
    assert any("extra one(s) were added" in n for n in notes)


def test_read_models_carry_decimal_money_and_round_half_up():
    r = evidence.ReceiptRead.model_validate_json('{"vendor": "V", "amount": 2.345, "currency": "MYR", "position": "left"}')
    assert isinstance(r.amount, Decimal) and evidence._money_text(r.amount) == "2.35"
    r2 = evidence.ReceiptRead.model_validate_json('{"vendor": "V", "amount": "45.10", "currency": "MYR", "position": "left"}')
    assert r2.amount == Decimal("45.10")
    t = evidence.TripRead.model_validate_json('{"date": "", "return_trip": false, "km_printed": 12.35}')
    assert isinstance(t.km_printed, Decimal)
    assert evidence.PageRead.model_validate(evidence.PageRead(kind="map", trips=[t], why="x").model_dump(mode="json"))
    merged, _ = evidence._merge_trip_reads(_pr("map", []).model_copy(update={"trips": [t]}),
                                           _pr("map", []).model_copy(update={"trips": [t]}))
    assert merged[0]["km_printed"] == "12.4"
    out, _ = evidence._merge_receipt_reads(_pr("receipts", [_rr("V", Decimal("0.005"), "left")]),
                                           _pr("receipts", [_rr("V", Decimal("0.005"), "left")]))
    assert out[0]["amount"] == "0.01"


# ---- 6. the row span is bounded ------------------------------------------------------

def _sheet(lines=3):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    ws["A1"], ws["B1"] = "Name:", "Aegene Ong"
    for c, h in zip("ABCDEFGH", ["Date", "Item", "Reason", "Receipt?", "Amount", "Curr", "Rate", "Total"]):
        ws[f"{c}6"] = h
    r = 7
    for i in range(lines):
        ws[f"A{r}"], ws[f"B{r}"], ws[f"E{r}"], ws[f"F{r}"], ws[f"G{r}"], ws[f"H{r}"] = \
            f"2026-07-0{i + 1}", "Taxi", 10, "MYR", 1, 10
        r += 1
    ws[f"G{r + 1}"], ws[f"H{r + 1}"] = "Total", 10 * lines
    return ws, r - 1, r + 1


def _reading(last_row, total_row, **over):
    base = dict(columns=ReportColumns(date="A", item="B", reason="C", receipt_included="D", amount="E",
                                      currency="F", rate="G", total="H"),
                header_row=6, first_row=7, last_row=last_row, total_cell=f"H{total_row}", name_cell="B1",
                why="x")
    return ReportReading(**{**base, **over})


def test_row_span_is_clamped_and_wide_spans_refused():
    ws, last, total_row = _sheet()
    with pytest.raises(Exception):
        ReportReading.model_validate(_reading(last, total_row).model_dump() | {"last_row": 10**9})
    # a last_row past the sheet's end: clamped, no cells created, the lines still read
    reading = _reading(200, total_row)
    rows = report_reader.extract_rows(ws, reading)
    assert len(rows) == 3 and ws.max_row < 20
    problems = report_reader.audit_report(ws, reading, rows, "Aegene Ong", "")
    assert not [t for k, t in problems if k == report_reader.STRUCTURE], problems
    # a total cell far below the sheet does not make the audit walk there
    reading = _reading(last, total_row, total_cell=f"H{report_reader.MAX_ROW}")
    rows = report_reader.extract_rows(ws, reading)
    problems = report_reader.audit_report(ws, reading, rows, "Aegene Ong", "")
    assert ws.max_row < 20 and any("holds no number" in t for _, t in problems)
    # a span wider than the reader supports is refused before a cell is read
    wide = _reading(7 + report_reader.MAX_SHEET_ROWS + 5, total_row)   # first_row 7, last_row far below
    with pytest.raises(ValueError) as exc:
        report_reader.extract_rows(ws, wide)
    assert "wider" in str(exc.value)
    problems = report_reader.audit_report(ws, wide, [], "Aegene Ong", "")
    assert any(k == report_reader.STRUCTURE and "wider" in t for k, t in problems)
    assert ws.max_row < 20
    km = report_reader.KMReading(has_trips=True, columns=report_reader.KMColumns(date="A", km="E", amount="H"),
                                 header_row=6, first_row=7, last_row=200, why="x")
    assert report_reader.extract_trips(ws, km) and ws.max_row < 20
    km_wide = report_reader.KMReading(has_trips=True, columns=report_reader.KMColumns(date="A", km="E", amount="H"),
                                      header_row=6, first_row=7, last_row=7 + report_reader.MAX_SHEET_ROWS + 5, why="x")
    with pytest.raises(ValueError):
        report_reader.extract_trips(ws, km_wide)
    assert any("wider" in t for t in report_reader.audit_km(ws, km_wide, []))


def test_blank_name_cell_fails_the_name_audit():
    ws, last, total_row = _sheet()
    ws["B1"] = None
    reading = _reading(last, total_row, name_cell="B1")
    rows = report_reader.extract_rows(ws, reading)
    problems = report_reader.audit_report(ws, reading, rows, "Aegene Ong", "")
    assert any("name_cell" in t and "B1" in t for _, t in problems)
    ws["B1"] = "   "
    problems = report_reader.audit_report(ws, reading, rows, "Aegene Ong", "")
    assert any("name_cell" in t for _, t in problems)


# ---- suggestion: Confirm grouping says how many claimants it confirms -----------------

def test_confirm_grouping_reports_the_claimants_it_confirms(Session):
    s = Session()
    run = ClaimsRun(client="c", folder_url="", listing_url="", received_date="2026-08-19",
                    instructions="", snapshot={}, status="map_ready", survey={"files": []})
    s.add(run)
    s.commit()
    for i, (name, state) in enumerate([("Aegene Ong", "proposed"), ("Nick Goh", "proposed"), ("Zed", "confirmed")]):
        c = ClaimCase(id=f"c{i}", run_id=run.id, label=name, claimant_name=name, claimant_state=state,
                      state="proposed", roles={"no_report": True, "receipt_files": ["x.pdf"]})
        s.add(c)
        s.add(_art(run.id, f"a{i}", f"{name}_receipts.pdf", f"c{i}"))
    s.commit()
    g = grouping.gate(s, run)
    assert g["ok"] and g["counts"]["claimants_to_confirm"] == 2
    n, g = cases_mod.confirm_grouping(s, run)
    assert n == 3 and g["counts"]["claimants_to_confirm"] == 2
    assert {c.claimant_state for c in s.query(ClaimCase).filter(ClaimCase.run_id == run.id)} == {"confirmed"}


def test_mirror_clears_a_stale_reported_total_cite():
    case = ClaimCase(id="c", run_id="r", label="A", reported_total="100.00", reported_total_cite={"sheet": "S", "cell": "H9"})
    emp = ClaimEmployee(run_id="r", folder="A", name="A", report_total="100.00", status="verified")
    cases_mod._mirror(case, emp)
    assert case.reported_total_cite == {"sheet": "S", "cell": "H9"}  # same figure: the cite still holds
    emp.report_total = "120.00"
    cases_mod._mirror(case, emp)
    assert case.reported_total == "120.00" and case.reported_total_cite == {}
    case.reported_total_cite = {"sheet": "S", "cell": "H9"}
    emp.report_total = ""
    cases_mod._mirror(case, emp)
    assert case.reported_total == "" and case.reported_total_cite == {}


# ---- suggestion: a foreign unclaimed receipt is always a decision -------------------

def test_foreign_unclaimed_receipt_is_always_open():
    res = _checks([], [_receipt("e1", "2026-07-01", "5.00", currency="USD")])
    unc = [f for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"]
    assert len(unc) == 1 and unc[0]["status"] == "open" and "USD" in unc[0]["reason"]
    res = _checks([], [_receipt("e1", "2026-07-01", "5.00")])
    assert [f["status"] for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"] == ["info"]


# ---- nits: input rows untouched; MILEAGE_NO_MAP behind its switch --------------------

def test_checks_do_not_mutate_the_input_rows():
    line = _row("r1", "2026-07-01", "6.40", item="Mileage claim")
    km = {"id": "k1", "kind": "mileage", "sheet": "KM", "row": 5,
          "values": {"date": "2026-07-01", "from": "A", "to": "B", "km": "10", "rate": "0.64", "amount": "6.40"}}
    rows = [line, km]
    before = copy.deepcopy(rows)
    res = _checks(rows, [])
    assert rows == before
    assert res["verdicts"]["r1"][0] == "matched"
    assert res["km_pairs"] == {"r1": "k1"}


def test_unmatched_map_trip_note_is_gated_by_the_switch():
    trip = {"id": "t1", "kind": "map_trip", "file": "m.pdf", "page": 1, "position": "",
            "values": {"date": "2026-07-01", "purpose": "p", "from": "A", "to": "B", "return_trip": False,
                       "km_printed": "12.0"}, "confidence": {}}
    res = _checks([], [trip])
    assert [f["code"] for f in res["flags"]] == ["MILEAGE_NO_MAP"]
    res = _checks([], [trip], {**PROFILE, "checks": {"MILEAGE_NO_MAP": False}})
    assert res["flags"] == []


# ---- suggestion: the zip's cumulative bytes are capped as they are written ----------

def test_unpack_zip_caps_the_bytes_actually_written(tmp_path, monkeypatch):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for i in range(3):
            z.writestr(f"A/f{i}.bin", b"\0" * (1024 * 1024))
    zp = tmp_path / "b.zip"
    zp.write_bytes(buf.getvalue())
    # the headers are taken at their word here; what comes out is what counts
    monkeypatch.setattr(source, "_check_listing_quotas", lambda entries: None)
    monkeypatch.setattr(source, "MAX_TOTAL_MB", 2)
    with pytest.raises(source.QuotaExceeded) as exc:
        source.unpack_zip(zp, tmp_path / "dest")
    assert "MB limit for a run" in str(exc.value)
    assert not list((tmp_path / "dest").rglob("f2.bin"))
