"""Reading one employee's expense report (and its KM tab).

Same division of labour as the payment-listing reader, strictly kept:

  - The AI READS STRUCTURE ONLY ("judge" role, temperature 0). It sees the
    tab as a labelled grid ('B3: Aegene Ong') and answers with
    coordinates: which column is the date / item / reason / receipt
    included / amount / currency / rate / total; which rows are expense
    lines; where the total is; where the name, period and business-reason
    header cells are. It never retypes a number.
  - CODE pulls every value out of the cells the AI pointed at and audits
    the reading against the file's own arithmetic: amount × rate = total
    on every line (to the cent), the lines sum to the total, the header
    name is the employee the map named, the dates fall inside the ER
    period. What fails goes back to the AI with the mismatch, up to
    MAX_ROUNDS. Still failing → ReportUnreadable, which the worker turns
    into a REPORT_UNREADABLE flag and carries on with receipts only.

The KM tab is read the same way (trips: date, from, to, purpose, vehicle,
km, rate, amount; audit: km × rate = amount). Whether the rate is one of
the client's rates is a CHECK (checks.py), not an audit: a wrong rate is
the employee's mistake, not the reader's.

Money is Decimal end to end and stored as text ("45.00").
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from pydantic import BaseModel, Field, model_validator

from ..model_layer import USAGE_LIMITS, create_agent
from .evidence import ai_call
from ..pipeline.listing_agent import content_cells, grid_text

log = logging.getLogger("claims.report_reader")

MAX_ROUNDS = 3
CENT = Decimal("0.01")
# The last row an .xlsx sheet can hold, and the widest span of lines this
# reader will walk. A reading that names a row beyond the first is
# refused by the model; one that names a span beyond the second is
# refused before a single cell is touched — reading a cell that does not
# exist CREATES it, so an unbounded span rewrites the workbook in memory
# and walks a million empty rows to say nothing.
MAX_ROW = 1048576
MAX_SHEET_ROWS = 20000
_COL = r"^[A-Za-z]{1,3}$"
_CELL = r"^[A-Za-z]{1,3}[0-9]{1,7}$"


class ReportUnreadable(Exception):
    """The reading never verified; the flag is REPORT_UNREADABLE."""


def _distinct(model) -> None:
    seen: dict[str, str] = {}
    for role, letter in model.model_dump().items():
        if letter is None:
            continue
        key = letter.upper()
        if key in seen:
            raise ValueError(f"column {letter} is given two roles ({seen[key]} and {role})")
        seen[key] = role


class ReportColumns(BaseModel):
    """Which column letter plays which role on the report tab."""

    @model_validator(mode="after")
    def _distinct_columns(self):
        _distinct(self)
        return self

    date: str | None = Field(default=None, pattern=_COL)
    item: str | None = Field(default=None, pattern=_COL, description="expense item / category, often with a GL code")
    reason: str | None = Field(default=None, pattern=_COL, description="detailed business reason")
    receipt_included: str | None = Field(default=None, pattern=_COL, description="Y/N receipt included")
    amount: str | None = Field(default=None, pattern=_COL, description="amount per receipt, in the row's currency")
    currency: str | None = Field(default=None, pattern=_COL)
    rate: str | None = Field(default=None, pattern=_COL, description="exchange rate")
    total: str | None = Field(default=None, pattern=_COL, description="total in the home currency (MYR)")


class ReportReading(BaseModel):
    """The AI's structural answer for a report tab."""
    columns: ReportColumns
    header_row: int = Field(ge=1, le=MAX_ROW, description="the row holding the column headings")
    first_row: int = Field(ge=1, le=MAX_ROW, description="first expense line")
    last_row: int = Field(ge=1, le=MAX_ROW, description="last expense line (blank rows inside are fine)")
    total_cell: str | None = Field(default=None, pattern=_CELL, description="the cell holding the report total")
    name_cell: str | None = Field(default=None, pattern=_CELL, description="the cell holding the employee's name")
    period_cell: str | None = Field(default=None, pattern=_CELL)
    purpose_cell: str | None = Field(default=None, pattern=_CELL,
                                     description="the cell holding the report's stated business reason / purpose")
    skip_rows: list[int] = Field(default_factory=list, max_length=40,
                                 description="rows INSIDE first_row..last_row that are not expense lines "
                                             "(a subtotal, a section heading, a blank-with-a-note); a "
                                             "dated row with an amount is never a skip")
    why: str = Field(max_length=300)
    observations: list[str] = Field(default_factory=list, max_length=10)


class KMColumns(BaseModel):
    @model_validator(mode="after")
    def _distinct_columns(self):
        _distinct(self)
        return self

    date: str | None = Field(default=None, pattern=_COL)
    from_: str | None = Field(default=None, pattern=_COL, alias="from", description="from / origin")
    to: str | None = Field(default=None, pattern=_COL)
    purpose: str | None = Field(default=None, pattern=_COL)
    vehicle: str | None = Field(default=None, pattern=_COL)
    km: str | None = Field(default=None, pattern=_COL)
    rate: str | None = Field(default=None, pattern=_COL, description="rate per km")
    amount: str | None = Field(default=None, pattern=_COL)

    model_config = {"populate_by_name": True}


class KMReading(BaseModel):
    """The AI's structural answer for a mileage (KM) tab."""
    has_trips: bool = Field(description="False when the tab holds no trip rows at all")
    columns: KMColumns | None = None
    header_row: int | None = Field(default=None, ge=1, le=MAX_ROW)
    first_row: int | None = Field(default=None, ge=1, le=MAX_ROW)
    last_row: int | None = Field(default=None, ge=1, le=MAX_ROW)
    total_cell: str | None = Field(default=None, pattern=_CELL)
    skip_rows: list[int] = Field(default_factory=list, max_length=40,
                                 description="rows inside first_row..last_row that are not trips (a subtotal, a heading)")
    why: str = Field(max_length=300)


_REPORT_INSTRUCTIONS = (
    "You are reading one tab of an employee expense report workbook. The "
    "tab is given as a grid of cells labelled like 'B3: some value'. Do NOT "
    "copy any numbers into your answer — answer only with structure: which "
    "column letter plays which role, which rows are the expense lines, and "
    "which single cells hold the total, the employee's name, the period, "
    "and the report's stated business reason / purpose.\n"
    "- The expense lines are the dated rows under the column headings; "
    "first_row..last_row spans them all (blank rows inside are fine). A "
    "subtotal or section-heading row INSIDE the span goes in skip_rows — "
    "never a dated row with an amount.\n"
    "- amount is the per-receipt amount in the row's currency; total is the "
    "home-currency (MYR) figure; rate is the exchange rate. If the sheet "
    "has only one money column, name it as amount and leave total null.\n"
    "- total_cell is the cell holding the sum of the lines (the report "
    "total), NOT the heading beside it.\n"
    "- name_cell / period_cell / purpose_cell are the VALUE cells (e.g. the "
    "name itself), not the label cells ('Name:').\n"
    "- In observations, list briefly anything a reviewer should know."
)

_KM_INSTRUCTIONS = (
    "You are reading the mileage tab of an employee expense report. The tab "
    "is a grid of cells labelled like 'F5: 12.4'. Answer with structure only: "
    "which column is the date, from, to, purpose, vehicle, km, rate per km, "
    "amount; which rows are the trips (first_row..last_row); which cell holds "
    "the total if any; subtotal or heading rows inside the span go in "
    "skip_rows. If the tab has headings but no trip rows, answer "
    "has_trips=false."
)


# ---- helpers ------------------------------------------------------------------

_CURRENCY_PREFIX = re.compile(r"^\s*(RM|MYR|SGD|USD|S\$|\$)\s*", re.IGNORECASE)
_GL_IN_ITEM = re.compile(r"\((\d{4,8})\)\s*$")
_ER_PERIOD = re.compile(r"ER\((\d{2}[A-Za-z]{3}\d{2})-(\d{2}[A-Za-z]{3}\d{2})\)", re.IGNORECASE)


def money(value) -> Decimal | None:
    """A cell's money value as a Decimal, or None when it is not a number.
    Text like 'RM 1,200.00' and accountant negatives '(500.00)' are read."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            d = Decimal(repr(float(value))) if isinstance(value, float) else Decimal(value)
        except InvalidOperation:
            return None
    else:
        s = _CURRENCY_PREFIX.sub("", str(value).strip()).replace(",", "")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            d = Decimal(s)
        except InvalidOperation:
            return None
    return d if d.is_finite() else None


def cents(d: Decimal) -> Decimal:
    return d.quantize(CENT)


def cell_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%b-%Y", "%d %b %Y",
                "%d %B %Y", "%d.%m.%Y", "%d-%b-%y", "%d %b %y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def er_period(er_code: str) -> tuple[date, date] | None:
    """'ER(01JUL26-21JUL26)' → (2026-07-01, 2026-07-21); None if the code
    does not follow that convention (then the date check is skipped)."""
    m = _ER_PERIOD.search(er_code or "")
    if not m:
        return None
    try:
        a = datetime.strptime(m.group(1).upper(), "%d%b%y").date()
        b = datetime.strptime(m.group(2).upper(), "%d%b%y").date()
    except ValueError:
        return None
    return (a, b) if a <= b else (b, a)


def gl_of(item: str) -> str:
    m = _GL_IN_ITEM.search(item or "")
    return m.group(1) if m else ""


def item_name(item: str) -> str:
    """'Taxi (713070)' → 'Taxi'."""
    return _GL_IN_ITEM.sub("", item or "").strip()


def row_of(ref: str | None) -> int:
    """The row number of a cell reference ('H11' → 11), 0 when there is none."""
    return int("".join(ch for ch in (ref or "") if ch.isdigit()) or 0)


def _at(ws, ref: str | None):
    """One cell's value, or None when the reference sits below the sheet's
    last row — asking openpyxl for a cell that is not there CREATES it."""
    if not ref or row_of(ref) > (ws.max_row or 0):
        return None
    return ws[ref].value


def _text(ws, ref: str | None) -> str:
    v = _at(ws, ref)
    return str(v).strip() if v is not None else ""


def _val(ws, col: str | None, row: int):
    if not col or row > (ws.max_row or 0):
        return None
    return ws[f"{col}{row}"].value


def _too_wide(first_row: int, last_row: int) -> str:
    """The reason a span is refused, or "" — checked before any cell is read."""
    if last_row - first_row + 1 > MAX_SHEET_ROWS:
        return (f"first_row..last_row spans rows {first_row}..{last_row} — wider than the "
                f"{MAX_SHEET_ROWS} rows one tab may hold; name the rows the lines are actually on")
    return ""


def _span(ws, first_row: int, last_row: int, total_cell: str | None = None) -> range:
    """The rows to walk: refused when the reading's span is wider than the
    reader supports, and never past the sheet's own last row or the total
    row (a total is not a line)."""
    why = _too_wide(first_row, last_row)
    if why:
        raise ValueError(why)
    last = min(last_row, ws.max_row or last_row)
    total_row = row_of(total_cell)
    if total_row and total_row <= last:
        last = total_row - 1
    return range(first_row, last + 1)


def _norm_currency(text: str) -> str:
    t = (text or "").strip().upper()
    if t in ("", "RM", "MYR"):
        return "MYR"
    if t in ("S$", "SGD"):
        return "SGD"
    if t in ("$", "USD", "US$"):
        return "USD"
    return t[:3] if len(t) >= 3 else t


# ---- extraction + audit ---------------------------------------------------------

STRUCTURE, SOFT, TOTAL = "structure", "soft", "total"


def total_check(ws, reading: ReportReading, rows: list[dict]) -> dict | None:
    """The lines' sum against the total cell, both as cent strings, or None
    when the total cell holds no number: {"lines", "cell", "column"}."""
    if not reading.total_cell:
        return None
    total = money(_at(ws, reading.total_cell))
    if total is None:
        return None
    col = "total" if reading.columns.total else "amount"
    summed = sum((Decimal(row[col]) for row in rows if row.get(col) is not None), Decimal("0"))
    return {"lines": str(cents(summed)), "cell": str(cents(total)), "column": col}


def uncomputed_formulas(path, tab: str, max_rows: int = 400) -> int:
    """How many cells of the tab hold a formula with no saved value — a
    workbook written by a script and never opened in Excel. Reading such a
    tab with data_only=True yields None everywhere a number should be."""
    from openpyxl import load_workbook

    try:
        with_formulas = load_workbook(path, read_only=True)
        with_values = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return 0
    try:
        if tab not in with_formulas.sheetnames:
            return 0
        n = 0
        rows_f = with_formulas[tab].iter_rows(min_row=1, max_row=max_rows, values_only=True)
        rows_v = with_values[tab].iter_rows(min_row=1, max_row=max_rows, values_only=True)
        for row_f, row_v in zip(rows_f, rows_v):
            for f, v in zip(row_f, row_v):
                if isinstance(f, str) and f.startswith("=") and v is None:
                    n += 1
        return n
    finally:
        with_formulas.close()
        with_values.close()


def extract_rows(ws, reading: ReportReading) -> list[dict]:
    """The expense lines as dicts, from the cells the AI pointed at.
    A row is a line when it has a date or an amount."""
    cols = reading.columns
    rows = []
    skip = set(reading.skip_rows or [])
    for r in _span(ws, reading.first_row, reading.last_row, reading.total_cell):
        if r in skip:
            continue
        d = cell_date(_val(ws, cols.date, r)) if _val(ws, cols.date, r) is not None else None
        amount = money(_val(ws, cols.amount, r))
        total = money(_val(ws, cols.total, r)) if cols.total else None
        if d is None and amount is None and total is None:
            continue
        rate = money(_val(ws, cols.rate, r)) if cols.rate else None
        item = str(_val(ws, cols.item, r) or "").strip()
        rows.append({
            "row": r, "date": d.isoformat() if d else "", "date_raw": str(_val(ws, cols.date, r) or ""),
            "item": item, "item_name": item_name(item), "gl": gl_of(item),
            "reason": str(_val(ws, cols.reason, r) or "").strip(),
            "receipt_included": str(_val(ws, cols.receipt_included, r) or "").strip().upper()[:1],
            "amount": None if amount is None else str(cents(amount)),
            "currency": _norm_currency(str(_val(ws, cols.currency, r) or "")),
            "rate": None if rate is None else str(rate),
            "total": None if total is None else str(cents(total)),
        })
    return rows


def audit_report(ws, reading: ReportReading, rows: list[dict], employee_name: str,
                 er_code: str) -> list[tuple[str, str]]:
    """Problems with the reading, as (kind, text). STRUCTURE problems mean
    the reading cannot be trusted; SOFT ones (dates outside the period)
    are fed back but do not, on their own, make the report unreadable."""
    problems: list[tuple[str, str]] = []
    wide = _too_wide(reading.first_row, reading.last_row)
    if wide:
        return [(STRUCTURE, wide)]
    cols = reading.columns
    for role in ("date", "item", "amount"):
        if getattr(cols, role) is None:
            problems.append((STRUCTURE, f"no {role} column named — an expense report needs one"))
    if reading.first_row > reading.last_row:
        problems.append((STRUCTURE, "first_row is after last_row"))
    if reading.header_row >= reading.first_row:
        problems.append((STRUCTURE, "header_row must be above first_row"))
    if not rows:
        problems.append((STRUCTURE, "no expense lines found in first_row..last_row"))
    for r in reading.skip_rows or []:
        if not (reading.first_row <= r <= reading.last_row):
            problems.append((STRUCTURE, f"skip_rows names row {r}, which is outside first_row..last_row"))
            continue
        raw_date = _val(ws, cols.date, r) if cols.date else None
        if raw_date is not None and cell_date(raw_date) and money(_val(ws, cols.amount, r)) is not None:
            problems.append((STRUCTURE, f"row {r} is in skip_rows but has a date and an amount — it looks "
                                        "like an expense line, not a subtotal"))
    if problems:
        return problems
    for row in rows:
        r = row["row"]
        if not row["date"]:
            problems.append((STRUCTURE, f"row {r} has an amount but no readable date "
                                        f"({row['date_raw']!r}) — is the date column right?"))
        if row["amount"] is None:
            problems.append((STRUCTURE, f"row {r} has a date but no amount in column {cols.amount}"))
            continue
        if row["rate"] is not None and row["total"] is not None:
            expect = cents(Decimal(row["amount"]) * Decimal(row["rate"]))
            if expect != Decimal(row["total"]):
                # A wrong reading looks like this on EVERY row; a single
                # row that is off is the employee's arithmetic (a check
                # flag, CURRENCY_MISMATCH). Only a widespread miss is a
                # reading problem — see below.
                row["arith_off"] = str(expect)
    off = [row for row in rows if row.get("arith_off")]
    if off and len(off) > max(1, len(rows) // 2):
        problems.append((STRUCTURE, f"amount × rate ≠ total on {len(off)} of {len(rows)} lines "
                                    f"(e.g. row {off[0]['row']}) — the amount / rate / total columns "
                                    "are probably wrong"))
    # A dated row with an amount just outside the span — between the
    # headings and first_row, or between last_row and the total — is a
    # line the reading missed. That is what tells a short span from a
    # mistyped total: with it, the total mismatch below stays structural.
    total_row = min(row_of(reading.total_cell), ws.max_row or 0)
    outside = list(range(reading.header_row + 1, reading.first_row))
    if total_row > reading.last_row:
        outside += list(range(reading.last_row + 1, total_row))
    for r in outside:
        raw_date = _val(ws, cols.date, r)
        if raw_date is not None and cell_date(raw_date) and money(_val(ws, cols.amount, r)) is not None:
            problems.append((STRUCTURE, f"row {r} has a date and an amount but is outside "
                                        f"first_row..last_row — extend the span to include it"))
            break
    if reading.total_cell:
        check = total_check(ws, reading, rows)
        if check is None:
            problems.append((STRUCTURE, f"total_cell {reading.total_cell} holds no number"))
        elif check["lines"] != check["cell"]:
            # A wrong span usually shows other structural signs (undated
            # rows, no amounts). When the reading is otherwise sound, the
            # mismatch is more likely the employee's total than the
            # reader's span — so it is TOTAL, not STRUCTURE: fed back once
            # (a missed line is the other cause), accepted if the reading
            # comes back the same, and then flagged for a person.
            sound = not any(k == STRUCTURE for k, _ in problems)
            problems.append((TOTAL if sound else STRUCTURE,
                             f"the {len(rows)} lines' {check['column']}s sum to {check['lines']} but "
                             f"{reading.total_cell} holds {check['cell']} — either a line is missing "
                             "from first_row..last_row (extend the span), or the total cell is mistyped. "
                             "If your reading is right, answer the same reading again and the mismatch "
                             "will be flagged for a person."))
    else:
        problems.append((STRUCTURE, "total_cell is missing — name the cell holding the report total"))
    if reading.name_cell:
        name = _text(ws, reading.name_cell)
        if not name.strip():
            # an empty cell is inside every name: without this it passes
            # the comparison below and the reading is never checked
            problems.append((STRUCTURE, f"name_cell {reading.name_cell} is empty — point at the cell "
                                        "holding the employee's name"))
        elif employee_name and employee_name.strip().lower() not in name.lower() \
                and name.lower() not in employee_name.strip().lower():
            problems.append((STRUCTURE, f"name_cell {reading.name_cell} holds {name!r}, not the "
                                        f"employee {employee_name!r} — point at the name cell"))
    else:
        problems.append((SOFT, "name_cell is missing — point at the cell holding the employee's name"))
    period = er_period(er_code)
    if period:
        outside = [row["row"] for row in rows
                   if row["date"] and not (period[0] <= date.fromisoformat(row["date"]) <= period[1])]
        if outside:
            problems.append((SOFT, f"row(s) {outside[:6]} are dated outside the ER period "
                                   f"{period[0]}..{period[1]} — check the date column"))
    return problems


def run_context(text: str) -> str:
    """The reviewer's instructions for this run, as the readers see them
    (H1): appended to the grid or page, plainly marked as steering — they
    say where to look and never override the sheet's own arithmetic or a
    page's own figures. Empty when there are none."""
    text = (text or "").strip()
    if not text:
        return ""
    return ("\n\n# Instructions for this run (from the reviewer; they steer where to look and "
            "what to expect, and never override what the sheet or page itself says)\n" + text[:4000])


async def read_report(ws, employee_name: str, er_code: str, usage=None,
                      context: str = "") -> tuple[list[dict], dict, list[tuple[str, str]]]:
    """The reason/act loop for the report tab.

    Returns (rows, header, notes): the extracted lines, the header facts
    (name, period, purpose, total), and (level, text) notes for the diary.
    Raises ReportUnreadable when the STRUCTURE never verifies. `context`
    is the run's instructions already wrapped by run_context() (the worker
    wraps once for every reader), shown but never trusted over the audit.
    """
    grid = grid_text(ws) + (context or "")
    agent = create_agent("judge", ReportReading, _REPORT_INSTRUCTIONS, temperature=0)
    feedback = ""
    notes: list[tuple[str, str]] = []
    problems: list[tuple[str, str]] = []
    previous: dict | None = None
    for round_no in range(1, MAX_ROUNDS + 1):
        if usage is not None:
            usage.reserve()
        result = await ai_call(agent.run(grid + feedback, usage_limits=USAGE_LIMITS), "the report reader")
        if usage is not None:
            usage.add(result)
        reading = result.output
        try:
            rows = extract_rows(ws, reading)
            problems = audit_report(ws, reading, rows, employee_name, er_code)
        except Exception as exc:
            rows, problems = [], [(STRUCTURE, f"applying your reading failed: {exc}")]
        structural = [t for k, t in problems if k == STRUCTURE]
        soft = [t for k, t in problems if k in (SOFT, TOTAL)]
        # A reading with only soft problems is accepted on the last round —
        # or as soon as the AI, shown the problem, answers the SAME reading
        # again: it has looked and stands by it, and pushing further only
        # tempts it to move the span to make a typo add up.
        structure = reading.model_dump(exclude={"why", "observations"})
        same_again = previous is not None and structure == previous
        previous = structure
        if not structural and (not soft or round_no == MAX_ROUNDS or same_again):
            check = total_check(ws, reading, rows)
            header = {
                "name": _text(ws, reading.name_cell), "period": _text(ws, reading.period_cell),
                "purpose": _text(ws, reading.purpose_cell),
                "total": check["cell"] if check else None,
                "total_cell": reading.total_cell,
                # set only when the lines do not add up to the total cell
                "total_check": check if check and check["lines"] != check["cell"] else None,
                "columns": reading.columns.model_dump(), "header_row": reading.header_row,
                "observations": list(reading.observations),
            }
            notes.append(("INFO", f"Report tab {ws.title!r}: {len(rows)} line(s), total "
                                  f"{header['total']}, confirmed on round {round_no}"
                                  + (" (same reading twice)" if same_again else "") + "."))
            for text in soft:
                notes.append(("WARNING", f"Report tab {ws.title!r}: {text}"))
            return rows, header, notes
        log.info("report %r round %d: %s", ws.title, round_no, "; ".join(t for _, t in problems)[:400])
        feedback = ("\n\nYour previous reading failed verification against the sheet's own "
                    "arithmetic:\n- " + "\n- ".join(t for _, t in problems)
                    + "\nLook at the grid again and correct the columns, rows or cells.")
    raise ReportUnreadable("Could not confirm a reading of the report tab after "
                           f"{MAX_ROUNDS} attempts. Last problems: "
                           + "; ".join(t for _, t in problems))


def extract_trips(ws, reading: KMReading) -> list[dict]:
    cols = reading.columns
    if not reading.has_trips or cols is None or reading.first_row is None or reading.last_row is None:
        return []
    trips = []
    skip = set(reading.skip_rows or [])
    for r in _span(ws, reading.first_row, reading.last_row, reading.total_cell):
        if r in skip:
            continue
        km = money(_val(ws, cols.km, r))
        amount = money(_val(ws, cols.amount, r))
        raw_date = _val(ws, cols.date, r)
        d = cell_date(raw_date) if raw_date is not None else None
        if km is None and amount is None and d is None:
            continue
        rate = money(_val(ws, cols.rate, r)) if cols.rate else None
        trips.append({
            "row": r, "date": d.isoformat() if d else "",
            "from": str(_val(ws, cols.from_, r) or "").strip(),
            "to": str(_val(ws, cols.to, r) or "").strip(),
            "purpose": str(_val(ws, cols.purpose, r) or "").strip(),
            "vehicle": str(_val(ws, cols.vehicle, r) or "").strip(),
            "km": None if km is None else str(km),
            "rate": None if rate is None else str(rate),
            "amount": None if amount is None else str(cents(amount)),
        })
    return trips


def audit_km(ws, reading: KMReading, trips: list[dict]) -> list[str]:
    if not reading.has_trips:
        return []
    cols = reading.columns
    problems = []
    if reading.first_row is not None and reading.last_row is not None:
        wide = _too_wide(reading.first_row, reading.last_row)
        if wide:
            return [wide]
    if cols is None or cols.km is None or cols.amount is None or cols.date is None:
        return ["name the date, km and amount columns (or answer has_trips=false)"]
    if not trips:
        return ["no trip rows found in first_row..last_row"]
    for r in reading.skip_rows or []:
        raw_date = _val(ws, cols.date, r)
        if raw_date is not None and cell_date(raw_date) and money(_val(ws, cols.km, r)) is not None:
            problems.append(f"row {r} is in skip_rows but has a date and km — it looks like a trip")
    for t in trips:
        if not t["date"]:
            problems.append(f"row {t['row']}: no readable date")
        if t["km"] is None or t["amount"] is None:
            problems.append(f"row {t['row']}: km or amount missing")
            continue
        if t["rate"] is not None and cents(Decimal(t["km"]) * Decimal(t["rate"])) != Decimal(t["amount"]):
            t["arith_off"] = str(cents(Decimal(t["km"]) * Decimal(t["rate"])))
    off = [t for t in trips if t.get("arith_off")]
    if off and len(off) > max(1, len(trips) // 2):
        problems.append(f"km × rate ≠ amount on {len(off)} of {len(trips)} rows — the columns are probably wrong")
    if reading.total_cell:
        total = money(_at(ws, reading.total_cell))
        summed = sum((Decimal(t["amount"]) for t in trips if t["amount"]), Decimal("0"))
        if total is None or cents(summed) != cents(total):
            problems.append(f"the trips' amounts sum to {cents(summed)} but {reading.total_cell} holds {total}")
    return problems


async def read_km(ws, usage=None, context: str = "") -> tuple[list[dict], list[tuple[str, str]]]:
    """The reason/act loop for the KM tab. Returns (trips, notes). A tab
    that never verifies raises ReportUnreadable (the worker records it and
    treats the employee as having no readable trips)."""
    grid = grid_text(ws) + (context or "")
    agent = create_agent("judge", KMReading, _KM_INSTRUCTIONS, temperature=0)
    feedback = ""
    problems: list[str] = []
    for round_no in range(1, MAX_ROUNDS + 1):
        if usage is not None:
            usage.reserve()
        result = await ai_call(agent.run(grid + feedback, usage_limits=USAGE_LIMITS), "the mileage reader")
        if usage is not None:
            usage.add(result)
        reading = result.output
        try:
            trips = extract_trips(ws, reading)
            problems = audit_km(ws, reading, trips)
        except Exception as exc:
            trips, problems = [], [f"applying your reading failed: {exc}"]
        if not problems:
            note = (f"KM tab {ws.title!r}: {len(trips)} trip(s), confirmed on round {round_no}."
                    if reading.has_trips else f"KM tab {ws.title!r}: no trips.")
            return trips, [("INFO", note)]
        feedback = ("\n\nYour previous reading failed verification:\n- " + "\n- ".join(problems)
                    + "\nLook at the grid again and correct it.")
    raise ReportUnreadable(f"Could not confirm a reading of the mileage tab {ws.title!r} after "
                           f"{MAX_ROUNDS} attempts. Last problems: " + "; ".join(problems))


def read_categories(wb) -> list[dict]:
    """The client's category list from an 'Expense Types'-style tab: rows of
    (item, GL code). Code only — the tab is two columns."""
    for ws in wb.worksheets:
        title = ws.title.lower()
        if "type" not in title and "categor" not in title:
            continue
        out = []
        for row in ws.iter_rows(min_row=1, max_row=400, max_col=6, values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if len(cells) < 1:
                continue
            item = cells[0]
            gl = next((c for c in cells[1:] if re.fullmatch(r"\d{4,8}", c)), gl_of(item))
            if item.lower() in ("expense item", "item", "category", "expense type"):
                continue
            out.append({"item": item_name(item), "gl": gl})
        if len(out) >= 3:
            return out
    return []


def grid_of(ws) -> str:
    """The tab as the AI sees it (shared with the listing reader)."""
    return grid_text(ws)


__all__ = ["read_report", "read_km", "read_categories", "ReportUnreadable", "content_cells",
           "money", "cents", "er_period", "gl_of", "item_name", "cell_date"]
