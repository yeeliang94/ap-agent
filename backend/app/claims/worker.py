"""One worker per employee, five at a time, failure isolated, retryable.

A worker = steps 7–9 for ONE employee, seeing only that employee's files:
read the report (+ KM tab) → inventory the evidence pages → match and
check → decide the category → write rows, evidence and flags. It runs
inside its own database session and its own try/except: one employee
failing (model error, request cap, unreadable file) marks that employee
failed with the reason and the others carry on. Per-employee timing and
AI cost go to the run diary and the employee summary.

The run turns ready when every employee is verified, failed or skipped.
"""
from __future__ import annotations

import asyncio
import logging
import time
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from .. import config, telemetry
from ..db import SessionLocal
from ..model_layer import USAGE_LIMITS, create_agent
from . import cases as cases_mod
from . import category as category_mod
from . import checks as checks_mod
from . import evidence as evidence_mod
from . import profile as profile_mod
from . import report_reader
from .models import ClaimEmployee, ClaimEvidence, ClaimFlag, ClaimRow, ClaimsRun

log = logging.getLogger("claims.worker")

POOL_SIZE = 5
# A worker's whole budget of AI requests. Each page read is its own tiny
# conversation (well inside the enterprise per-agent cap, which
# USAGE_LIMITS enforces on every call); this cap bounds the WORKER, so a
# runaway employee folder cannot multiply spend unnoticed.
WORKER_REQUEST_CAP = config.MAX_AGENT_REQUESTS * 4


# Raised by Usage.reserve() before any request past the cap is sent.
WorkerBudgetExceeded = evidence_mod.BudgetExceeded


def _files_dir(run_id: str) -> Path:
    from .runner import files_dir

    return files_dir(run_id)


async def verify_run(db, run: ClaimsRun) -> None:
    """Run the pool over every pending employee, then close the run."""
    from .runner import _set

    started = time.monotonic()
    employees = db.query(ClaimEmployee).filter(ClaimEmployee.run_id == run.id,
                                               ClaimEmployee.status == "pending").all()
    ids = [e.id for e in employees]
    total = db.query(ClaimEmployee).filter(ClaimEmployee.run_id == run.id).count()
    _set(db, run, progress={"done": total - len(ids), "total": total})
    telemetry.record(db, run.id, "verify", telemetry.INFO, "STAGE_STARTED",
                     f"Verifying {len(ids)} employee(s), {POOL_SIZE} at a time.")
    # The listing header map (Step 12) is read once, before the workers,
    # so the category judge can see past examples and the output has its
    # column order ready.
    try:
        from . import listing as listing_mod

        await listing_mod.prepare_listing(db, run)
    except Exception as exc:
        telemetry.record_failure(db, run.id, "output", "LISTING_UNREADABLE",
                                 "Could not read the linked listing", exc)
    # The client's category list, once per run: from the first report in
    # the batch that carries an Expense Types tab (they share a template),
    # so an employee with no report is still judged against the client's
    # own list. Code only.
    if not (profile_mod.profile_of(run.snapshot).get("categories") or []):
        cats = _batch_categories(run)
        if cats:
            run.survey = {**(run.survey or {}), "categories": cats}
            db.commit()
    sem = asyncio.Semaphore(POOL_SIZE)

    async def one(eid: str) -> None:
        async with sem:
            await verify_employee(run.id, eid)
            _bump(run.id)

    await asyncio.gather(*(one(eid) for eid in ids))
    _finish_run(run.id, started)


def _shared_receipt_flags(s, run_id: str, profile: dict) -> int:
    """SHARED_RECEIPT: the same receipt (vendor, date, amount, currency)
    matched to a row under two or more employees. Each worker sees only its
    own folder, so this is the one pass over the whole batch — at run
    close. Idempotent by (code, row, evidence); a flag a person already
    decided is not raised again. Returns how many were added."""
    if not profile_mod.check_enabled(profile, "SHARED_RECEIPT"):
        return 0
    receipts = s.query(ClaimEvidence).filter(ClaimEvidence.run_id == run_id, ClaimEvidence.kind == "receipt",
                                             ClaimEvidence.matched_row_id != "").all()
    if not receipts:
        return 0
    employees = {e.id: e for e in s.query(ClaimEmployee).filter(ClaimEmployee.run_id == run_id).all()}
    existing = {(f.code, f.row_id, f.evidence_id) for f in s.query(ClaimFlag).filter(
        ClaimFlag.run_id == run_id, ClaimFlag.code == "SHARED_RECEIPT").all()}
    groups: dict[tuple, list[ClaimEvidence]] = {}
    for e in receipts:
        groups.setdefault(checks_mod._receipt_key({"values": e.values or {}}), []).append(e)
    added = 0
    for group in groups.values():
        if len({e.employee_id for e in group}) < 2:
            continue
        for e in group:
            key = ("SHARED_RECEIPT", e.matched_row_id, e.id)
            if key in existing:
                continue
            others = [x for x in group if x.employee_id != e.employee_id]
            names = sorted({(employees.get(x.employee_id).name or employees.get(x.employee_id).folder)
                            if employees.get(x.employee_id) else "?" for x in others})
            as_dict = lambda x: {"file": x.file, "page": x.page, "position": x.position, "values": x.values or {}}  # noqa: E731
            v = e.values or {}
            s.add(ClaimFlag(
                run_id=run_id, employee_id=e.employee_id, case_id=e.case_id or cases_mod.case_id_for_employee(s, e.employee_id),
                row_id=e.matched_row_id, evidence_id=e.id,
                code="SHARED_RECEIPT",
                reason=(f"The receipt from {v.get('vendor', '?')} ({v.get('date') or 'no date'}, "
                        f"{v.get('currency', 'MYR')} {v.get('amount')}) at {checks_mod._where(as_dict(e))}"
                        + f" also supports a row of {', '.join(names)} ("
                        + "; ".join(checks_mod._where(as_dict(x)) for x in others)
                        + "). One expense, two claims — or two identical receipts. Decide whose it is."),
                basis="universal rule: one receipt supports one row, across the whole batch",
                cite=checks_mod._ev_cite(as_dict(e))))
            existing.add(key)
            added += 1
    return added


def _unreadable_reason(path: Path, tab: str, report_file: str, exc: Exception) -> str:
    """The sentence for a report tab that never verified. When the tab
    holds formulas with no saved value (a workbook written by a script and
    never opened in Excel), that is the cause and the fix — say so instead
    of the generic three-rounds message."""
    n = report_reader.uncomputed_formulas(path, tab)
    if n:
        return (f"The report tab {tab!r} of {report_file} has {n} formula cell(s) with no saved value — "
                "the workbook was never opened and saved in Excel, so its totals read as empty. Open it "
                "in Excel, save it, re-upload and re-verify. Continuing with receipts only.")
    return (f"The report tab {tab!r} of {report_file} could not be read reliably: {exc} "
            "Continuing with receipts only.")


def _report_total_flags(header: dict, tab: str, profile: dict | None = None) -> list[dict]:
    """REPORT_TOTAL_MISMATCH when the report's lines were read and checked
    but do not add up to its total cell — the lines are kept and paid; a
    person confirms which figure is right. Off per client like any check."""
    check = (header or {}).get("total_check")
    if not check or (profile is not None and not profile_mod.check_enabled(profile, "REPORT_TOTAL_MISMATCH")):
        return []
    cell = (header or {}).get("total_cell") or ""
    row_no = int("".join(ch for ch in cell if ch.isdigit()) or 0)
    return [checks_mod._flag(
        "REPORT_TOTAL_MISMATCH",
        f"The report's {check['column']} column adds up to {check['lines']} but its total cell "
        f"{cell} says {check['cell']} — the lines were read and checked and are what the listing "
        "will pay; the total is usually a typo, sometimes a line the reader missed. Compare the two "
        "on the report; acknowledge if the lines are right, or fix the file and re-verify.",
        "universal rule: the lines and the total cell must agree; a person settles which is right",
        {"sheet": tab, "row": row_no})]


def _batch_categories(run: ClaimsRun) -> list[dict]:
    files = _files_dir(run.id)
    for e in (run.map or {}).get("employees", []):
        rf = e.get("report_file")
        if e.get("is_employee") and rf and (files / rf).is_file():
            try:
                cats = report_reader.read_categories(load_workbook(files / rf, read_only=True, data_only=True))
            except Exception:
                continue
            if cats:
                return cats
    return []


def _bump(run_id: str) -> None:
    s = SessionLocal()
    try:
        run = s.get(ClaimsRun, run_id)
        total = s.query(ClaimEmployee).filter(ClaimEmployee.run_id == run_id).count()
        done = s.query(ClaimEmployee).filter(
            ClaimEmployee.run_id == run_id,
            ClaimEmployee.status.in_(("verified", "failed", "skipped"))).count()
        run.progress = {"done": done, "total": total}
        s.commit()
    finally:
        s.close()


def _finish_run(run_id: str, started: float) -> bool:
    """Close the run: run-level flags for controls the batch needed and
    could not find, then ready — with a diary line summing it up.

    Only when EVERY employee is verified, failed or skipped. A retry that
    ends while the pool is still working (or the pool ending while a
    retry is in flight) leaves the run verifying; whoever finishes last
    closes it. Returns whether the run was closed."""
    s = SessionLocal()
    try:
        run = s.get(ClaimsRun, run_id)
        employees = s.query(ClaimEmployee).filter(ClaimEmployee.run_id == run_id).all()
        if any(e.status in ("pending", "verifying") for e in employees):
            return False
        profile = profile_mod.profile_of(run.snapshot)
        rows = s.query(ClaimRow).filter(ClaimRow.run_id == run_id).all()
        row_dicts = [{"kind": r.kind} for r in rows]
        # A run-level flag is identified by its code AND what it is about
        # (cite.what), so a re-verification never raises the same one twice.
        existing = {(f.code, (f.cite or {}).get("what", ""))
                    for f in s.query(ClaimFlag).filter(ClaimFlag.run_id == run_id,
                                                        ClaimFlag.employee_id == "").all()}
        why = checks_mod.needs_missing_reference(row_dicts, profile)
        if why and ("MISSING_REFERENCE", "rates") not in existing:
            s.add(ClaimFlag(run_id=run_id, employee_id="", code="MISSING_REFERENCE", reason=why,
                            basis="client profile: mileage_rates (not set)", cite={"what": "rates"}))
        listing_state = (run.listing_headers or {}).get("state")
        if listing_state in (None, "missing", "unreadable") and ("MISSING_REFERENCE", "listing") not in existing:
            s.add(ClaimFlag(run_id=run_id, employee_id="", code="MISSING_REFERENCE",
                            reason=("No listing workbook could be read for this run"
                                    + (f" ({(run.listing_headers or {}).get('why')})" if (run.listing_headers or {}).get("why") else "")
                                    + ", so the output columns will follow the fallback set (Name, ER code, "
                                      "Category, GL, Amount MYR) instead of the client's own listing. Link "
                                      "the listing and start a new run, or acknowledge to proceed with the "
                                      "fallback."),
                            basis="run input: this month's listing link", cite={"what": "listing"}))
        _shared_receipt_flags(s, run_id, profile)
        s.flush()  # so the count below sees the flags just added
        n_flags = s.query(ClaimFlag).filter(ClaimFlag.run_id == run_id, ClaimFlag.status == "open").count()
        failed = [e for e in employees if e.status == "failed"]
        cost = sum(int((e.summary or {}).get("requests", 0)) for e in employees)
        tokens = sum(int((e.summary or {}).get("tokens", 0)) for e in employees)
        run.status = "ready"
        run.progress = {"done": len(employees), "total": len(employees)}
        s.commit()
        telemetry.record(s, run_id, "verify", telemetry.INFO, "STAGE_DONE",
                         f"Verification finished in {time.monotonic() - started:.0f}s: "
                         f"{len(employees) - len(failed)} of {len(employees)} employee(s) verified"
                         + (f", {len(failed)} failed" if failed else "")
                         + f"; {n_flags} open flag(s); AI cost {cost} request(s), {tokens} tokens.")
        telemetry.record(s, run_id, "run", telemetry.INFO, "RUN_READY",
                         "Run is ready for review.")
        return True
    finally:
        s.close()


async def verify_employee(run_id: str, employee_id: str) -> None:
    """The whole worker for one employee, in its own session, never raising."""
    s = SessionLocal()
    started = time.monotonic()
    usage = evidence_mod.Usage(cap=WORKER_REQUEST_CAP)
    try:
        run = s.get(ClaimsRun, run_id)
        emp = s.get(ClaimEmployee, employee_id)
        if emp is None or emp.status == "skipped":
            return
        emp.status, emp.error = "verifying", ""
        cases_mod.sync_case_from_employee(s, emp)
        s.commit()
        # A retry starts clean: the previous attempt's rows, evidence and
        # open flags go; decided flags are kept for the record.
        s.query(ClaimRow).filter(ClaimRow.employee_id == employee_id).delete()
        s.query(ClaimEvidence).filter(ClaimEvidence.employee_id == employee_id).delete()
        s.query(ClaimFlag).filter(ClaimFlag.employee_id == employee_id,
                                  ClaimFlag.status.in_(("open", "info"))).delete()
        s.commit()
        try:
            await _work(s, run, emp, usage)
        except WorkerBudgetExceeded as exc:
            emp = _discard_partial(s, employee_id)
            _fail_employee(s, run_id, emp, f"AI request cap reached ({exc})", started, usage)
            return
        except Exception as exc:
            # Whatever _work had staged (rows, evidence, flags) goes with
            # the failure; only the failed status is recorded.
            emp = _discard_partial(s, employee_id)
            reason = telemetry.record_failure(s, run_id, "verify", "EMPLOYEE_FAILED",
                                              f"Could not verify {emp.name or emp.folder}", exc)
            _fail_employee(s, run_id, emp, reason, started, usage)
            return
        emp.status = "verified"
        emp.summary = {**(emp.summary or {}), "seconds": round(time.monotonic() - started, 1),
                       "requests": usage.requests, "tokens": usage.tokens}
        cases_mod.sync_case_from_employee(s, emp)
        s.commit()
        n_flags = s.query(ClaimFlag).filter(ClaimFlag.employee_id == employee_id,
                                            ClaimFlag.status == "open").count()
        telemetry.record(s, run_id, "verify", telemetry.INFO, "EMPLOYEE_DONE",
                         f"{emp.name or emp.folder}: verified in {time.monotonic() - started:.0f}s, "
                         f"{(emp.summary or {}).get('rows', 0)} row(s), {n_flags} open flag(s), "
                         f"{usage.requests} AI request(s), {usage.tokens} tokens.")
    finally:
        s.close()


def _discard_partial(s, employee_id: str) -> ClaimEmployee:
    """Roll back everything the worker staged but had not committed, and
    hand back a fresh employee record to write the failure on."""
    s.rollback()
    return s.get(ClaimEmployee, employee_id)


def _fail_employee(s, run_id: str, emp: ClaimEmployee, reason: str, started: float, usage) -> None:
    emp.status, emp.error = "failed", reason
    emp.summary = {**(emp.summary or {}), "seconds": round(time.monotonic() - started, 1),
                   "requests": usage.requests, "tokens": usage.tokens}
    cases_mod.sync_case_from_employee(s, emp)
    s.commit()


def run_context_for(run: ClaimsRun) -> str:
    """The run's instructions + the client's playbook, as the readers and
    the category judge see them (H1). Steering only: the audits and the
    checks never read this."""
    parts = [(run.instructions or "").strip(), ((run.snapshot or {}).get("playbook") or "").strip()]
    return report_reader.run_context("\n\n".join(p for p in parts if p))


async def _work(s, run: ClaimsRun, emp: ClaimEmployee, usage: evidence_mod.Usage) -> None:
    profile = profile_mod.profile_of(run.snapshot)
    files = _files_dir(run.id)
    roles = emp.roles or {}
    notes: list[tuple[str, str]] = []
    case_id = cases_mod.sync_case_from_employee(s, emp).id
    context = run_context_for(run)
    if context:
        notes.append(("INFO", f"the run's instructions/playbook ({len(context)} chars) were shown to the "
                              "report reader, the page reader and the category judge (steering only; the "
                              "audits and checks are unchanged)"))
    rows: list[dict] = []
    header: dict = {}
    categories: list[dict] = []
    flags: list[dict] = []

    def budget() -> None:
        # Every page read reserves before it sends; this is the same check
        # between the stages, so a report reader that used the budget up
        # stops the worker before the evidence is opened.
        usage.reserve()

    # ---- 7. the report (+ KM tab) --------------------------------------------------
    report_ok = False
    if not roles.get("no_report") and roles.get("report_file"):
        path = files / roles["report_file"]
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as exc:
            flags.append(checks_mod._flag("REPORT_UNREADABLE",
                                          f"{roles['report_file']} could not be opened as a workbook "
                                          f"({type(exc).__name__}). Continuing with receipts only.",
                                          "universal rule: a report that cannot be read is said so, never guessed",
                                          {"file": roles["report_file"], "page": 0}))
            wb = None
        if wb is not None:
            categories = report_reader.read_categories(wb) or list(profile.get("categories") or [])
            tab = roles.get("report_tab")
            if tab in wb.sheetnames:
                try:
                    r_rows, header, r_notes = await report_reader.read_report(
                        wb[tab], emp.name, emp.er_code, usage)
                    notes += r_notes
                    for r in r_rows:
                        rows.append({"kind": "expense", "sheet": tab, "row": r["row"], "values": r})
                    report_ok = True
                    flags += _report_total_flags(header, tab, profile)
                except report_reader.ReportUnreadable as exc:
                    why = _unreadable_reason(path, tab, roles["report_file"], exc)
                    flags.append(checks_mod._flag("REPORT_UNREADABLE", why,
                                                  "universal rule: three attempts, then a person",
                                                  {"sheet": tab, "row": 0}))
                    notes.append(("WARNING", f"Report tab {tab!r}: unreadable — {exc}"))
            else:
                flags.append(checks_mod._flag("REPORT_UNREADABLE",
                                              f"Tab {tab!r} is not in {roles['report_file']} (tabs: "
                                              f"{wb.sheetnames}). Continuing with receipts only.",
                                              "the confirmed map", {"sheet": tab or "", "row": 0}))
            budget()
            km_tab = roles.get("mileage_tab")
            if km_tab and km_tab in wb.sheetnames and km_tab != tab:
                try:
                    trips_rows, k_notes = await report_reader.read_km(wb[km_tab], usage, context=context)
                    notes += k_notes
                    for t in trips_rows:
                        rows.append({"kind": "mileage", "sheet": km_tab, "row": t["row"], "values": t})
                except report_reader.ReportUnreadable as exc:
                    flags.append(checks_mod._flag("REPORT_UNREADABLE",
                                                  f"The mileage tab {km_tab!r} could not be read reliably: {exc}",
                                                  "universal rule: three attempts, then a person",
                                                  {"sheet": km_tab, "row": 0}))
                    notes.append(("WARNING", f"KM tab {km_tab!r}: unreadable — {exc}"))
            budget()
    if not categories:
        categories = list(profile.get("categories") or (run.survey or {}).get("categories") or [])

    # ---- 8. the evidence pages ---------------------------------------------------------
    receipts: list[dict] = []
    trips: list[dict] = []
    pages_read = 0
    files_read = 0
    page_sem = evidence_mod.page_semaphore()  # shared by every worker
    for rel in roles.get("receipt_files") or []:
        path = files / rel
        if not path.is_file():
            notes.append(("WARNING", f"{rel}: file missing from the workspace"))
            continue
        try:
            r, t, pages, n = await evidence_mod.read_bundle(path, rel, usage, page_sem, context=context)
        except Exception as exc:
            reason = telemetry.describe_failure(exc)
            notes.append(("WARNING", f"{rel}: could not be read ({reason})"))
            raise
        receipts += r
        trips += t
        pages_read += len(pages)
        files_read += 1
        notes += [("INFO", x) for x in n]
        kinds = {}
        for p in pages:
            kinds[p["kind"]] = kinds.get(p["kind"], 0) + 1
        notes.append(("INFO", f"{rel}: {len(pages)} page(s) — "
                              + ", ".join(f"{n} {k}" for k, n in sorted(kinds.items()))
                              + f"; {len(r)} receipt(s), {len(t)} map trip(s)."))
        budget()

    # ---- no report: the receipts become the row list ------------------------------------
    if not report_ok and not roles.get("no_report") and not any(f["code"] == "REPORT_UNREADABLE" for f in flags):
        # a report was named but nothing came of it — say so
        flags.append(checks_mod._flag("REPORT_UNREADABLE", "No report tab could be read.",
                                      "the confirmed map", {}))
    derived = False
    if not report_ok:
        derived = True
        for i, r in enumerate(receipts, 1):
            rows.append({"kind": "derived", "sheet": "", "row": i,
                         "values": {"date": r.get("date", ""), "item": "", "item_name": "", "gl": "",
                                    "reason": f"from receipt: {r.get('vendor', '')}", "receipt_included": "Y",
                                    "amount": r["amount"], "currency": r["currency"], "rate": "1",
                                    "total": r["amount"] if r["currency"] == "MYR" else None,
                                    "vendor": r.get("vendor", "")}})
        if roles.get("no_report"):
            flags.append(checks_mod._flag("NO_REPORT",
                                          f"{emp.name or emp.folder} has no expense report in the folder; "
                                          f"{len(receipts)} row(s) were built from the receipts found. Confirm "
                                          "the derived list is what should be paid.",
                                          "universal rule: receipts but no report → rows from receipts, "
                                          "flagged for a person", {}))

    # ---- 9. checks — in memory, so no database lock is held while the
    # tie-break may call the AI; then everything is written in one commit.
    from .models import _id as new_id

    row_dicts = []
    for r in rows:
        r["id"] = new_id()
        row_dicts.append({"id": r["id"], "kind": r["kind"], "sheet": r["sheet"], "row": r["row"],
                          "values": dict(r["values"])})
    ev_dicts = []
    for r in receipts:
        r["id"] = new_id()
        ev_dicts.append({"id": r["id"], "kind": "receipt", "file": r["file"], "page": r["page"],
                         "position": r["position"],
                         "values": {"vendor": r["vendor"], "date": r["date"], "amount": r["amount"],
                                    "currency": r["currency"],
                                    **{k: r[k] for k in ("date_alt", "amount_alt") if r.get(k)}},
                         "confidence": dict(r.get("confidence") or {})})
    for t in trips:
        t["id"] = new_id()
        ev_dicts.append({"id": t["id"], "kind": "map_trip", "file": t["file"], "page": t["page"],
                         "position": "",
                         "values": {"date": t["date"], "purpose": t["purpose"], "from": t["from"],
                                    "to": t["to"], "return_trip": t["return_trip"],
                                    "km_printed": t["km_printed"]},
                         "confidence": dict(t.get("confidence") or {})})
    if derived:
        # each derived row IS its receipt (rows were built from receipts in order)
        derived_rows = [r for r in rows if r["kind"] == "derived"]
        for row, rec in zip(derived_rows, receipts):
            row["matched_evidence_id"] = rec["id"]
            rec["matched_row_id"] = row["id"]
    result = await checks_mod.run_checks(row_dicts, ev_dicts, profile,
                                         {"name": emp.name, "er_code": emp.er_code},
                                         (pages_read, files_read), TieBreak(usage))
    for r, rd in zip(rows, row_dicts):
        verdict, eid = result["verdicts"].get(r["id"], ("unchecked", ""))
        s.add(ClaimRow(id=r["id"], run_id=run.id, employee_id=emp.id, case_id=case_id, kind=r["kind"], sheet=r["sheet"],
                       row=r["row"], values=rd["values"],
                       verdict="matched" if r["kind"] == "derived" else verdict,
                       matched_evidence_id=r.get("matched_evidence_id") or eid))
    for ed in ev_dicts:
        s.add(ClaimEvidence(id=ed["id"], run_id=run.id, employee_id=emp.id, case_id=case_id, kind=ed["kind"], file=ed["file"],
                            page=ed["page"], position=ed["position"], values=ed["values"],
                            confidence=ed["confidence"],
                            matched_row_id=result["matches"].get(ed["id"], "")
                            or next((r["id"] for r in rows if r.get("matched_evidence_id") == ed["id"]), "")))
    for f in flags:
        s.add(ClaimFlag(run_id=run.id, employee_id=emp.id, case_id=case_id, **f))
    for f in result["flags"]:
        s.add(ClaimFlag(run_id=run.id, employee_id=emp.id, case_id=case_id, **f))
    for level, text in notes + [("INFO", n) for n in result["notes"]]:
        telemetry.record(s, run.id, "verify", telemetry.WARNING if level == "WARNING" else telemetry.INFO,
                         "EMPLOYEE_NOTE", f"{emp.name or emp.folder}: {text}")

    # ---- category ---------------------------------------------------------------------
    emp.category, emp.gl, emp.category_basis = "", "", ""
    if categories and profile_mod.check_enabled(profile, "CATEGORY_UNCLEAR"):
        row_values = [r["values"] for r in rows if r["kind"] in ("expense", "derived")]
        examples = list((run.listing_headers or {}).get("past_examples") or [])
        try:
            judgment, gl = await category_mod.judge_category(
                categories, header.get("purpose", "") if header else "", row_values,
                profile.get("category_rule") or "", examples, usage, context=context)
        except Exception as exc:
            judgment, gl = None, ""
            notes.append(("WARNING", f"category judgement failed: {telemetry.describe_failure(exc)}"))
        if judgment is not None and judgment.sure:
            emp.category, emp.gl = judgment.category, gl
            emp.category_basis = (f"{judgment.category}: quoted \"{judgment.quoted_text}\" — {judgment.why}"
                                  + (f" (rule: {profile.get('category_rule')})" if profile.get("category_rule") else ""))
        else:
            s.add(ClaimFlag(run_id=run.id, employee_id=emp.id, case_id=case_id, code="CATEGORY_UNCLEAR",
                            reason=("The listing category for this employee could not be settled"
                                    + (f": {judgment.why}" if judgment else " (the judge failed)")
                                    + (f" (best guess: {judgment.category})" if judgment and judgment.category else "")
                                    + ". Choose the category on the employee's summary."),
                            basis=("client profile: category rule — " + (profile.get("category_rule") or "none confirmed yet")
                                   + f"; list from {'the report' if categories else 'nowhere'}"),
                            cite={"sheet": roles.get("report_tab") or "", "row": 0}))
            if judgment and judgment.category:
                emp.category_basis = f"unsure — best guess {judgment.category}: {judgment.why}"
    elif not categories:
        s.add(ClaimFlag(run_id=run.id, employee_id=emp.id, case_id=case_id, code="CATEGORY_UNCLEAR",
                        reason="No category list is known for this client (no Expense Types tab in the "
                               "report and none in the profile). Choose the category on the employee's "
                               "summary, or add the client's list in Settings → Claims.",
                        basis="client profile: categories (not set)", cite={}))

    # ---- summary -----------------------------------------------------------------------
    # report_total is the figure the REPORT states (its total cell) — the
    # independent side of the output reconciliation. An employee with no
    # such figure (no report, or no total cell) has none, and the output
    # says so rather than reconciling the rows against themselves. The
    # rows' own sum is kept beside it for the screen.
    rows_total = str(sum((Decimal(r["values"]["total"] or r["values"]["amount"]) for r in rows
                          if r["kind"] in ("expense", "derived") and (r["values"].get("total") or r["values"].get("amount"))),
                         Decimal("0")).quantize(Decimal("0.01")))
    emp.report_total = (header.get("total") if header else None) or ""
    n_rows = sum(1 for r in rows if r["kind"] != "mileage")
    open_flags = sum(1 for f in flags + result["flags"] if f["status"] == "open")
    emp.summary = {**(emp.summary or {}), "rows": n_rows, "rows_total": rows_total,
                   "km_rows": sum(1 for r in rows if r["kind"] == "mileage"),
                   "receipts": len(receipts), "map_trips": len(trips), "pages": pages_read,
                   "flagged": len({f["row_id"] for f in result["flags"] if f["status"] == "open" and f["row_id"]}),
                   "open_flags": open_flags, "purpose": (header or {}).get("purpose", ""),
                   "categories_from": "report" if categories and not profile.get("categories") else ("profile" if categories else "")}
    cases_mod.sync_case_from_employee(s, emp)
    s.commit()


class TieBreak:
    """The AI tie-break among candidate receipts — one call per tie."""

    def __init__(self, usage: evidence_mod.Usage) -> None:
        self.usage = usage

    async def __call__(self, row: dict, candidates: list[dict]) -> str:
        from pydantic import BaseModel, Field

        class Pick(BaseModel):
            index: int = Field(ge=0, description="which candidate (0-based), or -1 if unsure")
            why: str = Field(max_length=200)

        v = row["values"]
        text = (f"A claim row: {v.get('date')} | {v.get('item_name') or v.get('item')} | "
                f"{v.get('reason')} | {v.get('currency', 'MYR')} {v.get('amount')}\n"
                "Candidate receipts (same day, amount and currency):\n"
                + "\n".join(f"{i}. vendor {c['values'].get('vendor')} at {c.get('file')} p.{c.get('page')} "
                            f"{c.get('position')}" for i, c in enumerate(candidates))
                + "\nWhich candidate is this row's receipt, judging by the row's reason and the "
                  "vendors? Answer -1 if you cannot tell.")
        self.usage.reserve()  # the budget is checked before any model is built
        agent = create_agent("judge", Pick, "Pick the receipt that supports the row, or -1 if unsure.",
                             temperature=0)
        result = await evidence_mod.ai_call(agent.run(text, usage_limits=USAGE_LIMITS), "the receipt tie-break")
        self.usage.add(result)
        i = result.output.index
        return candidates[i]["id"] if 0 <= i < len(candidates) else ""


async def run_checks_for(s, run: ClaimsRun, emp: ClaimEmployee, profile: dict,
                         searched: tuple[int, int], usage=None) -> dict:
    """Run the checks over the employee's STORED rows and evidence, and
    write the verdicts and matches back. Used by the worker and by every
    correction (instant re-check)."""
    row_objs = s.query(ClaimRow).filter(ClaimRow.employee_id == emp.id).all()
    ev_objs = s.query(ClaimEvidence).filter(ClaimEvidence.employee_id == emp.id).all()
    rows = [{"id": r.id, "kind": r.kind, "sheet": r.sheet, "row": r.row, "values": dict(r.values)}
            for r in row_objs]
    evidence = [{"id": e.id, "kind": e.kind, "file": e.file, "page": e.page, "position": e.position,
                 "values": dict(e.values), "confidence": dict(e.confidence or {})} for e in ev_objs]
    tie = TieBreak(usage or evidence_mod.Usage(cap=WORKER_REQUEST_CAP))
    result = await checks_mod.run_checks(rows, evidence, profile,
                                         {"name": emp.name, "er_code": emp.er_code}, searched, tie)
    by_id = {r.id: r for r in row_objs}
    for rid, (verdict, eid) in result["verdicts"].items():
        if rid in by_id and by_id[rid].kind != "derived":
            by_id[rid].verdict = verdict
            by_id[rid].matched_evidence_id = eid
    for e in ev_objs:
        if e.kind == "receipt" or e.kind == "map_trip":
            if e.matched_row_id and by_id.get(e.matched_row_id) and by_id[e.matched_row_id].kind == "derived":
                continue
            e.matched_row_id = result["matches"].get(e.id, "")
    s.flush()
    return result


async def retry_employee(run_id: str, employee_id: str) -> None:
    """Re-run one worker (Retry on a failed employee, or Re-verify)."""
    s = SessionLocal()
    try:
        run = s.get(ClaimsRun, run_id)
        was_ready = run.status == "ready"
        if was_ready:
            run.status = "verifying"
            s.commit()
    finally:
        s.close()
    started = time.monotonic()
    await verify_employee(run_id, employee_id)
    _bump(run_id)
    # Closes the run only if this was the last employee still working;
    # otherwise the pool (or a later retry) closes it.
    _finish_run(run_id, started)
