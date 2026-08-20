"""The month's listing ("Summary of Invoices"): its columns, and the batch rows.

Columns come from the client's own listing (PRD decision 2): the workbook
linked at run start is opened, the AI maps its header row (which column is
Received Date, Category, GL Account, Name of Vendor, Invoice Number,
amount, …) — the same way the invoice pipeline maps a template's headers
— and code emits one row per employee in that order. A header the app has
no value for is left blank for a clean paste. If the header row cannot be
mapped, the output falls back to a fixed minimal set and says so.

The listing's past tabs also give the category judge its precedent: rows
whose invoice number is an ER(...) code, as "Name — Category (GL) — MYR".

Totals are recomputed independently from the emitted text with Decimal
and reconciled against the employees' report totals.
"""
from __future__ import annotations

import logging
import re
from decimal import Decimal
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from pydantic import BaseModel, Field, model_validator

from .. import telemetry
from ..docsource import SourceUnavailable, get_source
from ..model_layer import USAGE_LIMITS, create_agent
from .evidence import ai_call
from .models import ClaimEvidence, ClaimEmployee, ClaimFlag, ClaimRow, ClaimsRun

log = logging.getLogger("claims.listing")

ROLES = ("serial", "processed_by", "received_date", "p2p_ref", "po_number", "cost_center",
         "category", "gl_account", "vendor_name", "invoice_number", "amount", "remarks")

# The minimal set when the header row cannot be read.
FALLBACK_HEADER = ["Name", "ER code", "Category", "GL", "Amount (MYR)"]
FALLBACK_ROLES = ["vendor_name", "invoice_number", "category", "gl_account", "amount"]

_ER = re.compile(r"ER\s*\([^)]*\)", re.IGNORECASE)


class HeaderMap(BaseModel):
    """Which header (by its 0-based index in the header row) plays which role."""

    @model_validator(mode="after")
    def _distinct(self):
        seen: dict[int, str] = {}
        for role, idx in self.model_dump().items():
            if role in ("header_row", "why"):
                continue
            if isinstance(idx, int):
                if idx in seen:
                    raise ValueError(f"header {idx} given two roles ({seen[idx]} and {role})")
                seen[idx] = role
        return self

    header_row: int = Field(ge=1, description="the 1-based row holding the column headings")
    serial: int | None = Field(default=None, ge=0, description="S/N running number")
    processed_by: int | None = Field(default=None, ge=0)
    received_date: int | None = Field(default=None, ge=0)
    p2p_ref: int | None = Field(default=None, ge=0, description="helpline / P2P reference")
    po_number: int | None = Field(default=None, ge=0)
    cost_center: int | None = Field(default=None, ge=0)
    category: int | None = Field(default=None, ge=0)
    gl_account: int | None = Field(default=None, ge=0)
    vendor_name: int | None = Field(default=None, ge=0, description="name of vendor / payee (an employee, for claims)")
    invoice_number: int | None = Field(default=None, ge=0, description="invoice number (the ER(...) code, for claims)")
    amount: int | None = Field(default=None, ge=0, description="the amount column")
    remarks: int | None = Field(default=None, ge=0)
    why: str = Field(max_length=200)


_INSTRUCTIONS = (
    "You are shown the first rows of one tab of an accounts-payable 'Summary "
    "of Invoices' listing as a grid ('C3: Received Date'). Say which row holds "
    "the column headings and, for each role, the 0-based position of its "
    "heading in that row (column A = 0). Roles: serial (S/N), processed_by, "
    "received_date, p2p_ref, po_number, cost_center, category, gl_account, "
    "vendor_name (name of vendor / payee), invoice_number, amount, remarks. "
    "Leave a role null when no heading plays it. Answer with positions only."
)


def listing_path(run: ClaimsRun) -> Path | None:
    """The run's private copy of the listing, fetching it once if a link
    was given. None when there is neither a file nor a link."""
    from .runner import workspace_for

    ws = workspace_for(run.id)
    local = ws / "listing.xlsx"
    if local.is_file():
        return local
    if not run.listing_url:
        return None
    url = run.listing_url
    if Path(url).expanduser().is_file():
        local.write_bytes(Path(url).expanduser().read_bytes())
        return local
    # A SharePoint link to the file: its folder is the parent, its name the
    # last segment.
    from urllib.parse import unquote, urlsplit

    parts = urlsplit(url)
    path = unquote(parts.path)
    folder_url, _, name = url.rpartition("/")
    if not name:
        raise SourceUnavailable("The listing link does not end in a file name.")
    source = get_source(folder_url)
    # Finding the workbook and fetching it are one visit to one folder, so
    # they share one session and one site lookup (docsource.batch).
    with source.batch(folder_url):
        entries = source.list_folder(folder_url, "")
        entry = next((e for e in entries if e["name"] == unquote(name)), None)
        if entry is None:
            raise SourceUnavailable(
                f"{unquote(name)!r} is not in the linked folder ({path}).")
        local.write_bytes(source.download(folder_url, entry))
    return local


async def prepare_listing(db, run: ClaimsRun) -> None:
    """Read the linked listing once: header map + past examples. Stores the
    result on run.listing_headers with a state: ok / fallback / missing /
    unreadable. Never raises — a run without a listing still verifies; the
    output then uses the fallback set and a run-level flag says so."""
    try:
        path = listing_path(run)
    except SourceUnavailable as exc:
        run.listing_headers = {"state": "unreadable", "why": str(exc)}
        db.commit()
        telemetry.record(db, run.id, "output", telemetry.WARNING, "LISTING_UNREADABLE",
                         f"The linked listing could not be fetched: {exc}")
        return
    if path is None:
        run.listing_headers = {"state": "missing", "why": "no listing link or file was given"}
        db.commit()
        telemetry.record(db, run.id, "output", telemetry.WARNING, "LISTING_MISSING",
                         "No listing was linked to this run — the output will use the fallback columns.")
        return
    try:
        result = await read_listing(path)
    except Exception as exc:
        reason = telemetry.record_failure(db, run.id, "output", "LISTING_UNREADABLE",
                                          "Could not read the listing workbook", exc)
        run.listing_headers = {"state": "unreadable", "why": reason}
        db.commit()
        return
    from . import profile as profile_mod

    result = apply_listing_columns(result, profile_mod.profile_of(run.snapshot))
    run.listing_headers = result
    db.commit()
    telemetry.record(db, run.id, "output", telemetry.INFO, "LISTING_READ",
                     (f"Listing header read from tab {result['tab']!r}: {len(result['header'])} "
                      f"columns, roles for {sum(1 for r in result['roles'].values() if r is not None)}; "
                      f"{len(result['past_examples'])} past employee row(s) found as precedent.")
                     if result["state"] == "ok" else
                     f"Listing header could not be mapped ({result.get('why')}); the fallback "
                     "columns will be used and the output says so.")


async def read_listing(path: Path, usage=None) -> dict:
    """AI maps the header row of the CURRENT tab — the last one in the
    workbook, which is the month being filled; code reads the past ER rows
    from every tab."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets = wb.worksheets
        if not sheets:
            return {"state": "unreadable", "why": "the workbook has no tabs"}
        current = sheets[-1]
        grid = _grid(current)
        agent = create_agent("judge", HeaderMap, _INSTRUCTIONS, temperature=0)
        result = await ai_call(agent.run(grid, usage_limits=USAGE_LIMITS), "the listing header reader")
        if usage is not None:
            usage.add(result)
        hm = result.output
        header = _row_values(current, hm.header_row)
        roles = {r: getattr(hm, r) for r in ROLES}
        # Code checks the answer: positions inside the row, and the roles
        # the output cannot do without (a name, a number, an amount).
        for role, idx in roles.items():
            if idx is not None and (idx >= len(header) or not header[idx]):
                roles[role] = None
        must = ("vendor_name", "amount")
        if any(roles[r] is None for r in must) or not any(header):
            return {"state": "fallback", "tab": current.title, "header": FALLBACK_HEADER,
                    "roles": dict(zip(FALLBACK_ROLES, range(len(FALLBACK_ROLES)))),
                    "why": f"the header row of tab {current.title!r} did not yield a vendor and an "
                           f"amount column (AI: {hm.why})",
                    "past_examples": _past_examples(sheets, hm, header)}
        return {"state": "ok", "tab": current.title, "header_row": hm.header_row, "header": header,
                "roles": roles, "why": hm.why, "past_examples": _past_examples(sheets, hm, header)}
    finally:
        wb.close()


def apply_listing_columns(result: dict, profile: dict) -> dict:
    """The reviewer's pinned columns (profile listing_columns, H9) applied
    over the AI's header map: a role moves to the named header, 'blank'
    removes any role from it, '=text' writes that text in every row
    (stored as literals by column index). Unknown headers stay blank and
    visible; the universal roles stay required (a map that loses its
    vendor or amount column falls back)."""
    pins = {str(k).strip().lower(): str(v).strip() for k, v in (profile.get("listing_columns") or {}).items() if str(k).strip()}
    if not pins or result.get("state") not in ("ok", "fallback"):
        return result
    header = list(result.get("header") or [])
    roles = dict(result.get("roles") or {})
    literals: dict[str, str] = {}
    applied: list[str] = []
    for idx, text in enumerate(header):
        pin = pins.get(str(text).strip().lower())
        if pin is None:
            continue
        for role, i in list(roles.items()):
            if i == idx:
                roles[role] = None
        if pin == "blank":
            applied.append(f"{text!r}: blank")
        elif pin.startswith("="):
            literals[str(idx)] = pin[1:]
            applied.append(f"{text!r}: literal {pin[1:]!r}")
        elif pin in ROLES:
            roles[pin] = idx
            applied.append(f"{text!r}: {pin}")
    out = {**result, "roles": roles, "literals": literals, "pinned_columns": applied}
    if result.get("state") == "ok" and (roles.get("vendor_name") is None or roles.get("amount") is None):
        out["state"] = "fallback"
        out["header"], out["roles"] = list(FALLBACK_HEADER), dict(zip(FALLBACK_ROLES, range(len(FALLBACK_ROLES))))
        out["literals"] = {}
        out["why"] = "the pinned listing columns leave no vendor or amount column — check Settings → Claims"
    return out


def _row_values(ws, row: int) -> list[str]:
    for r in ws.iter_rows(min_row=row, max_row=row, values_only=True):
        vals = ["" if v is None else str(v).strip() for v in r]
        while vals and not vals[-1]:
            vals.pop()
        return vals
    return []


def _grid(ws, max_rows: int = 12) -> str:
    from openpyxl.utils import get_column_letter

    lines = [f"Sheet name: {ws.title!r}"]
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), 1):
        cells = [f"{get_column_letter(c)}{r}: {str(v).strip()[:60]}"
                 for c, v in enumerate(row, 1) if v is not None and str(v).strip()]
        if cells:
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def _past_examples(sheets, hm: HeaderMap, header: list[str]) -> list[str]:
    """'Nick Goh — Taxi (713070) — MYR 240.00' for every past row whose
    invoice number is an ER(...) code. Assumes the past tabs share the
    current tab's layout (they are the same file)."""
    out = []
    if hm.invoice_number is None or hm.vendor_name is None:
        return out
    for ws in sheets:
        for row in ws.iter_rows(min_row=hm.header_row + 1, max_row=hm.header_row + 400, values_only=True):
            vals = ["" if v is None else str(v).strip() for v in row]
            inv = vals[hm.invoice_number] if hm.invoice_number < len(vals) else ""
            if not _ER.search(inv):
                continue
            name = vals[hm.vendor_name] if hm.vendor_name < len(vals) else ""
            cat = vals[hm.category] if hm.category is not None and hm.category < len(vals) else ""
            gl = vals[hm.gl_account] if hm.gl_account is not None and hm.gl_account < len(vals) else ""
            amt = vals[hm.amount] if hm.amount is not None and hm.amount < len(vals) else ""
            out.append(f"{name} — {cat}" + (f" ({gl})" if gl else "") + (f" — MYR {amt}" if amt else "")
                       + f" [tab {ws.title}]")
    return out[:40]


# ---- the batch rows -----------------------------------------------------------------

def _cell(value) -> str:
    """One TSV cell, made safe: no tabs/newlines, no leading formula characters."""
    s = re.sub(r"[\x00-\x1f]", " ", str(value if value is not None else "")).strip()
    if s[:1] in ("=", "+", "-", "@"):
        s = "'" + s
    return s


def build_outputs(db, run: ClaimsRun) -> dict:
    """One Payment Listing Row per confirmed, verified Claim Case, in the
    client's own column order (H9).

    Included = confirmed cases whose worker verified, with a confirmed
    claimant, not excluded. An accepted flag excludes its ROW (the case
    stays with the remaining rows); excluded rows are subtracted and named.
    Cases that failed, were skipped, or have no confirmed claimant are
    listed under not_included with the reason. Three totals are kept
    apart and named: the Reported Total (what each source states, or
    absent), the Calculated Lines Total (the Decimal sum of the lines to
    be paid), and the emitted total re-summed from the output text; every
    missing comparison is named rather than counted as a match.
    """
    from .models import ClaimCase

    lh = run.listing_headers or {}
    if lh.get("state") == "ok":
        header, roles = list(lh["header"]), dict(lh["roles"])
        fallback, note = False, ""
    else:
        header, roles = list(FALLBACK_HEADER), dict(zip(FALLBACK_ROLES, range(len(FALLBACK_ROLES))))
        fallback = True
        note = ("The listing's header row could not be read"
                + (f" ({lh.get('why')})" if lh.get("why") else "")
                + ", so these rows use the fallback columns; paste them by hand into the right places.")
    literals: dict[int, str] = {int(k): v for k, v in (lh.get("literals") or {}).items()}
    cases = db.query(ClaimCase).filter(ClaimCase.run_id == run.id).all()
    employees = {e.id: e for e in db.query(ClaimEmployee).filter(ClaimEmployee.run_id == run.id).all()}
    # A run made before the case model has employees and no cases: treat
    # each employee as its own confirmed case (the migration does the same).
    from . import cases as cases_mod

    if not cases and employees:
        cases = [cases_mod.sync_case_from_employee(db, e) for e in employees.values()]
    # The employee record is the worker's unit during the compatibility
    # period: its totals/category/status are mirrored onto the case here,
    # so a value changed on one is what the listing sees on the other.
    for c in cases:
        emp = employees.get(c.legacy_employee_id)
        if emp is not None:
            cases_mod._mirror(c, emp)
    rows_by_case: dict[str, list[ClaimRow]] = {}
    for r in db.query(ClaimRow).filter(ClaimRow.run_id == run.id).all():
        rows_by_case.setdefault(r.case_id or _case_of_employee(cases, r.employee_id), []).append(r)
    excluded_rows = {f.row_id for f in db.query(ClaimFlag).filter(
        ClaimFlag.run_id == run.id, ClaimFlag.status == "accepted") if f.row_id}
    # Evidence no row used — what will NOT be paid, on the same screen as
    # what will, with the reviewer's decision on it if any.
    case_name = {c.id: (c.claimant_name or c.label) for c in cases}
    emp_to_case = {c.legacy_employee_id: c.id for c in cases if c.legacy_employee_id}
    decisions = {}
    for f in db.query(ClaimFlag).filter(ClaimFlag.run_id == run.id, ClaimFlag.evidence_id != "",
                                        ClaimFlag.code.in_(("UNCLAIMED_RECEIPT", "MILEAGE_NO_MAP"))):
        if f.status not in ("open", "info"):
            decisions[f.evidence_id] = f.status.replace("_", " ") + (f" — {f.resolution}" if f.resolution else "")
    # A receipt two rows both claim, or one of several candidates for a
    # row, is CONTESTED, not spare: its flag is the decision to make, and
    # listing it again as evidence nobody used would double-count it.
    contested: set[str] = set()
    for f in db.query(ClaimFlag).filter(ClaimFlag.run_id == run.id,
                                        ClaimFlag.code.in_(("DUPLICATE_RECEIPT", "RECEIPT_AMBIGUOUS"))):
        if f.evidence_id:
            contested.add(f.evidence_id)
        contested.update(str(x) for x in ((f.cite or {}).get("candidates") or []))
    unused_evidence = []
    for ev in db.query(ClaimEvidence).filter(ClaimEvidence.run_id == run.id, ClaimEvidence.matched_row_id == ""):
        if ev.id in contested:
            continue
        v = ev.values or {}
        if ev.kind == "receipt":
            what = f"receipt from {v.get('vendor') or '?'} ({v.get('date') or 'no date'})"
            amount = f"{v.get('currency') or 'MYR'} {v.get('amount') or '?'}"
        else:
            what = f"map trip ({v.get('date') or 'no date'}, {v.get('km_printed') or '?'} km)"
            amount = ""
        cid = ev.case_id or emp_to_case.get(ev.employee_id, "")
        unused_evidence.append({"name": case_name.get(cid, "?"), "case_id": cid, "what": what,
                                "where": f"{ev.file} page {ev.page}" + (f", {ev.position}" if ev.position else ""),
                                "amount": amount, "decision": decisions.get(ev.id, "")})
    unused_evidence.sort(key=lambda u: (u["name"], u["where"]))

    out_rows: list[list[str]] = []
    included, not_included, exclusions = [], [], []
    reported_total_sum = Decimal("0")   # the independent side, where present
    lines_total_sum = Decimal("0")      # Calculated Lines Total of the lines to be paid
    reported_missing = 0
    held_lines = 0                      # lines held out of the sum for want of a rate
    differences: list[dict] = []
    n = 0                               # the S/N counts the rows emitted, not the cases seen
    for c in sorted(cases, key=lambda x: (x.claimant_name or x.label).lower()):
        name = c.claimant_name or c.label
        if c.state == "excluded" or c.status == "skipped":
            not_included.append({"name": name, "case_id": c.id, "why": "excluded by the reviewer at the map"})
            continue
        if c.status != "verified":
            not_included.append({"name": name, "case_id": c.id,
                                 "why": {"failed": f"verification failed: {c.error}", "pending": "not verified yet",
                                         "verifying": "still verifying"}.get(c.status, c.status)})
            continue
        if c.claimant_state != "confirmed":
            not_included.append({"name": name, "case_id": c.id,
                                 "why": "no confirmed claimant — nobody to pay; set the claimant on the case"})
            continue
        c_rows = [r for r in rows_by_case.get(c.id, []) if r.kind != "mileage"]
        kept = [r for r in c_rows if r.id not in excluded_rows]
        gone = [r for r in c_rows if r.id in excluded_rows]
        if c_rows and not kept:
            not_included.append({"name": name, "case_id": c.id, "why": "every row was excluded in review"})
            continue
        # A line whose figure is not ringgit and carries no MYR total has
        # had no rate applied: it is named and held out, never added to
        # the sum as though its number were ringgit.
        held, payable = [], []
        for r in kept:
            why = _held_why(r)
            if why:
                held.append(r)
                held_lines += 1
                not_included.append({"name": name, "case_id": c.id, "why": why})
            else:
                payable.append(r)
        if not payable:
            # the Reported Total is what the source SAYS; it never stands
            # in for lines nobody could read (H11: the two stay apart)
            not_included.append({"name": name, "case_id": c.id,
                                 "why": "no lines to pay" + (
                                     " — every line is held for want of an exchange rate" if held else
                                     "; the case's Reported Total is not paid on its own — read the "
                                     "lines, or exclude the case")})
            continue
        amount = sum((_money(r) for r in payable), Decimal("0")).quantize(Decimal("0.01"))
        n += 1
        for r in gone:
            exclusions.append({"name": name, "case_id": c.id, "row": r.row, "amount": str(_money(r)),
                               "why": "flag accepted in review"})
        values = {"serial": str(n), "processed_by": "", "received_date": run.received_date,
                  "p2p_ref": "", "po_number": "", "cost_center": "", "category": c.category,
                  "gl_account": c.gl, "vendor_name": c.claimant_name, "invoice_number": c.claimant_identifier,
                  "amount": f"{amount:.2f}",
                  "remarks": "employee claim" + (f" ({len(gone)} row(s) excluded)" if gone else "")
                             + (" — lines derived from evidence" if any(r.kind == "derived" for r in payable) else "")}
        line = [""] * len(header)
        for role, idx in roles.items():
            if idx is not None and idx < len(header):
                line[idx] = _cell(values.get(role, ""))
        for idx, text in literals.items():
            if 0 <= idx < len(header):
                line[idx] = _cell(text)
        out_rows.append(line)
        lines_total_sum += amount
        reported_total = _dec_or_none(c.reported_total)
        included.append({"name": c.claimant_name, "case_id": c.id, "er_code": c.claimant_identifier,
                         "amount": f"{amount:.2f}", "category": c.category, "gl": c.gl,
                         "reported_total": f"{reported_total:.2f}" if reported_total is not None else None,
                         "lines_total": f"{amount:.2f}",
                         "derived": any(r.kind == "derived" for r in payable)})
        if reported_total is None:
            # nothing independent to check against: say so rather than pretend
            reported_missing += 1
            differences.append({"name": name, "case_id": c.id, "expected": None, "emitted": f"{amount:.2f}",
                                "why": "the source carried no Reported Total to reconcile against"})
        else:
            expected = (reported_total - sum((_money(r) for r in gone), Decimal("0"))).quantize(Decimal("0.01"))
            reported_total_sum += expected
            if expected != amount:
                differences.append({"name": name, "case_id": c.id, "expected": f"{expected:.2f}",
                                    "emitted": f"{amount:.2f}",
                                    "why": (f"the Reported Total is {reported_total:.2f}"
                                            + (f" ({expected:.2f} after {len(gone)} excluded row(s))" if gone else "")
                                            + f", but the lines to be paid sum to {amount:.2f} — "
                                            "a corrected amount, or a line the source total does not cover")})
    amount_idx = roles.get("amount")
    emitted_total = sum((Decimal(r[amount_idx].lstrip("'") or "0") for r in out_rows), Decimal("0")) \
        if amount_idx is not None else Decimal("0")
    # The comparison is named per case; the batch figure compares only the
    # cases that HAVE a Reported Total, and says how many have none.
    compared = reported_total_sum + sum((Decimal(d["emitted"]) for d in differences if d["expected"] is None), Decimal("0"))
    match = (emitted_total == lines_total_sum and not any(d["expected"] is not None for d in differences))
    tsv = "\n".join("\t".join(_cell(h) for h in header) + "" for _ in [0]) + "\n" + \
        "\n".join("\t".join(r) for r in out_rows)
    return {"header": header, "rows": out_rows, "tsv": tsv,
            "totals": {"total_myr": f"{emitted_total:.2f}",
                       "lines_total": f"{lines_total_sum:.2f}",
                       "reported_total": f"{reported_total_sum:.2f}",
                       "reported_missing": reported_missing,
                       "held_lines": held_lines,
                       "source_total": f"{compared:.2f}",   # delivered name: reported where present, else the lines
                       "match": match,
                       "difference": f"{(emitted_total - compared):.2f}",
                       "differences": differences},
            "included": included, "not_included": not_included, "exclusions": exclusions,
            "unused_evidence": unused_evidence,
            "header_fallback": fallback, "header_note": note, "received_date": run.received_date}


def _case_of_employee(cases, employee_id: str) -> str:
    return next((c.id for c in cases if c.legacy_employee_id == employee_id), "")


def _dec_or_none(text) -> Decimal | None:
    if text is None or str(text).strip() == "":
        return None
    try:
        d = Decimal(str(text))
        return d if d.is_finite() else None
    except Exception:
        return None


def _money(r: ClaimRow) -> Decimal:
    """What the line pays, in MYR: its total where the source computed one,
    else its amount. A line _held_why() names pays nothing."""
    if _held_why(r):
        return Decimal("0")
    v = r.values or {}
    return _dec_or_none(v.get("total")) or _dec_or_none(v.get("amount")) or Decimal("0")


def _held_why(r: ClaimRow) -> str:
    """Why this line is held out of the payment, or "". A figure in a
    foreign currency with no MYR total has had no exchange rate applied:
    adding it to a ringgit sum would pay USD 50.00 as RM 50.00."""
    v = r.values or {}
    currency = (str(v.get("currency") or "MYR").strip().upper() or "MYR")
    if currency == "MYR" or _dec_or_none(v.get("total")) is not None:
        return ""
    amount = _dec_or_none(v.get("amount"))
    return (f"a line of {currency} {amount if amount is not None else '?'} carries no MYR total — no "
            "exchange rate was applied, so it is held out of the payment; enter the rate (or the "
            "converted total) on the source line and re-run")
