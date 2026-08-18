"""The report reader's audit and the check catalogue, against the sample.

The AI is scripted (report reader) or absent (checks — the tie-break is
stubbed), so these pin the CODE half: a correct reading extracts every
row of every report exactly as the ground truth says; wrong readings are
sent back; and, given a perfect page inventory, every planted error is
flagged with the expected code, the must-not-flag cases are not, false
flags stay under one per employee, and every flag cites a place.
"""
from __future__ import annotations

import asyncio
import json
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.claims import checks, report_reader
from app.claims.report_reader import ReportColumns, ReportReading, ReportUnreadable

GEN = Path(__file__).resolve().parents[2] / "samples" / "generated" / "claims"
needs_sample = pytest.mark.skipif(not GEN.is_dir(), reason="run samples/generate_claims_sample.py first")

PROFILE = {"mileage_rates": {"Car": "0.64", "Motorcycle": "0.35"}, "km_tolerance": "0",
           "receipt_date_window_days": 0, "receipt_optional_items": ["Mobile Allowance"],
           "mileage_item_pattern": "mileage", "categories": [], "category_rule": "",
           "file_role_patterns": [], "checks": {}, "set_by": {}}


def _truth():
    return json.loads((GEN / "ground_truth_claims.json").read_text())


def _good_reading(ws) -> ReportReading:
    """The reading a competent AI gives for the sample's report tab."""
    last = max(r for r in range(7, ws.max_row + 1)
               if ws.cell(row=r, column=1).value is not None)
    total_row = next(r for r in range(last, ws.max_row + 1) if ws.cell(row=r, column=7).value == "Total (MYR)")
    return ReportReading(
        columns=ReportColumns(date="A", item="B", reason="C", receipt_included="D", amount="E",
                              currency="F", rate="G", total="H"),
        header_row=6, first_row=7, last_row=last, total_cell=f"H{total_row}", name_cell="B1",
        period_cell="B2", purpose_cell="B3", why="header block, dated lines, total")


class _Scripted:
    def __init__(self, outputs):
        self._outputs = list(outputs)
        self.prompts = []

    async def run(self, prompt, **kw):
        self.prompts.append(prompt)

        class R:
            output = self._outputs.pop(0)

            def usage(self):
                class U:
                    total_tokens = 10
                return U()
        return R()


@needs_sample
@pytest.mark.asyncio
async def test_correct_reading_extracts_every_row_of_every_report(monkeypatch):
    truth = _truth()
    for e in truth["employees"]:
        if not e["files"]["report"]:
            continue
        wb = load_workbook(GEN / "batch" / e["folder"] / e["files"]["report"], data_only=True)
        ws = wb["Expense Report"]
        agent = _Scripted([_good_reading(ws)])
        monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
        rows, header, notes = await report_reader.read_report(ws, e["name"], e["er_code"])
        assert len(rows) == len(e["report_rows"]), e["name"]
        for got, want in zip(rows, e["report_rows"]):
            assert got["row"] == want["row"]
            assert got["date"] == want["date"]
            assert got["item"] == want["item"]
            assert got["receipt_included"] == want["receipt"]
            assert Decimal(got["amount"]) == Decimal(str(want["amount"]))
            assert got["currency"] == want["currency"]
            assert Decimal(got["rate"]) == Decimal(str(want["rate"]))
            assert Decimal(got["total"]) == Decimal(str(want["total"]))
        assert Decimal(header["total"]) == Decimal(str(e["report_total"]))
        assert header["name"] == e["name"] and header["purpose"] == e["purpose"]
        assert len(agent.prompts) == 1


@needs_sample
@pytest.mark.asyncio
async def test_wrong_reading_is_sent_back_then_accepted(monkeypatch):
    e = _truth()["employees"][0]
    wb = load_workbook(GEN / "batch" / e["folder"] / e["files"]["report"], data_only=True)
    ws = wb["Expense Report"]
    good = _good_reading(ws)
    wrong = good.model_copy(update={"columns": good.columns.model_copy(update={"amount": "H", "total": "E"}),
                                    "total_cell": None, "name_cell": "A1"})
    agent = _Scripted([wrong, good])
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    rows, header, notes = await report_reader.read_report(ws, e["name"], e["er_code"])
    assert len(agent.prompts) == 2
    fb = agent.prompts[1]
    assert "total_cell is missing" in fb and "name_cell A1 holds" in fb
    assert len(rows) == len(e["report_rows"])


@needs_sample
@pytest.mark.asyncio
async def test_scrambled_reading_ends_as_unreadable(monkeypatch):
    e = _truth()["employees"][0]
    wb = load_workbook(GEN / "batch" / e["folder"] / e["files"]["report"], data_only=True)
    ws = wb["Expense Report"]
    good = _good_reading(ws)
    bad = good.model_copy(update={"first_row": 8, "total_cell": good.total_cell})  # drops row 7
    agent = _Scripted([bad] * report_reader.MAX_ROUNDS)
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    with pytest.raises(ReportUnreadable) as exc:
        await report_reader.read_report(ws, e["name"], e["er_code"])
    assert "sum to" in str(exc.value)


@needs_sample
@pytest.mark.asyncio
async def test_km_tab_reads_and_no_trips_is_fine(monkeypatch):
    from app.claims.report_reader import KMColumns, KMReading

    truth = _truth()
    nick = next(e for e in truth["employees"] if e["name"] == "Nick Goh")
    wb = load_workbook(GEN / "batch" / nick["folder"] / nick["files"]["report"], data_only=True)
    ws = wb["KM"]
    reading = KMReading(has_trips=True, columns=KMColumns(date="A", **{"from": "B"}, to="C", purpose="D",
                                                          vehicle="E", km="F", rate="G", amount="H"),
                        header_row=4, first_row=5, last_row=8, total_cell="H9", why="trips")
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: _Scripted([reading]))
    trips, notes = await report_reader.read_km(ws)
    assert [Decimal(t["km"]) for t in trips] == [Decimal(str(k["km"])) for k in nick["km_rows"]]
    assert trips[0]["date"] == "2026-07-06"
    empty = KMReading(has_trips=False, why="headings only")
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: _Scripted([empty]))
    trips, notes = await report_reader.read_km(ws)
    assert trips == []


def test_helpers():
    assert report_reader.er_period("ER(01JUL26-21JUL26)")[0].isoformat() == "2026-07-01"
    assert report_reader.er_period("weird") is None
    assert report_reader.gl_of("Taxi (713070)") == "713070"
    assert report_reader.item_name("Taxi (713070)") == "Taxi"
    assert report_reader.money("RM 1,200.50") == Decimal("1200.50")
    assert report_reader.money("(5.00)") == Decimal("-5.00")
    assert report_reader.money("nan") is None


@needs_sample
def test_read_categories_from_expense_types_tab():
    e = _truth()["employees"][0]
    wb = load_workbook(GEN / "batch" / e["folder"] / e["files"]["report"], data_only=True)
    cats = report_reader.read_categories(wb)
    assert len(cats) == 23
    assert {"item": "Taxi", "gl": "713070"} in cats


# ---- the check catalogue over a perfect inventory ---------------------------------

def _rows_and_evidence(e: dict) -> tuple[list[dict], list[dict]]:
    rows, evidence = [], []
    for i, r in enumerate(e["report_rows"]):
        rows.append({"id": f"r{i}", "kind": "expense", "sheet": "Expense Report", "row": r["row"],
                     "values": {"date": r["date"], "item": r["item"],
                                "item_name": report_reader.item_name(r["item"]),
                                "gl": report_reader.gl_of(r["item"]), "reason": r["reason"],
                                "receipt_included": r["receipt"], "amount": f"{r['amount']:.2f}",
                                "currency": r["currency"], "rate": str(r["rate"]), "total": f"{r['total']:.2f}"}})
    for i, k in enumerate(e["km_rows"]):
        rows.append({"id": f"k{i}", "kind": "mileage", "sheet": "KM", "row": k["row"],
                     "values": {"date": k["date"], "from": k["from"], "to": k["to"], "vehicle": k["vehicle"],
                                "km": str(k["km"]), "rate": str(k["rate"]), "amount": f"{k['amount']:.2f}"}})
    for i, r in enumerate(e["receipts"]):
        evidence.append({"id": f"e{i}", "kind": "receipt", "file": r["file"], "page": r["page"],
                         "position": r["position"],
                         "values": {"vendor": r["vendor"], "date": r["date"], "amount": f"{r['amount']:.2f}",
                                    "currency": r["currency"]}, "confidence": {}})
    for i, t in enumerate(e["map_trips"]):
        evidence.append({"id": f"t{i}", "kind": "map_trip", "file": t["file"], "page": t["page"], "position": "",
                         "values": {"date": t["date"], "purpose": t["purpose"], "from": "", "to": "",
                                    "return_trip": t["return_trip"], "km_printed": str(t["km_printed"])},
                         "confidence": {}})
    return rows, evidence


@needs_sample
def test_every_planted_error_is_flagged_and_nothing_else():
    truth = _truth()
    total_false = 0
    for e in truth["employees"]:
        rows, evidence = _rows_and_evidence(e)
        pages = max([r["page"] for r in e["receipts"]] + [t["page"] for t in e["map_trips"]] + [0])
        result = asyncio.run(checks.run_checks(rows, evidence, PROFILE, {"name": e["name"]},
                                               (pages, len(e["files"]["receipts"])), tie_break=None))
        flags = result["flags"]
        for f in flags:
            assert f["cite"], f"{e['name']}: flag {f['code']} has no citation"
            assert f["basis"], f"{e['name']}: flag {f['code']} has no basis"
        open_codes = sorted(f["code"] for f in flags if f["status"] == "open")
        info_codes = sorted(f["code"] for f in flags if f["status"] == "info")
        expected = [p["code"] for p in e["expected_flags"] if p["code"] not in ("NO_REPORT",)]
        # every planted error found (info-level ones may be info)
        for code in expected:
            assert code in open_codes + info_codes, f"{e['name']}: expected {code}, got {open_codes} / {info_codes}"
        # false flags: open flags beyond the expected ones (NO_REPORT is the worker's)
        extra = list(open_codes)
        for code in expected:
            if code in extra:
                extra.remove(code)
        assert len(extra) <= 1, f"{e['name']}: unexpected open flags {extra}"
        total_false += len(extra)
        # must-not-flag: the return trip / mobile allowance N / clean SGD row
        for m in e["must_not_flag"]:
            if "return trip" in m["what"]:
                assert not any(f["code"] == "MILEAGE_DISCREPANCY" and "2026-07-06" in f["reason"] for f in flags)
            if "Mobile Allowance" in m["what"]:
                assert not any(f["code"] == "NO_RECEIPT" and f["status"] == "open"
                               and "Mobile Allowance" in f["reason"] for f in flags)
            if "clean foreign" in m["what"]:
                assert not any(f["code"] == "CURRENCY_MISMATCH" and "350" in f["reason"] for f in flags)
    assert total_false <= len(truth["employees"])


@needs_sample
def test_the_rm10_row_names_the_receipt_and_the_difference():
    e = next(x for x in _truth()["employees"] if x["name"] == "Aegene Ong")
    rows, evidence = _rows_and_evidence(e)
    result = asyncio.run(checks.run_checks(rows, evidence, PROFILE, {}, (4, 2)))
    nr = [f for f in result["flags"] if f["code"] == "NO_RECEIPT" and f["status"] == "open"]
    assert len(nr) == 1
    assert "RM 10.00 more" in nr[0]["reason"] and "page" in nr[0]["reason"]
    assert nr[0]["cite"].get("position") in ("left", "middle", "right")
    # fixing the amount to 35.00 resolves it: no NO_RECEIPT and no UNCLAIMED
    row = next(r for r in rows if r["values"]["amount"] == "45.00")
    row["values"]["amount"] = row["values"]["total"] = "35.00"
    result = asyncio.run(checks.run_checks(rows, evidence, PROFILE, {}, (4, 2)))
    assert not [f for f in result["flags"] if f["code"] in ("NO_RECEIPT", "UNCLAIMED_RECEIPT")
                and f["status"] == "open"]
    assert not [f for f in result["flags"] if f["code"] == "UNCLAIMED_RECEIPT"]


def test_missing_rates_is_a_run_level_control():
    rows = [{"kind": "mileage"}]
    assert "mileage rates" in checks.needs_missing_reference(rows, {**PROFILE, "mileage_rates": {}})
    assert checks.needs_missing_reference(rows, PROFILE) == ""
    assert checks.needs_missing_reference([{"kind": "expense"}], {**PROFILE, "mileage_rates": {}}) == ""


def test_ambiguous_receipts_go_to_the_tie_break_then_a_person():
    rows = [{"id": "r1", "kind": "expense", "sheet": "S", "row": 7,
             "values": {"date": "2026-07-01", "item": "Taxi", "item_name": "Taxi", "reason": "Grab to X",
                        "receipt_included": "Y", "amount": "20.00", "currency": "MYR", "rate": "1", "total": "20.00"}}]
    ev = [{"id": f"e{i}", "kind": "receipt", "file": "f.pdf", "page": 1, "position": p,
           "values": {"vendor": v, "date": "2026-07-01", "amount": "20.00", "currency": "MYR"}, "confidence": {}}
          for i, (p, v) in enumerate([("left", "Grab"), ("right", "AirAsia Ride")])]

    async def pick_first(row, cands):
        return cands[0]["id"]

    r = asyncio.run(checks.run_checks(rows, ev, PROFILE, {}, (1, 1), tie_break=pick_first))
    assert r["verdicts"]["r1"] == ("matched", "e0")
    assert [f["code"] for f in r["flags"]] == ["UNCLAIMED_RECEIPT"]

    async def unsure(row, cands):
        return ""

    r = asyncio.run(checks.run_checks(rows, ev, PROFILE, {}, (1, 1), tie_break=unsure))
    codes = [f["code"] for f in r["flags"] if f["status"] == "open"]
    assert codes == ["RECEIPT_AMBIGUOUS"]
    assert "left" in r["flags"][0]["reason"] and "right" in r["flags"][0]["reason"]


def test_checks_can_be_switched_off_per_client():
    rows = [{"id": "r1", "kind": "expense", "sheet": "S", "row": 7,
             "values": {"date": "2026-07-01", "item": "Taxi", "item_name": "Taxi", "reason": "",
                        "receipt_included": "Y", "amount": "20.00", "currency": "MYR", "rate": "1", "total": "20.00"}}]
    r = asyncio.run(checks.run_checks(rows, [], {**PROFILE, "checks": {"NO_RECEIPT": False}}, {}, (0, 0)))
    assert r["flags"] == []
