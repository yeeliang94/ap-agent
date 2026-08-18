"""Opt-in evaluation of the REAL model against a real (anonymised) listing.

Everything else in this suite scripts the AI. This one does not: it runs
the actual reason/act loop on a workbook you point it at, so a model change
(or a prompt change) is caught here before a client sees it. It costs AI
calls and needs a key, so it is skipped unless asked for:

    AP_LISTING_EVAL=/path/to/anonymised_listing.xlsx pytest tests/test_listing_eval.py -s

Beside the workbook, put <same name>.expected.json describing what a
correct reading must find — only what you are sure of; leave the rest out:

    {
      "tabs": {
        "Cover":  {"payment_sheet": false},
        "Apr'26": {"payment_sheet": true, "invoice_rows": 27,
                   "contains": ["4115", "580261111513"]},
        "Jul'26": {"payment_sheet": true, "contains": ["245DHNQL-0015"],
                   "amount_of": {"245DHNQL-0015": 1044.95}}
      },
      "max_warnings": 0
    }

  payment_sheet  whether the tab must be read as payments (or skipped)
  invoice_rows   exact number of invoice rows the tab must yield (optional)
  contains       invoice numbers that must be among the tab's rows
  amount_of      invoice number -> the amount the row must carry
  max_warnings   how many WARNING lines the whole reading may emit (optional)

The per-tab Activity lines are printed (-s) so a failure can be read.
"""
from __future__ import annotations

import io
import json
import os
from pathlib import Path

import pytest

WORKBOOK = os.environ.get("AP_LISTING_EVAL", "")

pytestmark = pytest.mark.skipif(
    not WORKBOOK, reason="set AP_LISTING_EVAL=<workbook.xlsx> to run the real-model evaluation")


@pytest.mark.asyncio
async def test_real_model_reads_the_anonymised_listing_as_expected():
    from openpyxl import load_workbook

    from app.pipeline import listing_agent

    path = Path(WORKBOOK)
    expected_path = path.with_name(path.stem + ".expected.json")
    assert path.is_file(), f"no such workbook: {path}"
    assert expected_path.is_file(), f"put the expectations in {expected_path}"
    expected = json.loads(expected_path.read_text())

    data = path.read_bytes()
    wb = load_workbook(io.BytesIO(data), data_only=True)
    wb_formulas = load_workbook(io.BytesIO(data))
    result = await listing_agent.ingest_workbook(wb, wb_formulas)

    print()
    for level, text in result.notes:
        print(f"  [{level}] {text}")

    by_tab: dict[str, list[dict]] = {}
    for r in result.rows:
        by_tab.setdefault(r["sheet"], []).append(r)
    failures: list[str] = []
    for tab, want in expected.get("tabs", {}).items():
        rows = by_tab.get(tab, [])
        if want.get("payment_sheet") is True and not rows:
            failures.append(f"{tab}: expected a payment sheet, got no rows")
        if want.get("payment_sheet") is False and rows:
            failures.append(f"{tab}: expected to be skipped, got {len(rows)} rows")
        if "invoice_rows" in want and len(rows) != want["invoice_rows"]:
            failures.append(f"{tab}: expected {want['invoice_rows']} rows, got {len(rows)}")
        numbers = {r["invoice_number"] for r in rows}
        for n in want.get("contains", []):
            if n not in numbers:
                failures.append(f"{tab}: invoice {n} not found")
        for n, amount in want.get("amount_of", {}).items():
            got = [r["amount"] for r in rows if r["invoice_number"] == n]
            if not got or got[0] is None or abs(got[0] - float(amount)) > 0.005:
                failures.append(f"{tab}: invoice {n} amount {got} != {amount}")
    if "max_warnings" in expected:
        warnings = [t for lvl, t in result.notes if lvl == "WARNING"]
        if len(warnings) > expected["max_warnings"]:
            failures.append(f"{len(warnings)} warning(s), allowed {expected['max_warnings']}")
    assert not failures, "\n".join(failures)
