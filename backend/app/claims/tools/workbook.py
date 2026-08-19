"""Workbook tools (H4): `inspect_workbook` and `read_cells`, read-only, over
the run snapshot. Nothing is ever executed: formulas are REPORTED as text
beside their saved value (data_only read), macros are never loaded
(keep_vba=False, and openpyxl has no VBA engine), links and embedded
objects are counted, not followed.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from .contracts import MAX_CELL_CHARS, MAX_CELLS

_CELL = re.compile(r"^([A-Za-z]{1,3})(\d{1,7})$")
INSPECT_ROWS = 2000     # how far inspect looks for the used range / formulas
INSPECT_COLS = 60


def _col_num(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _col_letters(n: int) -> str:
    out = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        out = chr(65 + r) + out
    return out


def parse_range(cell_range: str) -> tuple[int, int, int, int]:
    """'A1:H20' → (col0, row0, col1, row1); 'B3' is a single cell."""
    text = (cell_range or "").replace("$", "").strip()
    a, _, b = text.partition(":")
    b = b or a
    ma, mb = _CELL.match(a), _CELL.match(b)
    if not ma or not mb:
        raise ValueError("range must look like A1:H20")
    c0, r0, c1, r1 = _col_num(ma.group(1)), int(ma.group(2)), _col_num(mb.group(1)), int(mb.group(2))
    if c1 < c0 or r1 < r0:
        raise ValueError("range end is before its start")
    return c0, r0, c1, r1


def _text(v) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    return s if len(s) <= MAX_CELL_CHARS else s[:MAX_CELL_CHARS] + "…"


def inspect_workbook(path: Path) -> dict:
    """Sheets, used ranges, cell/formula counts, hidden state, merged
    ranges, defined names — structure only, no values."""
    from openpyxl import load_workbook

    wb_f = load_workbook(path, read_only=True, keep_vba=False)  # formulas as text
    try:
        sheets = []
        for ws in wb_f.worksheets:
            n_cells = n_formulas = 0
            max_row = max_col = 0
            for r, row in enumerate(ws.iter_rows(min_row=1, max_row=INSPECT_ROWS, max_col=INSPECT_COLS,
                                                 values_only=True), 1):
                for c, v in enumerate(row, 1):
                    if v is None or (isinstance(v, str) and not v.strip()):
                        continue
                    n_cells += 1
                    max_row, max_col = max(max_row, r), max(max_col, c)
                    if isinstance(v, str) and v.startswith("="):
                        n_formulas += 1
            merged = []
            try:
                merged = [str(m) for m in list(getattr(ws, "merged_cells", []).ranges)[:50]]
            except Exception:
                merged = []
            sheets.append({"name": ws.title, "state": getattr(ws, "sheet_state", "visible"),
                           "hidden": getattr(ws, "sheet_state", "visible") != "visible",
                           "max_row": max_row, "max_col": max_col,
                           "used_range": f"A1:{_col_letters(max_col)}{max_row}" if max_row else "",
                           "cells": n_cells, "formulas": n_formulas, "merged": merged,
                           "truncated": ws.max_row is not None and ws.max_row > INSPECT_ROWS})
        names = []
        try:
            names = [n for n in list(wb_f.defined_names)[:50]] if hasattr(wb_f.defined_names, "__iter__") else []
        except Exception:
            names = []
        return {"sheets": sheets, "defined_names": [str(n) for n in names],
                "macros_present": path.suffix.lower() == ".xlsm", "macros_executed": False}
    finally:
        wb_f.close()


def read_cells(path: Path, sheet: str, cell_range: str) -> dict:
    """Exact saved values (and the formula text where a cell holds one) for
    a bounded range. Never computes a formula."""
    from openpyxl import load_workbook

    c0, r0, c1, r1 = parse_range(cell_range)
    if (c1 - c0 + 1) * (r1 - r0 + 1) > MAX_CELLS:
        raise ValueError(f"range holds more than {MAX_CELLS} cells — read it in parts")
    wb_v = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    wb_f = load_workbook(path, read_only=True, keep_vba=False)
    try:
        if sheet not in wb_v.sheetnames:
            raise KeyError(f"no sheet {sheet!r} (sheets: {wb_v.sheetnames})")
        ws_v, ws_f = wb_v[sheet], wb_f[sheet]
        vals = list(ws_v.iter_rows(min_row=r0, max_row=r1, min_col=c0, max_col=c1, values_only=True))
        forms = list(ws_f.iter_rows(min_row=r0, max_row=r1, min_col=c0, max_col=c1, values_only=True))
        cells = []
        for i, (rv, rf) in enumerate(zip(vals, forms)):
            for j, (v, f) in enumerate(zip(rv, rf)):
                if v is None and (f is None or f == ""):
                    continue
                cell = {"cell": f"{_col_letters(c0 + j)}{r0 + i}", "row": r0 + i, "col": c0 + j, "value": _text(v)}
                if isinstance(f, str) and f.startswith("="):
                    cell["formula"] = _text(f)
                    if v is None:
                        cell["note"] = "formula with no saved value (workbook never recalculated/saved in Excel)"
                cells.append(cell)
        return {"sheet": sheet, "range": f"{_col_letters(c0)}{r0}:{_col_letters(c1)}{r1}", "cells": cells,
                "count": len(cells)}
    finally:
        wb_v.close()
        wb_f.close()


def iter_text(path: Path, max_rows: int = INSPECT_ROWS, max_cols: int = INSPECT_COLS):
    """(sheet, cell, text) for every non-empty cell — the search index."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        for ws in wb.worksheets:
            for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), 1):
                for c, v in enumerate(row, 1):
                    if v is None:
                        continue
                    text = str(v).strip()
                    if text:
                        yield ws.title, f"{_col_letters(c)}{r}", text
    finally:
        wb.close()
