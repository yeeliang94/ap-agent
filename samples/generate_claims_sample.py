"""Generate the synthetic employee-claims batch (Client A, LinkedIn-shaped).

No real employee data is ever used: every name, receipt and map below is
invented. Creates, under samples/generated/claims/:

  batch/<Name>_<n>/          one folder per employee, exactly as the client's
                             SharePoint batch folder is laid out:
      <Name>_ER(<period>).xlsx   the expense report workbook (tabs:
                                 Instructions, Expense Types, Expense
                                 Report, KM)
      <Name>_ER(<period>).pdf    a print of the report (to be IGNORED)
      <Name>_Approval.pdf        the e-mail approval (to be IGNORED)
      <Name>_Receipt .pdf        receipt bundles: receipts drawn three to a
      <Name>_Receipt 1.pdf       page in random order; map pages at the back
  Summary of Invoices JUL26.xlsx   the month's payment listing (header row
                                   from the client's own layout, plus two
                                   past tabs holding earlier ER rows)
  demo_claims_batch.zip      the batch folder tree, zipped, for local runs
  ground_truth_claims.json   every row, receipt (file + page + position),
                             trip, expected flag and expected listing row

Planted errors — one per kind — are listed in PLANTED below and recorded in
the ground truth so the verifier can score the run. Everything else is
clean. Run:  python samples/generate_claims_sample.py
Check:       python samples/generate_claims_sample.py --check
"""
from __future__ import annotations

import json
import random
import shutil
import sys
import zipfile
from datetime import date, datetime
from pathlib import Path

from PIL import Image, ImageDraw, ImageFilter, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font

OUT = Path(__file__).resolve().parent / "generated" / "claims"
BATCH = OUT / "batch"
LISTING_FILE = "Summary of Invoices JUL26.xlsx"

# Deterministic "random" order for receipts on pages: the same sample every
# time, so a ground-truth mismatch is a bug and not bad luck.
RNG = random.Random(2026)

# --------------------------------------------------------------- the client
# LinkedIn Malaysia's SAMPLE values. These are the SAMPLE'S facts, not the
# app's: the app discovers or is told them per client (PRD: universal rules
# vs company facts).
CATEGORIES = [  # (expense item, GL code) — the report's Expense Types tab
    ("Taxi", "713070"), ("Mileage", "713080"), ("Parking", "713060"),
    ("Toll", "713065"), ("Mobile Allowance", "713090"), ("Internet", "713095"),
    ("Meals", "711010"), ("Entertainment", "712010"), ("Company Event", "710010"),
    ("Hotel", "714010"), ("Airfare", "714020"), ("Train", "714030"),
    ("Conference & Seminars", "720030"), ("Training", "720010"),
    ("Office Supplies", "730010"), ("Courier", "730020"), ("Printing", "730030"),
    ("Software", "740010"), ("Subscriptions", "740020"), ("Medical", "750010"),
    ("Gifts", "712020"), ("Visa & Permits", "714040"), ("Miscellaneous", "790010"),
]
GL = dict(CATEGORIES)
RATE_CAR, RATE_MOTORCYCLE = 0.64, 0.35
RECEIPT_OPTIONAL = ["Mobile Allowance"]

VENDORS = {
    "Taxi": ["Grab Malaysia", "AirAsia Ride", "Sunlight Taxi"],
    "Parking": ["KLCC Parking", "Menara Prudential Carpark", "Mid Valley Parking"],
    "Toll": ["PLUS Malaysia Berhad", "LDP Toll Plaza"],
    "Meals": ["Restoran Selera", "Kopitiam Kita", "The Ganga Cafe", "Nasi Kandar Pelita"],
    "Entertainment": ["Marini's on 57", "Ruyi & Lyn", "Fuego KL"],
    "Hotel": ["Avillion Port Dickson", "Hotel Jen Singapore"],
    "Airfare": ["Malaysia Airlines"],
    "Conference & Seminars": ["FinTech Asia 2026"],
    "Training": ["Coursera Business"],
    "Office Supplies": ["MPH Bookstores", "Popular Bookstore"],
    "Courier": ["Pos Laju"],
    "Internet": ["TIME Internet"],
    "Company Event": ["Avillion Port Dickson", "PD Waterfront Grill"],
}


def _period(n: int) -> tuple[date, date, str]:
    """Each employee gets a DIFFERENT ER period, so ER codes are unique
    (the map audit insists on one code per employee, none shared)."""
    start = date(2026, 7, 1)
    end = date(2026, 7, 20 + n)
    code = f"ER({start:%d%b%y}-{end:%d%b%y})".upper()
    return start, end, code


def d(day: int) -> date:
    return date(2026, 7, day)


# --------------------------------------------------------------- the people
# Row = dict(date, item, reason, receipt("Y"/"N"), amount, currency, rate,
#            total, **planted). Planted keys:
#   no_receipt=True         the receipt is simply not in any bundle
#   overstated=10.0         the receipt shows amount - 10 (the RM 10 test)
#   shares_receipt=<idx>    this row's receipt is the SAME as row <idx>'s
#   blur=True               the receipt's amount is smudged (low confidence)
# Trip = dict(date, frm, to, purpose, vehicle, km, rate, amount, **planted):
#   map_km=<n>              the map page prints a different km
#   return_trip=True        narrative says "and back"; map prints km/2
#   no_map=True             no map page for this trip
#   no_report_line=True     the KM row has no matching report line
EMPLOYEES: list[dict] = []


def _row(day, item, reason, receipt="Y", amount=0.0, currency="MYR", rate=1.0,
         total=None, **planted):
    return dict(date=d(day), item=item, reason=reason, receipt=receipt,
                amount=round(amount, 2), currency=currency, rate=rate,
                total=round(amount * rate, 2) if total is None else total,
                **planted)


def _trip(day, frm, to, purpose, vehicle, km, rate=None, **planted):
    rate = RATE_CAR if rate is None else rate
    return dict(date=d(day), frm=frm, to=to, purpose=purpose, vehicle=vehicle,
                km=km, rate=rate, amount=round(km * rate, 2), **planted)


def _emp(n, name, purpose, category, rows, trips=(), bundles=2, **extra):
    start, end, code = _period(n)
    EMPLOYEES.append(dict(
        n=n, name=name, folder=f"{name}_{n}", purpose=purpose, category=category,
        gl=GL[category], period=(start, end), er_code=code, rows=rows,
        trips=list(trips), bundles=bundles, department="Sales", **extra))


# 1 — the owner's Copilot test: one taxi row overstated by RM 10.
_emp(1, "Aegene Ong", "Client meetings around KL — July 2026", "Taxi", [
    _row(2, "Taxi", "Grab to client office, KLCC", amount=24.00),
    _row(3, "Taxi", "Grab from client office back to office", amount=26.50),
    _row(7, "Taxi", "Taxi to Bangsar South client", amount=45.00, overstated=10.0),
    _row(9, "Taxi", "Grab to Menara Prudential", amount=18.20),
    _row(15, "Taxi", "Grab to Mid Valley client meeting", amount=31.00),
    _row(9, "Parking", "Parking at Menara Prudential", amount=8.00),
    _row(15, "Parking", "Parking at Mid Valley", amount=6.00),
    _row(1, "Mobile Allowance", "Monthly mobile allowance", receipt="N", amount=100.00),
])

# 2 — mileage: return trip (must NOT flag), km ≠ map, wrong rate.
_emp(2, "Nick Goh", "Client site visits by car — Klang Valley", "Mileage", [
    _row(6, "Toll", "LDP toll to Shah Alam", amount=4.50),
    _row(13, "Toll", "PLUS toll to Nilai", amount=7.20),
], trips=[
    _trip(6, "Home, Petaling Jaya", "Client site, Shah Alam", "Client site visit", "Car", 24.8,
          return_trip=True),                                       # 24.8 km claimed, map 12.4
    _trip(13, "Office, KL Sentral", "Client office, Nilai", "Client site visit", "Car", 20.0,
          map_km=15.2),                                            # km ≠ map
    _trip(17, "Office, KL Sentral", "Client office, Cyberjaya", "Client site visit", "Car", 32.5,
          rate=0.70),                                              # wrong rate
    _trip(21, "Office, KL Sentral", "Client office, Bangsar", "Client site visit", "Motorcycle", 6.1,
          rate=RATE_MOTORCYCLE),                                   # clean
], bundles=1)

# 3 — the mixed-category report whose stated purpose is an offsite.
_emp(3, "Priya Nair", "Team offsite retreat at Port Dickson (Halloween planning offsite)",
     "Company Event", [
    _row(10, "Hotel", "Avillion Port Dickson — 2 nights, team offsite", amount=480.00),
    _row(10, "Meals", "Team dinner, offsite day 1", amount=186.00),
    _row(11, "Meals", "Team lunch, offsite day 2", amount=142.50),
    _row(11, "Company Event", "Team activity — beach games kit", amount=95.00),
    _row(10, "Taxi", "Grab to Port Dickson meeting point", amount=38.00),
    _row(12, "Taxi", "Grab back from KL Sentral", amount=22.00),
])

# 4 — foreign currency: one clean SGD row, one whose arithmetic is wrong.
_emp(4, "Wei Ling Tan", "FinTech Asia 2026 conference — Singapore", "Conference & Seminars", [
    _row(14, "Airfare", "KUL–SIN return, Malaysia Airlines", amount=620.00),
    _row(15, "Conference & Seminars", "FinTech Asia 2026 — 2-day pass", amount=350.00,
         currency="SGD", rate=3.45),                                          # clean: 1207.50
    _row(15, "Meals", "Lunch near conference venue", amount=45.00, currency="SGD",
         rate=3.45, total=160.00),                                            # wrong: 45×3.45=155.25
    _row(16, "Hotel", "Hotel Jen Singapore — 1 night", amount=180.00, currency="SGD", rate=3.45),
    _row(16, "Taxi", "Grab to KLIA", amount=65.00),
])

# 5 — a receipt simply missing; and "receipt = N" on Taxi (must flag).
_emp(5, "Hafiz Rahman", "Client entertainment dinners — July", "Entertainment", [
    _row(8, "Entertainment", "Dinner with client X, 3 pax", amount=312.00),
    _row(16, "Entertainment", "Dinner with client Y, 2 pax", amount=228.50, no_receipt=True),
    _row(22, "Entertainment", "Drinks with client Z", amount=96.00),
    _row(8, "Taxi", "Grab to dinner venue", amount=19.00),
    _row(16, "Taxi", "Grab home after dinner", receipt="N", amount=23.00),
])

# 6 — the same receipt used for two rows.
_emp(6, "Mei Chen", "Office supplies for the team", "Office Supplies", [
    _row(3, "Office Supplies", "Printer paper, 5 reams", amount=64.90),
    _row(3, "Office Supplies", "Printer paper, 5 reams (second batch)", amount=64.90,
         shares_receipt=0),
    _row(20, "Office Supplies", "Whiteboard markers and pens", amount=38.40),
    _row(24, "Courier", "Pos Laju to client", amount=12.00),
])

# 7 — no expense report at all: receipts only.
_emp(7, "Arjun Pillai", "", "Taxi", [
    _row(9, "Taxi", "", amount=28.00),
    _row(14, "Taxi", "", amount=33.50),
    _row(21, "Taxi", "", amount=17.80),
    _row(14, "Parking", "", amount=5.00),
], no_report=True, bundles=1)

# 8 — a smudged receipt amount (low confidence, never a wrong number) and a
# stray file the map cannot place.
_emp(8, "Sofia Lim", "Professional training course — July", "Training", [
    _row(11, "Training", "Coursera Business — data analytics course", amount=890.00, blur=True),
    _row(11, "Meals", "Lunch during course day", amount=27.50),
    _row(18, "Meals", "Lunch during course day", amount=31.00),
], stray_file="notes.txt", bundles=1)

# 9 — mileage without a map, and a KM row that never made it to the report.
_emp(9, "Daniel Wong", "Mileage — client visits, Klang Valley", "Mileage", [
    _row(10, "Parking", "Parking at client, Puchong", amount=5.00),
], trips=[
    _trip(10, "Office, KL Sentral", "Client, Puchong", "Client visit", "Car", 18.3),
    _trip(16, "Office, KL Sentral", "Client, Subang Jaya", "Client visit", "Car", 21.7,
          no_map=True),                                            # MILEAGE_NO_MAP
    _trip(23, "Office, KL Sentral", "Client, Ampang", "Client visit", "Car", 9.9),
    _trip(28, "Office, KL Sentral", "Client, Klang", "Client visit", "Car", 35.0,
          no_report_line=True),                                    # MILEAGE_LINE_MISMATCH
], bundles=1)

# 10 — clean, with a receipt-optional N row and one unclaimed receipt.
_emp(10, "Kavitha Raj", "Monthly mobile allowance and internet", "Mobile Allowance", [
    _row(1, "Mobile Allowance", "Monthly mobile allowance", receipt="N", amount=100.00),
    _row(5, "Internet", "Home fibre — July", amount=139.00),
    _row(19, "Taxi", "Grab to client, Damansara", amount=21.30),
], extra_receipts=[dict(date=d(26), item="Meals", amount=15.90)], bundles=1)

# Which planted problem lives where — the verifier's checklist. Codes are
# the check catalogue's (PRD Flow 3c–3e). "match" locates the row/trip.
PLANTED = [
    {"employee": "Aegene Ong", "code": "NO_RECEIPT", "match": "2026-07-07 Taxi 45.00",
     "why": "row overstated by RM 10 (receipt shows 35.00) — the owner's Copilot test"},
    {"employee": "Aegene Ong", "code": "UNCLAIMED_RECEIPT", "match": "2026-07-07 35.00",
     "why": "the RM 35 receipt supports no row"},
    {"employee": "Nick Goh", "code": "MILEAGE_DISCREPANCY", "match": "2026-07-13 20.0 km",
     "why": "map prints 15.2 km"},
    {"employee": "Nick Goh", "code": "MILEAGE_RATE", "match": "2026-07-17 0.70",
     "why": "rate 0.70 is neither car 0.64 nor motorcycle 0.35"},
    {"employee": "Wei Ling Tan", "code": "CURRENCY_MISMATCH", "match": "2026-07-15 SGD 45.00",
     "why": "45 × 3.45 = 155.25, row total says 160.00"},
    {"employee": "Hafiz Rahman", "code": "NO_RECEIPT", "match": "2026-07-16 Entertainment 228.50",
     "why": "receipt not in any bundle"},
    {"employee": "Hafiz Rahman", "code": "NO_RECEIPT", "match": "2026-07-16 Taxi 23.00",
     "why": "receipt = N on Taxi, which is not receipt-optional"},
    {"employee": "Mei Chen", "code": "DUPLICATE_RECEIPT", "match": "2026-07-03 64.90",
     "why": "two rows, one receipt"},
    {"employee": "Arjun Pillai", "code": "NO_REPORT", "match": "",
     "why": "folder holds receipts and an approval but no report"},
    {"employee": "Daniel Wong", "code": "MILEAGE_NO_MAP", "match": "2026-07-16 21.7 km",
     "why": "no map page for the trip"},
    {"employee": "Daniel Wong", "code": "MILEAGE_LINE_MISMATCH", "match": "2026-07-28 35.0 km",
     "why": "KM row with no report line"},
    {"employee": "Kavitha Raj", "code": "UNCLAIMED_RECEIPT", "match": "2026-07-26 15.90",
     "why": "receipt in the bundle that no row claims (warning, not a blocker)"},
]
MUST_NOT_FLAG = [
    {"employee": "Nick Goh", "what": "2026-07-06 return trip 24.8 km vs map 12.4 km"},
    {"employee": "Aegene Ong", "what": "Mobile Allowance receipt = N (receipt-optional)"},
    {"employee": "Kavitha Raj", "what": "Mobile Allowance receipt = N (receipt-optional)"},
    {"employee": "Wei Ling Tan", "what": "SGD 350 × 3.45 = 1207.50 (clean foreign row)"},
]

# --------------------------------------------------------------- drawing
# Pillow's built-in scalable font: no font files to ship, same look on
# every machine. Sizes are chosen so a receipt is legible after the app's
# 150 dpi render + 1400 px downsize (three receipts share a page).
F = {s: ImageFont.load_default(size=s) for s in (16, 18, 20, 22, 24, 26, 28, 30, 34, 40)}
PAGE = (1654, 2339)  # A4 at 200 dpi
RECEIPT = (500, 780)


def _t(dr: ImageDraw.ImageDraw, xy, text: str, **kw) -> None:
    """Draw text. The built-in font has no glyph for typographic dashes,
    which would render as a missing-character box the AI then reads."""
    dr.text(xy, text.replace("\u2014", "-").replace("\u2013", "-"), **kw)


def _receipt_image(vendor: str, when: date, amount: float, currency: str,
                   item: str, blur: bool = False) -> Image.Image:
    img = Image.new("RGB", RECEIPT, "#fdfcf7")
    dr = ImageDraw.Draw(img)
    _t(dr, (30, 30), vendor, font=F[30], fill="black")
    _t(dr, (30, 80), "Kuala Lumpur, Malaysia" if currency == "MYR" else "Singapore",
            font=F[18], fill="#333")
    _t(dr, (30, 110), "*** OFFICIAL RECEIPT ***", font=F[18], fill="black")
    dr.line((30, 145, 470, 145), fill="black", width=2)
    _t(dr, (30, 165), f"Date : {when:%d/%m/%Y}  {RNG.randint(8, 20):02d}:{RNG.randint(0, 59):02d}",
            font=F[22], fill="black")
    _t(dr, (30, 200), f"Ref  : {RNG.randint(100000, 999999)}", font=F[22], fill="black")
    dr.line((30, 240, 470, 240), fill="black", width=1)
    _t(dr, (30, 260), f"{item}", font=F[22], fill="black")
    _t(dr, (30, 295), f"1 x {currency} {amount:,.2f}", font=F[22], fill="black")
    dr.line((30, 340, 470, 340), fill="black", width=1)
    _t(dr, (30, 370), "TOTAL", font=F[28], fill="black")
    _t(dr, (200, 366), f"{currency} {amount:,.2f}", font=F[34], fill="black")
    _t(dr, (30, 440), "Paid by card", font=F[20], fill="#333")
    _t(dr, (30, 520), "Thank you, please come again", font=F[20], fill="#333")
    _t(dr, (30, 700), "".join(RNG.choice("|| |  ") for _ in range(60)), font=F[24], fill="black")
    if blur:
        # Smudge ONLY the amount region: the digits become genuinely
        # uncertain, so the double read is expected to disagree or to say
        # so — the point is "low confidence", never a confident wrong number.
        box = (190, 355, 470, 410)
        img.paste(img.crop(box).filter(ImageFilter.GaussianBlur(radius=4.5)), box)
    return img.rotate(RNG.uniform(-1.5, 1.5), expand=True, fillcolor="#e9e6df")


def _receipts_page(receipts: list[Image.Image]) -> Image.Image:
    """Three receipts side by side across the upper part of an A4 scan."""
    page = Image.new("RGB", PAGE, "white")
    x = 60
    for img in receipts[:3]:
        page.paste(img, (x, 120))
        x += 520
    _t(ImageDraw.Draw(page), (60, 2280), "scanned by office MFP", font=F[16], fill="#999")
    return page


def _route_image(km: float, mins: int) -> Image.Image:
    """A fake Google-Maps route screenshot with the km in SMALL text —
    the real screenshots print it small, which is why the app re-renders
    map pages at full resolution."""
    img = Image.new("RGB", (720, 400), "#eef2ea")
    dr = ImageDraw.Draw(img)
    for gx in range(0, 720, 60):
        dr.line((gx, 0, gx, 400), fill="#dfe6d8")
    for gy in range(0, 400, 60):
        dr.line((0, gy, 720, gy), fill="#dfe6d8")
    pts = [(80, 320)]
    for i in range(1, 8):
        pts.append((80 + i * 80 + RNG.randint(-15, 15), 320 - i * 35 + RNG.randint(-25, 25)))
    dr.line(pts, fill="#4285f4", width=6)
    dr.ellipse((70, 310, 90, 330), fill="#34a853")
    dr.ellipse((pts[-1][0] - 10, pts[-1][1] - 10, pts[-1][0] + 10, pts[-1][1] + 10), fill="#ea4335")
    _t(dr, (78, 335), "A", font=F[18], fill="black")
    _t(dr, (pts[-1][0] - 4, pts[-1][1] - 32), "B", font=F[18], fill="black")
    dr.rectangle((16, 16, 250, 60), fill="white", outline="#ccc")
    _t(dr, (26, 20), f"{km:.1f} km", font=F[16], fill="#1a73e8")   # deliberately small
    _t(dr, (26, 40), f"{mins} min via highway", font=F[16], fill="#555")
    return img


def _map_page(trips: list[dict]) -> Image.Image:
    """Two trips per page: a narrative line, then the route screenshot."""
    page = Image.new("RGB", PAGE, "white")
    dr = ImageDraw.Draw(page)
    _t(dr, (80, 60), "Mileage claim — supporting maps", font=F[34], fill="black")
    y = 140
    for t in trips:
        back = " and back to home" if t.get("return_trip") else ""
        _t(dr, (80, y), f"{t['date']:%d %b %Y} — {t['purpose']}", font=F[26], fill="black")
        _t(dr, (80, y + 40), f"From: {t['frm']}   To: {t['to']}{back}", font=F[24], fill="black")
        _t(dr, (80, y + 75), f"Vehicle: {t['vehicle']}", font=F[22], fill="#333")
        printed = t.get("map_km", t["km"] / 2 if t.get("return_trip") else t["km"])
        page.paste(_route_image(printed, int(printed * 1.8) + 5), (80, y + 115))
        y += 560
    return page


def _report_print(emp: dict, ws_rows: list[list]) -> Image.Image:
    """A print of the report tab — the file the map must IGNORE."""
    page = Image.new("RGB", PAGE, "white")
    dr = ImageDraw.Draw(page)
    _t(dr, (80, 60), "Expense Report", font=F[40], fill="black")
    y = 130
    for line in ws_rows[:6]:
        _t(dr, (80, y), "   ".join(str(c) for c in line if c is not None), font=F[22], fill="black")
        y += 34
    dr.line((80, y + 10, 1580, y + 10), fill="black", width=2)
    y += 30
    for line in ws_rows[6:]:
        _t(dr, (80, y), "  |  ".join(("" if c is None else str(c))[:22] for c in line), font=F[16], fill="black")
        y += 26
    return page


def _approval(emp: dict) -> Image.Image:
    page = Image.new("RGB", PAGE, "white")
    dr = ImageDraw.Draw(page)
    _t(dr, (80, 80), "From: Manager <manager@example.com>", font=F[24], fill="black")
    _t(dr, (80, 120), f"To: {emp['name']}", font=F[24], fill="black")
    _t(dr, (80, 160), f"Subject: RE: Expense report {emp['er_code']}", font=F[24], fill="black")
    _t(dr, (80, 240), "Approved. Please proceed with submission.", font=F[26], fill="black")
    _t(dr, (80, 300), "Regards,\nManager", font=F[24], fill="black")
    return page


def _pdf(path: Path, pages: list[Image.Image]) -> None:
    pages[0].save(path, save_all=True, append_images=pages[1:], resolution=200.0)


# --------------------------------------------------------------- workbooks
HEADER = ["Date", "Expense Item", "Detailed Business Reason",
          "Itemized Receipt Included (Y/N)", "Total Amount (per receipt)",
          "Currency", "Exchange Rate", "Total (MYR)"]
KM_HEADER = ["Date", "From", "To", "Purpose", "Vehicle", "KM", "Rate (RM/km)", "Amount (MYR)"]


def _report_workbook(emp: dict, path: Path) -> list[list]:
    """The client's report template: Instructions, Expense Types (with GL
    codes), Expense Report (header block + rows + total), KM (trips).
    Values, not formulas: openpyxl computes nothing, and the reader reads
    cached values (a real file saved by Excel carries them)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "How to fill in this expense report"
    ws["A3"] = "1. One line per receipt. Attach itemized receipts."
    ws["A4"] = "2. Choose the Expense Item from the Expense Types tab; the GL code follows it."
    ws["A5"] = "3. Mileage: list every trip on the KM tab and carry each trip to the report."
    ws["A6"] = "4. Foreign currency: type the exchange rate used; Total is in MYR."

    et = wb.create_sheet("Expense Types")
    et.append(["Expense Item", "GL Code"])
    for item, gl in CATEGORIES:
        et.append([f"{item} ({gl})", gl])

    rep = wb.create_sheet("Expense Report")
    start, end = emp["period"]
    head = [
        ["Name:", emp["name"], None, "Department:", emp["department"]],
        ["Period:", f"{start:%d %b %Y} – {end:%d %b %Y}"],
        ["Business Reason for the Report:", emp["purpose"]],
        ["Approved By:", "Manager"],
        [],
        HEADER,
    ]
    for line in head:
        rep.append(line)
    rep["A6"].font = Font(bold=True)
    for cell in rep[6]:
        cell.font = Font(bold=True)
    rows_out = []
    for r in emp["rows"]:
        rows_out.append([datetime.combine(r["date"], datetime.min.time()),
                         f"{r['item']} ({GL[r['item']]})", r["reason"], r["receipt"],
                         r["amount"], r["currency"], r["rate"], r["total"]])
    for t in emp["trips"]:
        if t.get("no_report_line"):
            continue
        rows_out.append([datetime.combine(t["date"], datetime.min.time()),
                         f"Mileage ({GL['Mileage']})",
                         f"{t['purpose']}: {t['frm']} to {t['to']}"
                         + (" and back" if t.get("return_trip") else "")
                         + f" ({t['km']:.1f} km)", "Y", t["amount"], "MYR", 1.0, t["amount"]])
    rows_out.sort(key=lambda x: x[0])
    for line in rows_out:
        rep.append(line)
    total = round(sum(line[7] for line in rows_out), 2)
    rep.append([])
    rep.append([None, None, None, None, None, None, "Total (MYR)", total])
    for cell in rep[rep.max_row]:
        cell.font = Font(bold=True)
    for col, width in zip("ABCDEFGH", (14, 26, 46, 14, 16, 10, 12, 14)):
        rep.column_dimensions[col].width = width
    emp["report_total"] = total
    emp["report_rows"] = [
        {"row": 7 + i, "date": line[0].date().isoformat(), "item": line[1],
         "reason": line[2], "receipt": line[3], "amount": line[4],
         "currency": line[5], "rate": line[6], "total": line[7]}
        for i, line in enumerate(rows_out)]

    km = wb.create_sheet("KM")
    km.append([f"Mileage claim — {emp['name']}"])
    km.append([f"Rates: car RM {RATE_CAR}/km, motorcycle RM {RATE_MOTORCYCLE}/km"])
    km.append([])
    km.append(KM_HEADER)
    for cell in km[4]:
        cell.font = Font(bold=True)
    for i, t in enumerate(emp["trips"]):
        km.append([datetime.combine(t["date"], datetime.min.time()), t["frm"],
                   t["to"] + (" and back" if t.get("return_trip") else ""), t["purpose"],
                   t["vehicle"], t["km"], t["rate"], t["amount"]])
        t["km_row"] = 5 + i
    if emp["trips"]:
        km.append([None, None, None, None, None, None, "Total",
                   round(sum(t["amount"] for t in emp["trips"]), 2)])
    wb.save(path)
    return head + rows_out + [[None] * 6 + ["Total (MYR)", total]]


def _listing_workbook(path: Path) -> dict:
    """The month's Summary of Invoices: the header row in the client's own
    order (this is what the output columns follow), and two past tabs
    holding earlier ER rows so a run has precedent to look at."""
    header = ["S/N", "Processed by", "Received Date", "HelpLine P2P Ref", "PO #",
              "Cost Center", "Category", "GL Account", "Name of Vendor",
              "Invoice Number", "Amount (MYR)", "Remarks"]
    past = {
        "MAY'26": [
            ("Nick Goh", "ER(04MAY26-25MAY26)", "Taxi", 240.00),
            ("Priya Nair", "ER(01MAY26-31MAY26)", "Company Event", 812.40),
            ("Kavitha Raj", "ER(01MAY26-31MAY26)", "Mobile Allowance", 239.00),
        ],
        "JUN'26": [
            ("Aegene Ong", "ER(01JUN26-30JUN26)", "Taxi", 188.70),
            ("Daniel Wong", "ER(02JUN26-29JUN26)", "Mileage", 96.64),
            ("Hafiz Rahman", "ER(01JUN26-30JUN26)", "Entertainment", 540.00),
            ("Wei Ling Tan", "ER(08JUN26-12JUN26)", "Conference & Seminars", 1450.00),
        ],
    }
    wb = Workbook()
    first = True
    for tab, rows in past.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = tab
        ws["A1"] = f"Summary of Invoices for the month of {tab}"
        ws.append([])
        ws.append(header)
        n = 1
        # a few vendor rows first — employees and vendors are mixed in the real file
        for vendor, inv, cat, gl, amt in [
            ("Tenaga Nasional Berhad", "TNB-5100", "Utilities", "760010", 3390.10),
            ("Maxis Bhd", "MX-2214", "Telecoms", "760020", 1240.00),
        ]:
            ws.append([n, "MYTAX", "2026-06-05", "", "", "MY-001", cat, gl, vendor, inv, amt, ""])
            n += 1
        for name, code, cat, amt in rows:
            ws.append([n, "MYTAX", "2026-06-05", "", "", "MY-001", cat, GL[cat], name, code, amt,
                       "employee claim"])
            n += 1
    ws = wb.create_sheet("JUL'26")
    ws["A1"] = "Summary of Invoices for the month of JUL'26"
    ws.append([])
    ws.append(header)
    ws.append([1, "MYTAX", "2026-08-03", "", "", "MY-001", "Utilities", "760010",
               "Tenaga Nasional Berhad", "TNB-5520", 3480.50, ""])
    wb.save(path)
    return {"header": header, "current_tab": "JUL'26", "past_tabs": list(past)}


# --------------------------------------------------------------- assembly
def _build_employee(emp: dict, truth: dict) -> None:
    folder = BATCH / emp["folder"]
    folder.mkdir(parents=True)
    base = f"{emp['name']}_{emp['er_code']}"
    files: dict = {"report": None, "report_print": None, "approval": None,
                   "receipts": [], "stray": None}

    if not emp.get("no_report"):
        ws_rows = _report_workbook(emp, folder / f"{base}.xlsx")
        files["report"] = f"{base}.xlsx"
        _pdf(folder / f"{base}.pdf", [_report_print(emp, ws_rows)])
        files["report_print"] = f"{base}.pdf"
    else:
        emp["report_total"] = None
        emp["report_rows"] = []
    _pdf(folder / f"{emp['name']}_Approval.pdf", [_approval(emp)])
    files["approval"] = f"{emp['name']}_Approval.pdf"
    if emp.get("stray_file"):
        (folder / emp["stray_file"]).write_text("reminder: submit claims by Friday\n")
        files["stray"] = emp["stray_file"]

    # Every receipt that should exist, with the row it belongs to.
    receipts: list[dict] = []
    for i, r in enumerate(emp["rows"]):
        if r["receipt"] != "Y" or r.get("no_receipt") or "shares_receipt" in r:
            continue
        vendor = RNG.choice(VENDORS.get(r["item"], ["Sundry Store"]))
        amount = round(r["amount"] - r.get("overstated", 0.0), 2)
        receipts.append(dict(row=i, vendor=vendor, date=r["date"], amount=amount,
                             currency=r["currency"], item=r["item"], blur=bool(r.get("blur"))))
    for x in emp.get("extra_receipts", []):
        receipts.append(dict(row=None, vendor=RNG.choice(VENDORS[x["item"]]), date=x["date"],
                             amount=x["amount"], currency="MYR", item=x["item"], blur=False))
    RNG.shuffle(receipts)  # random order across pages, as scanned

    # Split across bundle files; map pages go at the back of the LAST bundle.
    n_bundles = max(1, min(emp["bundles"], len(receipts)))
    per = -(-len(receipts) // n_bundles)
    chunks = [receipts[i:i + per] for i in range(0, len(receipts), per)] or [[]]
    trips_with_maps = [t for t in emp["trips"] if not t.get("no_map")]
    truth_receipts = []
    truth_trips = []
    for b, chunk in enumerate(chunks):
        name = f"{emp['name']}_Receipt{' ' if b == 0 else ' ' + str(b)}.pdf"
        pages: list[Image.Image] = []
        for p in range(0, len(chunk), 3):
            group = chunk[p:p + 3]
            imgs = [_receipt_image(g["vendor"], g["date"], g["amount"], g["currency"],
                                   g["item"], g["blur"]) for g in group]
            pages.append(_receipts_page(imgs))
            for pos, g in zip(("left", "middle", "right"), group):
                truth_receipts.append({**g, "date": g["date"].isoformat(), "file": name,
                                       "page": len(pages), "position": pos})
        if b == len(chunks) - 1 and trips_with_maps:
            for p in range(0, len(trips_with_maps), 2):
                group = trips_with_maps[p:p + 2]
                pages.append(_map_page(group))
                for t in group:
                    printed = t.get("map_km", t["km"] / 2 if t.get("return_trip") else t["km"])
                    truth_trips.append({"file": name, "page": len(pages), "date": t["date"].isoformat(),
                                        "km_printed": printed, "return_trip": bool(t.get("return_trip")),
                                        "km_claimed": t["km"], "km_row": t.get("km_row"),
                                        "purpose": t["purpose"]})
        if pages:
            _pdf(folder / name, pages)
            files["receipts"].append(name)

    truth["employees"].append({
        "folder": emp["folder"], "name": emp["name"], "er_code": emp["er_code"],
        "period": [emp["period"][0].isoformat(), emp["period"][1].isoformat()],
        "purpose": emp["purpose"], "category": emp["category"], "gl": emp["gl"],
        "files": files, "report_tab": None if emp.get("no_report") else "Expense Report",
        "mileage_tab": "KM" if emp["trips"] else None,
        "report_total": emp["report_total"], "report_rows": emp["report_rows"],
        "km_rows": [{"row": t.get("km_row"), "date": t["date"].isoformat(), "from": t["frm"],
                     "to": t["to"], "vehicle": t["vehicle"], "km": t["km"], "rate": t["rate"],
                     "amount": t["amount"], "return_trip": bool(t.get("return_trip")),
                     "no_map": bool(t.get("no_map")), "no_report_line": bool(t.get("no_report_line"))}
                    for t in emp["trips"]],
        "receipts": truth_receipts, "map_trips": truth_trips,
        "expected_flags": [p for p in PLANTED if p["employee"] == emp["name"]],
        "must_not_flag": [m for m in MUST_NOT_FLAG if m["employee"] == emp["name"]],
        # What the listing row should say once the reviewer has cleared the
        # flags the intended way (fix the RM 10 row to 35; exclude nothing
        # else). Amount = report total after that fix; for the no-report
        # employee, the sum of the receipts.
        "expected_listing": {
            "name": emp["name"], "er_code": emp["er_code"], "category": emp["category"],
            "gl": emp["gl"],
            "amount": (round(sum(r["amount"] for r in truth_receipts if r["row"] is not None), 2)
                       if emp.get("no_report")
                       else round(emp["report_total"]
                                  - sum(r.get("overstated", 0.0) for r in emp["rows"]), 2)),
        },
    })


def _check() -> int:
    """Assert the sample has the shape the plan promises. Legibility is
    checked by construction — the smallest text sizes, rendered at the
    app's own settings — because there is no OCR engine here."""
    import pymupdf

    truth = json.loads((OUT / "ground_truth_claims.json").read_text())
    problems = []
    folders = sorted(p for p in BATCH.iterdir() if p.is_dir())
    if len(folders) != 10:
        problems.append(f"expected 10 employee folders, found {len(folders)}")
    from openpyxl import load_workbook
    for e in truth["employees"]:
        folder = BATCH / e["folder"]
        if e["files"]["report"]:
            wb = load_workbook(folder / e["files"]["report"], read_only=True)
            if wb.sheetnames != ["Instructions", "Expense Types", "Expense Report", "KM"]:
                problems.append(f"{e['folder']}: tabs are {wb.sheetnames}")
        for name in e["files"]["receipts"]:
            with pymupdf.open(folder / name) as pdf:
                want = max([r["page"] for r in e["receipts"] if r["file"] == name]
                           + [t["page"] for t in e["map_trips"] if t["file"] == name])
                if pdf.page_count != want:
                    problems.append(f"{e['folder']}/{name}: {pdf.page_count} pages, truth says {want}")
    # Receipt text at the app's render: page 2339 px tall at 200 dpi → 150 dpi
    # is 1754 px, downsized to 1400 → scale 1400/2339 ≈ 0.6. The amount is
    # drawn at 34 px → ≈ 20 px, the body at 22 px → ≈ 13 px. Both readable.
    scale = 1400 / PAGE[1]
    if 34 * scale < 18 or 22 * scale < 12:
        problems.append("receipt text too small at 150 dpi")
    # The km on map pages is 16 px at source (≈ 10 px after downsize —
    # deliberately hard); at full resolution (the 300 dpi re-render the app
    # uses for map pages) it is 16 × 1.5 = 24 px. Readable.
    if 16 * 1.5 < 20:
        problems.append("map km text too small at full resolution")
    if len(truth["employees"]) != 10:
        problems.append("ground truth does not list 10 employees")
    if not (OUT / "demo_claims_batch.zip").is_file() or not (OUT / LISTING_FILE).is_file():
        problems.append("zip or listing workbook missing")
    for p in problems:
        print("PROBLEM:", p)
    print("sample check:", "OK" if not problems else f"{len(problems)} problem(s)")
    return 1 if problems else 0


def main() -> None:
    if OUT.exists():
        shutil.rmtree(OUT)
    BATCH.mkdir(parents=True)
    truth: dict = {"client": "Client A (LinkedIn-shaped sample)", "employees": [],
                   "profile": {"mileage_rates": {"Car": RATE_CAR, "Motorcycle": RATE_MOTORCYCLE},
                               "receipt_optional_items": RECEIPT_OPTIONAL,
                               "categories": CATEGORIES,
                               "category_rule": "the report's stated purpose"},
                   "planted": PLANTED, "must_not_flag": MUST_NOT_FLAG}
    for emp in EMPLOYEES:
        _build_employee(emp, truth)
    truth["listing"] = _listing_workbook(OUT / LISTING_FILE)
    truth["listing"]["file"] = LISTING_FILE
    with zipfile.ZipFile(OUT / "demo_claims_batch.zip", "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(BATCH.rglob("*")):
            if f.is_file():
                z.write(f, f.relative_to(BATCH).as_posix())
    (OUT / "ground_truth_claims.json").write_text(json.dumps(truth, indent=2, default=str))
    n_files = sum(1 for f in BATCH.rglob("*") if f.is_file())
    print(f"Generated {len(EMPLOYEES)} employee folders, {n_files} files, "
          f"{len(PLANTED)} planted problems → {OUT}")


if __name__ == "__main__":
    if "--check" in sys.argv:
        sys.exit(_check())
    main()
    sys.exit(_check())
