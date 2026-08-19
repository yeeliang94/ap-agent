"""The immutable run manifest: every file of the snapshot, hashed (H1/H3).

Built once, right after the batch is copied into the workspace and before
anything looks inside a file. From then on a Citation resolves to the
hash captured here (universal control 1, "snapshot integrity"), tools
resolve manifest ids rather than paths the model typed, and nothing that
was uploaded can vanish: an artifact with no disposition is visible until
a reviewer settles it.

Ids are stable across restarts of the same run: 'a' + the first 10 hex
digits of sha256(path + NUL + content hash), so the same file at the same
place gets the same id and two identical files at different places do not
collide.
"""
from __future__ import annotations

import hashlib
from pathlib import Path

from .investigator.contracts import ManifestEntry
from . import source as batch_source
from . import survey as survey_mod

CHUNK = 1024 * 1024


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(CHUNK)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def artifact_id(rel_path: str, sha256: str) -> str:
    return "a" + hashlib.sha256(f"{rel_path}\0{sha256}".encode()).hexdigest()[:10]


def sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True)
    except Exception:
        return []
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def build_manifest(files_dir: Path, files: list[dict]) -> list[ManifestEntry]:
    """One entry per copied file, sorted by path. `files` are the source
    walker's file entries ({path, size, ...}); anything missing from disk
    is still listed (size 0, no hash) so nothing disappears silently."""
    out: list[ManifestEntry] = []
    for entry in sorted(files, key=lambda f: f["path"]):
        rel = entry["path"]
        path = files_dir / rel
        if not path.is_file():
            out.append(ManifestEntry(id=artifact_id(rel, ""), path=rel, size=int(entry.get("size") or 0),
                                     media_type=survey_mod.file_type(rel), snapshot=f"files/{rel}"))
            continue
        digest = sha256_of(path)
        kind = survey_mod.file_type(rel)
        out.append(ManifestEntry(
            id=artifact_id(rel, digest), path=rel, size=path.stat().st_size, sha256=digest,
            media_type=kind, pages=batch_source.page_count(path) if kind in ("pdf", "image") else None,
            sheets=sheet_names(path) if kind == "workbook" else [], snapshot=f"files/{rel}"))
    return out


def lock_snapshot(files_dir: Path) -> int:
    """Make every file of the snapshot read-only once it is inventoried:
    nothing in the run (tools, workers, a careless process) rewrites the
    bytes the manifest hashes name. Returns how many files were locked.
    Best effort per file — a filesystem that refuses is logged, not fatal."""
    import logging
    import stat as stat_mod

    n = 0
    for path in files_dir.rglob("*"):
        if path.is_file():
            try:
                path.chmod(stat_mod.S_IRUSR | stat_mod.S_IRGRP | stat_mod.S_IROTH)
                n += 1
            except OSError as exc:
                logging.getLogger("claims.manifest").warning("could not lock %s: %s", path.name, exc)
    return n


def verify_snapshot(files_dir: Path, manifest: list[dict]) -> list[str]:
    """Re-hash the bytes on disk against the manifest: every entry must be
    present with the recorded hash. Returns the problems, plain sentences."""
    problems: list[str] = []
    for m in manifest:
        rel, want = m.get("path"), m.get("sha256")
        if not rel or not want:
            continue
        path = files_dir / rel
        if not path.is_file():
            problems.append(f"snapshot file {rel!r} is missing from disk")
            continue
        got = sha256_of(path)
        if got != want:
            problems.append(f"snapshot file {rel!r}: bytes on disk hash {got[:12]}…, the manifest recorded {want[:12]}…")
    return problems


def by_path(manifest: list[ManifestEntry]) -> dict[str, ManifestEntry]:
    return {m.path: m for m in manifest}


def by_id(manifest: list[ManifestEntry]) -> dict[str, ManifestEntry]:
    return {m.id: m for m in manifest}


def to_dicts(manifest: list[ManifestEntry]) -> list[dict]:
    return [m.model_dump() for m in manifest]


def from_dicts(items: list[dict]) -> list[ManifestEntry]:
    return [ManifestEntry(**i) for i in items or []]
