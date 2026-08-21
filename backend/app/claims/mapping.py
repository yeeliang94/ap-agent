"""The map: the agent works out which file is what, and code checks it.

The map AI ("judge" role) is shown the survey — every subfolder, every
file with its facts, the tab names and first rows of every workbook, the
page-1 thumbnail of every PDF/image — plus, only if they exist, the
client's playbook, the reviewer's instructions for this run and the last
confirmed map as a worked example. It answers with a proposed claim map:
per subfolder, whether it is an employee, the name and ER code, which
file and tab is the report, which tab holds mileage, which files are
receipt bundles, which to ignore, which it cannot place — EACH WITH A
ONE-LINE REASON, so a wrong guess is obvious at a glance.

Then code audits the guess (audit_map): every folder and file placed; one
ER code per employee and none shared; the tab called "the report" really
yields dated rows with amounts that add up to a total; a file called
"receipts" is a readable PDF or image with at least one page. Whatever
does not fit goes back to the AI with the mismatch, up to MAX_ROUNDS.
Leftovers become map warnings shown beside the map; the reviewer fixes
them on screen before confirming.

The AI reads and reasons; code measures and decides; a person confirms.
"""
from __future__ import annotations

import fnmatch
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Literal

from pydantic import BaseModel, Field
from pydantic_ai import BinaryContent

from ..model_layer import USAGE_LIMITS, create_agent
from . import profile as profile_mod
from . import survey as survey_mod
from .evidence import ai_call

log = logging.getLogger("claims.mapping")

MAX_ROUNDS = 3
MAX_THUMBNAILS = 60

ROLES = ("report", "receipts", "ignore", "unplaced")


class FileRole(BaseModel):
    """One file's role, with the reason the AI gives for it."""
    path: str = Field(max_length=400)
    role: Literal["report", "receipts", "ignore", "unplaced"]
    reason: str = Field(max_length=240)


class FolderMap(BaseModel):
    """One subfolder of the batch."""
    folder: str = Field(max_length=300)
    is_employee: bool
    name: str = Field(default="", max_length=120)
    er_code: str = Field(default="", max_length=60)
    # The workbook holding the expense report, and the tab inside it.
    report_file: str | None = Field(default=None, max_length=400)
    report_tab: str | None = Field(default=None, max_length=100)
    # The tab holding the mileage trips (often "KM"), if any.
    mileage_tab: str | None = Field(default=None, max_length=100)
    # True when the folder holds receipts but no claim list: the worker
    # then builds the rows from the receipts and flags NO_REPORT.
    no_report: bool = False
    files: list[FileRole] = Field(default_factory=list)
    reason: str = Field(max_length=300)


class ClaimMap(BaseModel):
    """The AI's proposed map of the whole batch."""
    employees: list[FolderMap] = Field(default_factory=list)
    # Files sitting directly in the batch folder, not in any subfolder.
    root_files: list[FileRole] = Field(default_factory=list)
    notes: list[str] = Field(default_factory=list, max_length=10)


_INSTRUCTIONS = (
    "You are mapping a folder of employee expense claims so that a verifier "
    "can check each employee's claims. You are given a SURVEY: every "
    "subfolder of the batch folder, every file with its type, size, page "
    "count and any ER(...) code in its name; for each workbook, its tab "
    "names and the first rows of each tab; for each PDF or image, a small "
    "picture of its first page (attached, labelled with the file path).\n"
    "Answer with a map. For EVERY subfolder say whether it is one employee's "
    "claim folder; if so give the employee's name and ER(...) code (from "
    "file names or the report's header), which workbook file AND which tab "
    "inside it is the expense report (the tab with a header block — name, "
    "period, business reason — and dated expense lines with amounts and a "
    "total), which tab holds the mileage trips (km, rate, amount), which "
    "files are receipt bundles (scans of receipts, possibly with map "
    "screenshots for mileage at the back), which files to ignore because "
    "they need no reading (an e-mail approval; a PDF print of the same "
    "report the workbook holds; a template), and which files you cannot "
    "place (role 'unplaced'). Use 'ignore' ONLY for a file you have "
    "positively recognised from its picture or its tabs; a file you could "
    "not look inside (no picture, no tabs — a .txt, .docx, .zip, an "
    "unreadable file) must be 'unplaced', so a person looks at it. Every "
    "file in a folder must be listed once. "
    "If a folder has receipts but no claim list, set no_report=true and say "
    "so — the verifier will build the rows from the receipts.\n"
    "Give a ONE-LINE reason for every role and every folder judgement, in "
    "plain words, quoting what you saw (e.g. \"tab 'Expense Report': name "
    "header, Date / Expense Item / Amount columns and a Total row\"; \"page 1 "
    "shows three till receipts side by side\"; \"subject line 'RE: Expense "
    "report' — an approval e-mail\").\n"
    "Use the client notes and instructions, if given, only for things the "
    "survey cannot show. Never invent files or tabs that are not in the "
    "survey. Copy file paths and tab names EXACTLY as given."
)


async def propose_map(survey: dict, files_dir: Path, snapshot: dict | None = None,
                      instructions: str = "") -> tuple[dict, list[str], list[tuple[str, str]]]:
    """Propose the map, audit it, look again on mismatches.

    Returns (map as a dict, warnings, notes). Notes are (level, text)
    lines for the run diary; warnings are the audit's leftovers, shown
    with the map for the reviewer to fix.
    """
    snapshot = snapshot or {}
    profile = profile_mod.profile_of(snapshot)
    prompt_parts: list = [_prompt_text(survey, snapshot, instructions)]
    prompt_parts += _thumbnails(survey, files_dir)
    agent = create_agent("judge", ClaimMap, _INSTRUCTIONS, temperature=0)
    feedback = ""
    notes: list[tuple[str, str]] = []
    problems: list[str] = []
    claim_map: ClaimMap | None = None
    for round_no in range(1, MAX_ROUNDS + 1):
        parts = [prompt_parts[0] + feedback, *prompt_parts[1:]]
        result = await ai_call(agent.run(parts, usage_limits=USAGE_LIMITS), "the map agent")
        claim_map = result.output
        _apply_role_patterns(claim_map, profile)
        try:
            problems = audit_map(claim_map, survey, files_dir)
        except Exception as exc:
            problems = [f"applying your map failed: {exc}"]
        if not problems:
            notes.append(("INFO", f"Map confirmed by the audit on round {round_no}."))
            break
        notes.append(("WARNING" if round_no == MAX_ROUNDS else "INFO",
                      f"Map round {round_no}: {len(problems)} problem(s) — "
                      + "; ".join(problems)[:500]))
        feedback = (
            "\n\nYour previous map did not pass the audit:\n- " + "\n- ".join(problems)
            + "\nLook at the survey again and correct the map. Keep every "
              "file listed once, copy paths and tab names exactly."
        )
    assert claim_map is not None
    warnings = list(problems)
    out = _complete(claim_map, survey, warnings)
    out["rounds"] = round_no
    return out, warnings, notes


def _prompt_text(survey: dict, snapshot: dict, instructions: str) -> str:
    parts = ["# Survey of the batch folder", survey_mod.survey_text(survey)]
    playbook = (snapshot.get("playbook") or "").strip()
    if playbook:
        parts += ["\n# Client notes (playbook)", playbook]
    if instructions.strip():
        parts += ["\n# Instructions for this run", instructions.strip()]
    patterns = profile_mod.profile_of(snapshot).get("file_role_patterns") or []
    if patterns:
        parts += ["\n# File-role rules already confirmed for this client"]
        parts += [f"- files matching {p['pattern']!r} → {p['role']}" for p in patterns]
    last = (snapshot.get("last_map") or {}).get("map")
    if last:
        parts += ["\n# The last confirmed map for this client (a worked example — "
                  "the folder names and files may differ this time)"]
        for e in last.get("employees", [])[:30]:
            parts.append(f"- {e.get('folder')}: employee={e.get('is_employee')} "
                         f"name={e.get('name')!r} report={e.get('report_file')!r} "
                         f"tab={e.get('report_tab')!r} mileage_tab={e.get('mileage_tab')!r} "
                         f"receipts={[f['path'] for f in e.get('files', []) if f['role'] == 'receipts']} "
                         f"ignored={[f['path'] for f in e.get('files', []) if f['role'] == 'ignore']}")
    return "\n".join(parts)


def _thumbnails(survey: dict, files_dir: Path) -> list:
    """The page-1 pictures, each preceded by its label. Capped so a huge
    batch cannot turn one call into hundreds of images."""
    parts: list = []
    n = 0
    for f in survey["files"]:
        thumb = (f.get("peek") or {}).get("thumbnail")
        if not thumb:
            continue
        path = files_dir.parent / thumb
        if not path.is_file():
            continue
        n += 1
        if n > MAX_THUMBNAILS:
            parts.append(f"(thumbnail of {f['path']} omitted — over {MAX_THUMBNAILS} pictures)")
            continue
        parts.append(f"Page 1 of `{f['path']}`:")
        parts.append(BinaryContent(data=path.read_bytes(), media_type="image/png"))
    return parts


def _apply_role_patterns(claim_map: ClaimMap, profile: dict) -> None:
    """Confirmed file-role patterns from the client profile win over the
    AI's guess — that is what "remember for <client>" promised."""
    patterns = profile.get("file_role_patterns") or []
    if not patterns:
        return
    for folder in claim_map.employees:
        for fr in folder.files:
            for p in patterns:
                if p.get("role") in ROLES and fnmatch.fnmatch(Path(fr.path).name, p["pattern"]):
                    if fr.role != p["role"]:
                        fr.role = p["role"]
                        fr.reason = f"client profile rule: {p['pattern']} → {p['role']}"
                    if p["role"] != "report" and folder.report_file == fr.path:
                        folder.report_file = None
                        folder.report_tab = None


# ---- the audit ---------------------------------------------------------

def audit_map(claim_map: ClaimMap, survey: dict, files_dir: Path) -> list[str]:
    """Everything code can check about the AI's map. Returns problems in
    the AI's own terms (paths, tabs), empty when the map holds up."""
    problems: list[str] = []
    survey_folders = {fo["path"] for fo in survey["folders"]}
    files_by_folder: dict[str, set[str]] = {}
    for f in survey["files"]:
        files_by_folder.setdefault(f["folder"], set()).add(f["path"])
    by_path = {f["path"]: f for f in survey["files"]}

    mapped_folders = [e.folder for e in claim_map.employees]
    for folder in survey_folders - set(mapped_folders):
        problems.append(f"subfolder {folder!r} is missing from your map")
    for folder in set(mapped_folders) - survey_folders:
        problems.append(f"subfolder {folder!r} is not in the survey")
    seen: set[str] = set()
    for folder in mapped_folders:
        if folder in seen:
            problems.append(f"subfolder {folder!r} appears twice in your map")
        seen.add(folder)

    codes: dict[str, str] = {}
    for e in claim_map.employees:
        expected = files_by_folder.get(e.folder, set())
        listed = [fr.path for fr in e.files]
        for path in expected - set(listed):
            problems.append(f"file {path!r} in {e.folder!r} has no role")
        for path in set(listed) - expected:
            problems.append(f"file {path!r} is not in folder {e.folder!r} of the survey")
        dup = {p for p in listed if listed.count(p) > 1}
        for path in dup:
            problems.append(f"file {path!r} is listed more than once")
        if not e.is_employee:
            continue
        if not e.name.strip():
            problems.append(f"{e.folder!r}: employee name is empty")
        if e.er_code:
            # The code must be SEEN somewhere — a file name or the peek of
            # a workbook in that folder — never inferred from the pattern
            # of the other folders.
            seen_in = " ".join(files_by_folder.get(e.folder, set()))
            for path in files_by_folder.get(e.folder, set()):
                for rows in ((by_path.get(path, {}).get("peek") or {}).get("tabs") or {}).values():
                    seen_in += " " + " ".join(rows)
            if e.er_code.replace(" ", "").upper() not in seen_in.replace(" ", "").upper():
                problems.append(f"{e.folder!r}: ER code {e.er_code!r} does not appear in any file name "
                                "or workbook of that folder — copy it exactly from where you saw it, "
                                "or leave er_code empty")
            if e.er_code in codes and codes[e.er_code] != e.folder:
                problems.append(f"ER code {e.er_code!r} is used by both {codes[e.er_code]!r} "
                                f"and {e.folder!r} — one code per employee")
            codes.setdefault(e.er_code, e.folder)
        receipt_files = [fr for fr in e.files if fr.role == "receipts"]
        report_roles = [fr for fr in e.files if fr.role == "report"]
        if e.no_report:
            if e.report_file or report_roles:
                problems.append(f"{e.folder!r}: no_report is true but a report file is named")
            if not receipt_files:
                problems.append(f"{e.folder!r}: no report and no receipt files — nothing to verify")
        else:
            if not e.report_file:
                problems.append(f"{e.folder!r}: name the report_file and report_tab, or set "
                                "no_report=true if the folder holds no claim list")
            else:
                if e.report_file not in expected:
                    problems.append(f"{e.folder!r}: report_file {e.report_file!r} is not in that folder")
                elif not any(fr.path == e.report_file and fr.role == "report" for fr in e.files):
                    problems.append(f"{e.folder!r}: report_file {e.report_file!r} must also be "
                                    "listed with role 'report'")
                else:
                    problems += _check_report_tab(files_dir, e, by_path.get(e.report_file))
            if not receipt_files and not e.report_file:
                problems.append(f"{e.folder!r}: an employee needs a report or receipts")
        for fr in receipt_files:
            info = by_path.get(fr.path)
            if info and info["type"] not in ("pdf", "image"):
                problems.append(f"{fr.path!r}: a receipts file must be a PDF or an image, "
                                f"this is a {info['type']}")
            elif info and not info.get("pages"):
                problems.append(f"{fr.path!r}: no readable page — cannot be a receipts file")
        # More than one report per folder is a real ambiguity, not a typo.
        if len(report_roles) > 1:
            problems.append(f"{e.folder!r}: {len(report_roles)} files marked 'report'; "
                            "exactly one is the report — mark the rest ignore or unplaced")
    root_expected = set(survey.get("root_files") or [])
    root_listed = {fr.path for fr in claim_map.root_files}
    for path in root_expected - root_listed:
        problems.append(f"file {path!r} in the batch root has no role")
    return problems


def _check_report_tab(files_dir: Path, e: FolderMap, info: dict | None) -> list[str]:
    """Does the named tab look like a report: dated rows with amounts that
    add up to a total? Cheap and structural — the full read is the
    worker's job."""
    if info is None or info["type"] != "workbook":
        return [f"{e.folder!r}: report_file {e.report_file!r} is not a workbook"]
    tabs = list(((info.get("peek") or {}).get("tabs") or {}).keys())
    if not e.report_tab:
        return [f"{e.folder!r}: name the report_tab (tabs: {tabs})"]
    if tabs and e.report_tab not in tabs:
        return [f"{e.folder!r}: tab {e.report_tab!r} is not in {e.report_file!r} (tabs: {tabs})"]
    if e.mileage_tab and tabs and e.mileage_tab not in tabs:
        return [f"{e.folder!r}: mileage tab {e.mileage_tab!r} is not in {e.report_file!r} (tabs: {tabs})"]
    ok, why = report_tab_plausible(files_dir / e.report_file, e.report_tab)
    if not ok:
        return [f"{e.folder!r}: tab {e.report_tab!r} of {e.report_file!r} does not look like "
                f"an expense report: {why}"]
    return []


def report_tab_plausible(path: Path, tab: str) -> tuple[bool, str]:
    """True when the tab has dated rows with amounts, and some cell holds
    the sum of one of those amount columns (the total)."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return False, f"cannot open workbook ({type(exc).__name__})"
    try:
        if tab not in wb.sheetnames:
            return False, f"no tab {tab!r}"
        ws = wb[tab]
        dated_rows: list[list] = []
        all_numbers: list[float] = []
        for row in ws.iter_rows(min_row=1, max_row=400, max_col=30, values_only=True):
            has_date = any(isinstance(v, (date, datetime)) for v in row)
            nums = [(i, float(v)) for i, v in enumerate(row)
                    if isinstance(v, (int, float)) and not isinstance(v, bool)]
            all_numbers += [n for _, n in nums]
            if has_date and nums:
                dated_rows.append(nums)
        if len(dated_rows) < 1:
            return False, "no rows holding both a date and a number"
        cols: dict[int, float] = {}
        for nums in dated_rows:
            for i, n in nums:
                cols[i] = cols.get(i, 0.0) + n
        for total in cols.values():
            if any(abs(total - n) < 0.02 for n in all_numbers):
                return True, ""
        return False, ("the dated rows' amounts do not add up to any total on the tab "
                       "(a mileage or working tab, not the report?)")
    finally:
        wb.close()


def _complete(claim_map: ClaimMap, survey: dict, warnings: list[str]) -> dict:
    """The map as stored: the AI's answer with every survey folder and file
    present (missing ones added as unplaced, so nothing vanishes silently)."""
    out = claim_map.model_dump()
    files_by_folder: dict[str, list[str]] = {}
    for f in survey["files"]:
        files_by_folder.setdefault(f["folder"], []).append(f["path"])
    have = {e["folder"] for e in out["employees"]}
    for fo in survey["folders"]:
        if fo["path"] not in have:
            out["employees"].append({
                "folder": fo["path"], "is_employee": False, "name": "", "er_code": "",
                "report_file": None, "report_tab": None, "mileage_tab": None,
                "no_report": False, "files": [],
                "reason": "the AI did not place this folder — check it"})
    for e in out["employees"]:
        listed = {fr["path"] for fr in e["files"]}
        for path in files_by_folder.get(e["folder"], []):
            if path not in listed:
                e["files"].append({"path": path, "role": "unplaced",
                                   "reason": "not placed by the AI"})
        e["files"] = [fr for fr in e["files"] if fr["path"] in set(files_by_folder.get(e["folder"], []))]
    root_listed = {fr["path"] for fr in out["root_files"]}
    for path in survey.get("root_files") or []:
        if path not in root_listed:
            out["root_files"].append({"path": path, "role": "unplaced",
                                      "reason": "not placed by the AI"})
    return out


# ---- validation of a reviewer-edited map (before confirm) ---------------

def _folder_of(path: str) -> str:
    return path.split("/", 1)[0] if "/" in path else ""


def validate_confirmed_map(claim_map: dict, survey: dict) -> list[str]:
    """What must hold before Confirm & verify: every employee has a report
    (a workbook IN THEIR OWN FOLDER, with a tab) or at least one receipts
    file, or is skipped; every file an employee lists is in their folder;
    every surveyed file is placed exactly once (nothing vanishes, nothing
    is read for two people); no shared ER codes. Returns plain-language
    problems."""
    problems: list[str] = []
    by_path = {f["path"]: f for f in survey["files"]}
    codes: dict[str, str] = {}
    placed: dict[str, list[str]] = {}  # file -> who lists it
    for e in claim_map.get("employees", []):
        for fr in e.get("files", []):
            placed.setdefault(fr["path"], []).append(e.get("name") or e["folder"])
    for fr in claim_map.get("root_files", []) or []:
        placed.setdefault(fr["path"], []).append("root")
    for path in sorted(by_path):
        who = placed.get(path, [])
        if not who:
            problems.append(f"{path}: this file is in the batch but nowhere on the map — every file "
                            "must be placed (as a report, receipts, or ignore)")
        elif len(who) > 1:
            problems.append(f"{path}: listed under {' and '.join(who)} — a file belongs to one place")
    for path in placed:
        if path not in by_path:
            problems.append(f"{path}: this file is on the map but not in the batch")
    for e in claim_map.get("employees", []):
        if not e.get("is_employee") or e.get("skip"):
            continue
        label = e.get("name") or e["folder"]
        roles = {fr["path"]: fr["role"] for fr in e.get("files", [])}
        for path in roles:
            if _folder_of(path) != e["folder"]:
                problems.append(f"{label}: {path} is not in the folder {e['folder']!r} — "
                                "a file is read only for the employee whose folder holds it")
        receipts = [p for p, r in roles.items() if r == "receipts"]
        if e.get("no_report"):
            if not receipts:
                problems.append(f"{label}: marked 'no report' but has no receipt files — "
                                "add a receipts file or skip this employee")
        else:
            rf = e.get("report_file")
            if not rf:
                problems.append(f"{label}: choose the report file, or mark 'no report'")
            elif rf not in by_path or by_path[rf]["type"] != "workbook":
                problems.append(f"{label}: the report file must be a workbook in the folder")
            elif _folder_of(rf) != e["folder"]:
                problems.append(f"{label}: the report file {rf} is not in the folder {e['folder']!r}")
            elif roles.get(rf) != "report":
                problems.append(f"{label}: {rf} is the report file, so its role must be 'report' "
                                f"(it is {roles.get(rf) or 'not listed'})")
            elif not e.get("report_tab"):
                problems.append(f"{label}: choose the report tab")
            else:
                tabs = list(((by_path[rf].get("peek") or {}).get("tabs") or {}).keys())
                if tabs and e["report_tab"] not in tabs:
                    problems.append(f"{label}: tab {e['report_tab']!r} is not in the report file")
                if e.get("mileage_tab") and tabs and e["mileage_tab"] not in tabs:
                    problems.append(f"{label}: mileage tab {e['mileage_tab']!r} is not in the report file")
        code = (e.get("er_code") or "").strip()
        if code:
            if code in codes:
                problems.append(f"{label}: ER code {code} is also used by {codes[code]}")
            codes[code] = label
        if not (e.get("name") or "").strip():
            problems.append(f"{e['folder']}: the employee needs a name")
    # A batch with files but no case would "verify" nothing and still end
    # green — the do-nothing run must be refused here, not discovered later.
    if by_path and not any(e.get("is_employee") and not e.get("skip")
                           for e in claim_map.get("employees", [])):
        problems.append("this map creates no cases — nothing would be verified; mark at "
                        "least one folder as an employee, or cancel the run if this batch "
                        "truly holds no claims")
    return problems
