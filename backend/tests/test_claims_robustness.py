"""Phase 4b: robustness of the checks + the flag catalogue.

R1  every flag code the code can raise has words in the catalogue (a new
    flag without a title / meaning / what-to-do fails here, not on screen)
R2  a report whose lines do not add up to its total cell is KEPT and
    flagged (REPORT_TOTAL_MISMATCH), not thrown away; a workbook whose
    formulas were never computed is named as such
"""
from __future__ import annotations

import asyncio
import re
from datetime import date
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.claims import profile, report_reader, worker
from app.claims.report_reader import ReportColumns, ReportReading
from app.main import app

CLAIMS = Path(__file__).resolve().parents[1] / "app" / "claims"


class _Scripted:
    """An 'agent' that answers from a list, in order (like test_claims_checks)."""

    def __init__(self, outputs):
        self._outputs = list(outputs)
        self.prompts = []

    async def run(self, prompt, **kw):
        self.prompts.append(prompt)

        class R:
            output = self._outputs.pop(0)

            def usage(self):
                class U:
                    total_tokens = 10
                    requests = 1
                return U()
        return R()


def _report_sheet(lines, total):
    """A tiny expense report tab: header block, headings on row 6, lines
    from row 7, the typed total two rows under the last line."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    ws["A1"], ws["B1"] = "Name:", "Aegene Ong"
    ws["A2"], ws["B2"] = "Period:", "01JUL26-21JUL26"
    ws["A3"], ws["B3"] = "Purpose:", "Client visits"
    for c, h in zip("ABCDEFGH", ["Date", "Expense Item", "Reason", "Receipt?", "Amount", "Curr", "Rate", "Total (MYR)"]):
        ws[f"{c}6"] = h
    r = 7
    for d, item, amount in lines:
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"] = d, item, "x", "Y"
        ws[f"E{r}"], ws[f"F{r}"], ws[f"G{r}"], ws[f"H{r}"] = amount, "MYR", 1, amount
        r += 1
    ws[f"G{r + 1}"], ws[f"H{r + 1}"] = "Total (MYR)", total
    return wb, ws, r - 1, r + 1


def _reading(last_row, total_row, **over):
    base = dict(columns=ReportColumns(date="A", item="B", reason="C", receipt_included="D", amount="E",
                                      currency="F", rate="G", total="H"),
                header_row=6, first_row=7, last_row=last_row, total_cell=f"H{total_row}", name_cell="B1",
                period_cell="B2", purpose_cell="B3", why="header block, dated lines, total")
    return ReportReading(**{**base, **over})


def _codes_in_source() -> set[str]:
    """Every literal flag code in the claims package: `_flag("CODE"`,
    `code="CODE"`, or `checks_mod._flag("CODE"`."""
    found: set[str] = set()
    for path in CLAIMS.glob("*.py"):
        text = path.read_text()
        found |= set(re.findall(r'_flag\(\s*"([A-Z][A-Z_]+)"', text))
        found |= set(re.findall(r'code="([A-Z][A-Z_]+)"', text))
    return found


# ---- R1: the catalogue ------------------------------------------------------------

def test_every_raised_code_is_catalogued():
    raised = _codes_in_source()
    assert raised, "the scan found no codes — the regex is broken"
    missing = sorted(raised - set(profile.CATALOGUE))
    assert not missing, f"flag codes with no catalogue words: {missing}"


def test_catalogue_entries_are_complete_and_shaped():
    for code, entry in profile.CATALOGUE.items():
        for key in ("title", "meaning", "what_to_do", "kind", "blocks"):
            assert entry.get(key), f"{code}: {key} missing"
        assert entry["kind"] in profile.FLAG_KINDS, code
        assert entry["blocks"] in ("open", "info"), code
        assert "_" not in entry["title"], f"{code}: title looks like a code"
    # run-level controls cannot be switched off; everything else can
    assert "MISSING_REFERENCE" not in profile.CHECK_CODES
    assert "NO_RECEIPT" in profile.CHECK_CODES
    assert profile.describe("NOT_A_CODE")["title"] == "Not a code"


def test_catalogue_endpoint_and_run_detail_carry_it(monkeypatch, tmp_path):
    client = TestClient(app)
    r = client.get("/api/claims-settings/catalogue")
    assert r.status_code == 200
    body = r.json()
    assert set(body["codes"]) == set(profile.CATALOGUE)
    assert body["codes"]["NO_RECEIPT"]["title"] == "No receipt for this row"
    assert body["kinds"] == list(profile.FLAG_KINDS)
    assert "MISSING_REFERENCE" not in body["toggleable"]


# ---- R2: a wrong total is a flag, not a lost report ---------------------------

LINES = [(date(2026, 7, 1), "Taxi (713070)", 45.00), (date(2026, 7, 3), "Coffee (713070)", 18.90),
         (date(2026, 7, 5), "Toll (713070)", 10.00)]  # sum 73.90


def test_total_typo_keeps_the_lines_and_marks_the_mismatch(monkeypatch):
    wb, ws, last, total_row = _report_sheet(LINES, 74.00)  # typed 10 cents high
    agent = _Scripted([_reading(last, total_row), _reading(last, total_row)])
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    rows, header, notes = asyncio.run(report_reader.read_report(ws, "Aegene Ong", "ER(01JUL26-21JUL26)"))
    # fed back once (a missed line is the other cause), then the same
    # reading again is accepted — the lines are kept
    assert len(agent.prompts) == 2 and "answer the same reading again" in agent.prompts[1]
    assert len(rows) == 3
    assert header["total"] == "74.00" and header["total_cell"] == f"H{total_row}"
    assert header["total_check"] == {"lines": "73.90", "cell": "74.00", "column": "total"}
    assert any("same reading twice" in t for _, t in notes)
    # the worker turns it into a flag citing the total cell's row
    flags = worker._report_total_flags(header, ws.title)
    assert len(flags) == 1 and flags[0]["code"] == "REPORT_TOTAL_MISMATCH"
    assert flags[0]["cite"] == {"sheet": "Expense Report", "row": total_row}
    assert "73.90" in flags[0]["reason"] and "74.00" in flags[0]["reason"]


def test_a_correct_total_raises_nothing_and_a_moved_span_is_still_audited(monkeypatch):
    wb, ws, last, total_row = _report_sheet(LINES, 73.90)
    agent = _Scripted([_reading(last, total_row)])
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    rows, header, _ = asyncio.run(report_reader.read_report(ws, "Aegene Ong", "ER(01JUL26-21JUL26)"))
    assert len(rows) == 3 and header["total_check"] is None and len(agent.prompts) == 1
    assert worker._report_total_flags(header, ws.title) == []
    # a span that misses the last line: fed back; the AI extends it; accepted with no mismatch
    agent = _Scripted([_reading(last - 1, total_row), _reading(last, total_row)])
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    rows, header, _ = asyncio.run(report_reader.read_report(ws, "Aegene Ong", "ER(01JUL26-21JUL26)"))
    assert len(rows) == 3 and header["total_check"] is None and len(agent.prompts) == 2


def test_total_mismatch_with_other_structural_problems_stays_structural(monkeypatch):
    # a reading whose amount column is wrong AND whose sum is off: three
    # rounds of the same wrong answer → unreadable, not "accepted with a note"
    wb, ws, last, total_row = _report_sheet(LINES, 74.00)
    wrong = _reading(last, total_row, columns=ReportColumns(date="A", item="B", amount="C", total="H"))
    agent = _Scripted([wrong, wrong, wrong])
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    with pytest.raises(report_reader.ReportUnreadable):
        asyncio.run(report_reader.read_report(ws, "Aegene Ong", "ER(01JUL26-21JUL26)"))


def test_uncomputed_formulas_are_named(tmp_path):
    wb, ws, last, total_row = _report_sheet(LINES, 0)
    ws[f"H{total_row}"] = f"=SUM(H7:H{last})"  # openpyxl saves the formula, no cached value
    path = tmp_path / "r.xlsx"
    wb.save(path)
    assert report_reader.uncomputed_formulas(path, "Expense Report") == 1
    assert report_reader.uncomputed_formulas(path, "No such tab") == 0
    # a values-only workbook has none
    wb2, _, _, _ = _report_sheet(LINES, 73.90)
    wb2.save(tmp_path / "v.xlsx")
    assert report_reader.uncomputed_formulas(tmp_path / "v.xlsx", "Expense Report") == 0
    # and data_only really does read the formula cell as empty (the reason this exists)
    assert load_workbook(path, data_only=True)["Expense Report"][f"H{total_row}"].value is None


def test_a_missed_line_is_told_apart_from_a_typo(monkeypatch):
    # the same three lines, correct total, but the reading drops the last
    # line: that is a short span, not a typo — structural, never "accepted
    # with a mismatch note", even when the AI insists three times
    wb, ws, last, total_row = _report_sheet(LINES, 73.90)
    short = _reading(last - 1, total_row)
    agent = _Scripted([short, short, short])
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    with pytest.raises(report_reader.ReportUnreadable) as exc:
        asyncio.run(report_reader.read_report(ws, "Aegene Ong", "ER(01JUL26-21JUL26)"))
    assert "outside" in str(exc.value)


# ---- R3: rows the reader may skip ------------------------------------------------

def test_subtotal_row_inside_the_span_can_be_skipped(monkeypatch):
    wb, ws, last, total_row = _report_sheet(LINES, 73.90)
    # push a "Meals subtotal" row in between lines 2 and 3
    ws.insert_rows(9)
    ws["B9"], ws["H9"] = "Meals subtotal", 63.90  # no date: a subtotal, not a line
    last, total_row = last + 1, total_row + 1
    ws[f"H{total_row}"] = 73.90
    naive = _reading(last, total_row)                      # sum includes the subtotal → 137.80
    with_skip = _reading(last, total_row, skip_rows=[9])   # subtotal left out → 73.90
    agent = _Scripted([naive, with_skip])
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    rows, header, _ = asyncio.run(report_reader.read_report(ws, "Aegene Ong", "ER(01JUL26-21JUL26)"))
    assert len(agent.prompts) == 2
    assert [r["row"] for r in rows] == [7, 8, 10] and header["total_check"] is None
    # skipping a REAL line is refused
    bad = _reading(last, total_row, skip_rows=[9, 10])
    agent = _Scripted([bad, bad, bad])
    monkeypatch.setattr(report_reader, "create_agent", lambda *a, **k: agent)
    with pytest.raises(report_reader.ReportUnreadable) as exc:
        asyncio.run(report_reader.read_report(ws, "Aegene Ong", "ER(01JUL26-21JUL26)"))
    assert "row 10 is in skip_rows" in str(exc.value)


# ---- R4: the same receipt on two pages -----------------------------------------------

from app.claims import checks  # noqa: E402

PROFILE = {"mileage_rates": {"Car": "0.64"}, "km_tolerance": "0", "receipt_date_window_days": 0,
           "receipt_optional_items": [], "mileage_item_pattern": "mileage", "categories": [],
           "category_rule": "", "file_role_patterns": [], "checks": {}, "set_by": {}}


def _row(rid, date_, amount, item="Taxi"):
    return {"id": rid, "kind": "expense", "sheet": "S", "row": int(rid[1:]) + 6,
            "values": {"date": date_, "item": item, "item_name": item, "reason": "x", "receipt_included": "Y",
                       "amount": amount, "currency": "MYR", "rate": "1", "total": amount}}


def _receipt(eid, date_, amount, page=1, vendor="Grab", position="left"):
    return {"id": eid, "kind": "receipt", "file": "r.pdf", "page": page, "position": position,
            "values": {"vendor": vendor, "date": date_, "amount": amount, "currency": "MYR"}, "confidence": {}}


def _checks(rows, ev, profile=PROFILE):
    calls = {"n": 0}

    async def tie(row, cs):
        calls["n"] += 1
        return ""
    res = asyncio.run(checks.run_checks(rows, ev, profile, {}, (1, 1), tie_break=tie))
    res["tie_calls"] = calls["n"]
    return res


def test_double_claim_on_a_twice_scanned_receipt_is_flagged():
    rows = [_row("r1", "2026-07-01", "45.00"), _row("r2", "2026-07-01", "45.00")]
    ev = [_receipt("e1", "2026-07-01", "45.00", page=2), _receipt("e2", "2026-07-01", "45.00", page=5)]
    res = _checks(rows, ev)
    assert res["tie_calls"] == 0  # identical candidates never go to the AI
    assert res["verdicts"]["r1"][0] == "matched" and res["verdicts"]["r2"][0] == "matched"
    dup = [f for f in res["flags"] if f["code"] == "DUPLICATE_SCAN"]
    assert {f["row_id"] for f in dup} == {"r1", "r2"}
    assert all(f["status"] == "open" and "page 2" in f["reason"] and "page 5" in f["reason"] for f in dup)


def test_one_row_and_a_duplicate_page_is_just_a_note():
    rows = [_row("r1", "2026-07-01", "45.00")]
    ev = [_receipt("e1", "2026-07-01", "45.00", page=2), _receipt("e2", "2026-07-01", "45.00", page=5)]
    res = _checks(rows, ev)
    assert res["verdicts"]["r1"][0] == "matched"
    assert not [f for f in res["flags"] if f["code"] == "DUPLICATE_SCAN"]
    assert len([f for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"]) == 1


def test_two_real_receipts_with_different_vendors_are_untouched():
    rows = [_row("r1", "2026-07-01", "10.00"), _row("r2", "2026-07-01", "10.00")]
    ev = [_receipt("e1", "2026-07-01", "10.00", page=1, vendor="PLUS toll"),
          _receipt("e2", "2026-07-01", "10.00", page=1, vendor="LDP toll", position="right")]
    res = _checks(rows, ev)
    assert res["tie_calls"] >= 1  # different vendors: the AI is asked
    assert not [f for f in res["flags"] if f["code"] == "DUPLICATE_SCAN"]
    # the check can be switched off per client
    rows2 = [_row("r1", "2026-07-01", "45.00"), _row("r2", "2026-07-01", "45.00")]
    ev2 = [_receipt("e1", "2026-07-01", "45.00", page=2), _receipt("e2", "2026-07-01", "45.00", page=5)]
    res = _checks(rows2, ev2, {**PROFILE, "checks": {"DUPLICATE_SCAN": False}})
    assert not [f for f in res["flags"] if f["code"] == "DUPLICATE_SCAN"]


# ---- R5: one receipt, two employees ------------------------------------------------

from sqlalchemy import create_engine  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402

from app.claims.models import ClaimEmployee, ClaimEvidence, ClaimFlag, ClaimRow, ClaimsRun  # noqa: E402
from app.db import Base  # noqa: E402


@pytest.fixture()
def Session(tmp_path, monkeypatch):
    engine = create_engine(f"sqlite:///{tmp_path / 't.sqlite3'}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    S = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    monkeypatch.setattr(worker, "SessionLocal", S)
    return S


def _employee_with_matched_receipt(s, run, folder, name, vendor="Sushi King", page=1):
    e = ClaimEmployee(run_id=run.id, folder=folder, name=name, er_code="", roles={}, status="verified")
    s.add(e)
    s.commit()
    row = ClaimRow(run_id=run.id, employee_id=e.id, kind="expense", sheet="S", row=7,
                   values={"date": "2026-07-01", "amount": "120.00", "total": "120.00", "currency": "MYR"},
                   verdict="matched")
    s.add(row)
    s.commit()
    ev = ClaimEvidence(run_id=run.id, employee_id=e.id, kind="receipt", file=f"{folder}/r.pdf", page=page,
                       position="middle", values={"vendor": vendor, "date": "2026-07-01", "amount": "120.00",
                                                  "currency": "MYR"}, confidence={}, matched_row_id=row.id)
    s.add(ev)
    s.commit()
    row.matched_evidence_id = ev.id
    s.commit()
    return e, row, ev


def test_shared_receipt_across_employees_is_flagged_once_on_each(Session):
    s = Session()
    run = ClaimsRun(client="c", folder_url="", listing_url="", received_date="2026-08-19", instructions="",
                    snapshot={"profile": PROFILE}, status="verifying", listing_headers={"state": "ok"})
    s.add(run)
    s.commit()
    a, arow, _ = _employee_with_matched_receipt(s, run, "A_1", "Alice")
    b, brow, _ = _employee_with_matched_receipt(s, run, "B_2", "Bob", page=3)
    _employee_with_matched_receipt(s, run, "C_3", "Cara", vendor="Starbucks")  # different receipt
    for _ in range(2):  # closing twice must not double the flags
        assert worker._finish_run(run.id, 0.0) is True
    flags = s.query(ClaimFlag).filter(ClaimFlag.run_id == run.id, ClaimFlag.code == "SHARED_RECEIPT").all()
    assert {f.row_id for f in flags} == {arow.id, brow.id}
    by_row = {f.row_id: f for f in flags}
    assert "Bob" in by_row[arow.id].reason and "B_2/r.pdf page 3" in by_row[arow.id].reason
    assert "Alice" in by_row[brow.id].reason and by_row[brow.id].status == "open"
    assert by_row[arow.id].cite == {"file": "A_1/r.pdf", "page": 1, "position": "middle"}
    # a decided one stays decided across another close
    by_row[arow.id].status = "dismissed"
    s.commit()
    assert worker._finish_run(run.id, 0.0) is True
    flags = s.query(ClaimFlag).filter(ClaimFlag.run_id == run.id, ClaimFlag.code == "SHARED_RECEIPT").all()
    assert len(flags) == 2 and {f.status for f in flags} == {"dismissed", "open"}
    s.close()


def test_shared_receipt_check_can_be_off(Session):
    s = Session()
    run = ClaimsRun(client="c", folder_url="", listing_url="", received_date="2026-08-19", instructions="",
                    snapshot={"profile": {**PROFILE, "checks": {"SHARED_RECEIPT": False}}}, status="verifying",
                    listing_headers={"state": "ok"})
    s.add(run)
    s.commit()
    _employee_with_matched_receipt(s, run, "A_1", "Alice")
    _employee_with_matched_receipt(s, run, "B_2", "Bob")
    assert worker._finish_run(run.id, 0.0) is True
    assert s.query(ClaimFlag).filter(ClaimFlag.run_id == run.id, ClaimFlag.code == "SHARED_RECEIPT").count() == 0
    s.close()


# ---- R6: unclaimed receipts that matter ------------------------------------------------

def test_unclaimed_receipt_blocks_at_or_above_the_threshold():
    rows = [_row("r1", "2026-07-01", "45.00")]
    ev = [_receipt("e1", "2026-07-01", "45.00"),
          _receipt("e2", "2026-07-02", "240.00", page=2, vendor="Hilton"),
          _receipt("e3", "2026-07-02", "12.00", page=3, vendor="7-Eleven")]
    res = _checks(rows, ev)  # default threshold RM 100
    unc = {f["evidence_id"]: f for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"}
    assert unc["e2"]["status"] == "open" and "RM 100" in unc["e2"]["reason"] and "missed line" in unc["e2"]["reason"]
    assert unc["e3"]["status"] == "info"
    # the client's threshold moves the line
    res = _checks(rows, ev, {**PROFILE, "unclaimed_receipt_threshold": "300"})
    unc = {f["evidence_id"]: f for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"}
    assert unc["e2"]["status"] == "info" and unc["e3"]["status"] == "info"
    res = _checks(rows, ev, {**PROFILE, "unclaimed_receipt_threshold": "10"})
    unc = {f["evidence_id"]: f for f in res["flags"] if f["code"] == "UNCLAIMED_RECEIPT"}
    assert unc["e2"]["status"] == "open" and unc["e3"]["status"] == "open"


def test_threshold_is_saved_through_settings_and_validated(monkeypatch, tmp_path):
    from app import settings_store
    from app.claims import routes as claims_routes
    engine = create_engine(f"sqlite:///{tmp_path / 's.sqlite3'}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    S = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    for module in (claims_routes, profile, settings_store):
        monkeypatch.setattr(module, "SessionLocal", S)
    client = TestClient(app)
    assert client.put("/api/claims-settings", json={"profile": {"unclaimed_receipt_threshold": "abc"}}).status_code == 400
    r = client.put("/api/claims-settings", json={"profile": {"unclaimed_receipt_threshold": "250"}})
    assert r.status_code == 200
    assert client.get("/api/claims-settings").json()["profile"]["unclaimed_receipt_threshold"] == "250"
    assert profile.PROFILE_DEFAULTS["unclaimed_receipt_threshold"] == "100"
