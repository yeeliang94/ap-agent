"""Reference documents: the payment listing, policy sheet, bank template.

All bytes come through the DocumentSource adapter (local folder in
development, SharePoint MCP on Windows) — this module only parses them
into the shapes the pipeline uses. Each run takes its own copy of the files
at start (see snapshot_references); the AI-read listing is parsed once per
distinct file content.
"""
from __future__ import annotations

import hashlib
import io
import json
from pathlib import Path

from openpyxl import load_workbook

from .. import config
from ..docsource import get_source
from . import listing_agent

# --- finding the files ------------------------------------------------------
# Client folders use human names, so each role is matched by the words that
# must ALL appear in the file name. Deliberately keyword matching and not
# similarity scoring: "FY2025 Payment Listing" and "FY2026 Payment Listing"
# score almost identically, and quietly checking invoices against last
# year's listing is the worst outcome this module can produce.
ROLE_KEYWORDS = {
    "payment_listing": ("payment", "listing"),
    "policy_sheet": ("policy",),
    "bank_template": ("template",),
}

# The fixed names used by the samples folder and the tests. An exact match
# wins outright, so development and CI never depend on keyword matching.
CANONICAL_NAMES = {
    "payment_listing": "payment_listing.xlsx",
    "policy_sheet": "policy_sheet.xlsx",
    "bank_template": "maybank_template.xlsx",
}

# Only these roles stop a run when absent. Missing policy or bank template
# degrade the run instead: see load_policy_clauses / load_maybank_headers.
REQUIRED_ROLES = ("payment_listing",)

_SPREADSHEET_SUFFIXES = (".xlsx", ".xlsm")


class AmbiguousReference(Exception):
    """More than one file could be this role, so the pipeline refuses to pick."""


class MissingReference(Exception):
    """A file the pipeline cannot work without is not in the folder."""


def _candidates(role: str, names: list[str]) -> list[str]:
    keywords = ROLE_KEYWORDS[role]
    out = []
    for name in names:
        # "~$..." are Excel's lock files, created whenever someone has the
        # real workbook open. They are not documents and must never match.
        if name.startswith("~$") or not name.lower().endswith(_SPREADSHEET_SUFFIXES):
            continue
        stem = name.rsplit(".", 1)[0].lower()
        if all(word in stem for word in keywords):
            out.append(name)
    return out


def resolve_name(role: str, names: list[str]) -> str | None:
    """Which file in `names` plays `role`? None means there isn't one.

    Raises AmbiguousReference when several files match, rather than guessing.
    """
    canonical = CANONICAL_NAMES[role]
    for name in names:
        if name.lower() == canonical.lower():
            return name
    matches = _candidates(role, names)
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise AmbiguousReference(
            f"Found {len(matches)} files that could be the {role.replace('_', ' ')}: "
            f"{', '.join(sorted(matches))}. Leave one in the folder (or rename the "
            f"others) so there is no doubt which one this run is judged against."
        )
    return None


# --- one private copy per run --------------------------------------------------
# A run's reference files are downloaded ONCE, when the run starts, into the
# run's own workspace (<runs>/<run_id>/reference/), together with a manifest
# saying which file plays which role. Every later touch — the checks, the
# output builder, a re-check after a correction during review — reads that
# copy and never goes back to SharePoint. Two consequences, both wanted:
#   - review of a run never costs a network round trip;
#   - a run is judged against the files as they were when it started, even
#     if the client changed the workbook or another run started since.
# The AI reading of a listing is still cached across runs by content hash
# (_LISTING_CACHE), so an unchanged file is read by the AI once, ever.
MANIFEST = "manifest.json"


def run_refs(run_id: str) -> Path:
    """Where a run keeps its private copy of the reference files."""
    return config.RUNS_DIR / run_id / "reference"


def snapshot_references(folder_url: str | None, dest: Path) -> dict[str, str | None]:
    """Run start: list the folder, resolve each role, copy each role's file
    into dest, and write the manifest. Returns the role → file-name map,
    which the run records so its history shows what it was judged against.

    Raises MissingReference / AmbiguousReference before copying anything,
    so a run never starts against a folder it cannot be judged by.
    """
    source = get_source(folder_url)
    names = source.list_names()
    roles = {role: resolve_name(role, names) for role in ROLE_KEYWORDS}
    if roles["payment_listing"] is None:
        raise MissingReference(
            "No payment listing found in the reference folder. Expected a "
            "spreadsheet whose name contains: payment, listing.")
    dest.mkdir(parents=True, exist_ok=True)
    for name in roles.values():
        if name:
            (dest / name).write_bytes(source.get_reference(name))
    (dest / MANIFEST).write_text(json.dumps(roles))
    return roles


def ensure_snapshot(refs: Path, folder_url: str | None) -> dict[str, str | None]:
    """The manifest for a run — taking the snapshot now if it has none.

    Runs created before snapshots existed have no reference folder; the
    first review action snapshots the folder for them, once. New runs never
    hit this path: the runner snapshots explicitly at start.
    """
    manifest = refs / MANIFEST
    if manifest.is_file():
        return json.loads(manifest.read_text())
    return snapshot_references(folder_url, refs)


def _role_file(refs: Path, role: str) -> Path | None:
    roles = json.loads((refs / MANIFEST).read_text())
    name = roles.get(role)
    if name is None:
        if role in REQUIRED_ROLES:
            raise MissingReference(
                f"No {role.replace('_', ' ')} was in the reference folder when "
                "this run started.")
        return None
    return refs / name


def _open(role: str, refs: Path):
    """Open the run's copy of the workbook for `role`, or None if it had none."""
    path = _role_file(refs, role)
    if path is None:
        return None
    return load_workbook(io.BytesIO(path.read_bytes()), read_only=True)


# Parsed listings keyed by content hash, as {"rows": [...], "notes":
# [(level, text), ...]}. The AI reading must be paid for once per distinct
# file, not once per run or per touch. The notes are cached with the rows so
# a later run on an unchanged file can still say how it was read.
_LISTING_CACHE: dict[str, dict] = {}

# A payment listing larger than this is refused before it is opened. The
# real client file is well under a megabyte; the same ceiling as a single
# uploaded document keeps a wrong file from exhausting memory.
MAX_LISTING_BYTES = config.MAX_FILE_MB * 1024 * 1024


async def _listing_entry(refs: Path) -> dict:
    path = _role_file(refs, "payment_listing")
    size = path.stat().st_size
    if size > MAX_LISTING_BYTES:
        raise listing_agent.ListingUnreadable(
            f"The payment listing is {size / 1024 / 1024:.1f} MB — larger than "
            f"the {config.MAX_FILE_MB} MB this reader supports.")
    data = path.read_bytes()
    key = hashlib.sha256(data).hexdigest()
    if key in _LISTING_CACHE:
        return _LISTING_CACHE[key]

    # data_only=True: balance columns are usually formulas, and the
    # audit needs their cached values, not "=J5-K5" strings. The second
    # open keeps the formulas, so a formula whose value was never saved
    # can be named instead of silently reading as an empty cell.
    wb = load_workbook(io.BytesIO(data), data_only=True)
    wb_formulas = load_workbook(io.BytesIO(data))
    try:
        result = await listing_agent.ingest_workbook(wb, wb_formulas)
    finally:
        wb.close()
        wb_formulas.close()
    entry = {"rows": result.rows, "notes": result.notes}
    _LISTING_CACHE[key] = entry
    return entry


async def load_payment_listing(refs: Path) -> list[dict]:
    """Every past-payment row as {sheet, row, no, date, vendor,
    invoice_number, amount, status, note}.

    Every listing — the samples included — goes through listing_agent's
    reason/act loop: the AI maps the structure, code extracts and audits the
    numbers. One reader, one behaviour. amount can be None where the file
    genuinely doesn't pair an amount to an invoice — callers must treat that
    as "cannot compare", never zero.
    """
    return (await _listing_entry(refs))["rows"]


async def load_listing_notes(refs: Path) -> list[tuple[str, str]]:
    """How the listing was read, tab by tab, as (level, sentence) pairs for
    the run's Activity tab. Reading it (and paying for it) happens on the
    first call; later calls replay from the cache."""
    return (await _listing_entry(refs))["notes"]


def load_policy_clauses(refs: Path) -> list[dict]:
    """Every policy clause as {clause, category, cap, currency, text}.

    Returns [] when the folder has no policy sheet. Callers treat an empty
    list as "no policy to test against", which means no spending cap is
    checked — recorded on the run by snapshot_references.
    """
    wb = _open("policy_sheet", refs)
    if wb is None:
        return []
    ws = wb.active
    clauses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        clauses.append({
            "clause": str(row[0]), "category": str(row[1]),
            "cap": float(row[2]), "currency": str(row[3]), "text": str(row[4]),
        })
    wb.close()
    return clauses


def load_maybank_headers(refs: Path) -> list[str]:
    """The column layout of the bank upload template — learned, not hardcoded.

    Returns [] when the folder has no template. build_outputs then omits the
    bank upload block entirely rather than inventing a column layout.
    """
    wb = _open("bank_template", refs)
    if wb is None:
        return []
    ws = wb.active
    headers = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    wb.close()
    return headers
