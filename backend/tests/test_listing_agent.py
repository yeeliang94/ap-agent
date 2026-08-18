"""The AI payment-listing reader, with the AI faked out.

The workbook built here is a miniature of the real ICMR listing: title
rows, headers on row 4, monthly-style tab, one payment spanning several
rows with several invoices inside, balance b/f and fund-received lines,
and a summary block. The "AI" is a scripted stand-in, so these tests cost
nothing and never flake — they pin the CODE half of the loop:

  - a correct reading extracts, pairs and flattens the grouped entries
  - the audit catches a wrong column, a wrongly-cut span, a merged span,
    an omitted payment, and missing required columns — the adversarial
    cases where pure arithmetic alone would be fooled
  - the loop feeds problems back and accepts the corrected second answer
  - a reading that never verifies raises ListingUnreadable, not a guess
  - unpairable invoice groups keep amount=None (checks skips, never
    zero-compares)
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from app.pipeline import listing_agent
from app.pipeline.listing_agent import (
    ColumnRoles, EntrySpan, ListingUnreadable, SheetReading,
    flatten_reading, verify_reading,
)


def _icmr_sheet():
    """Rows mirror the structure (not the data) of the real client file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Jul'26"
    ws["A1"] = "Name:"; ws["B1"] = "Institute For Capital Market Research"
    ws["A2"] = "A/C No:"; ws["B2"] = "514712417644"
    # headers on row 4, not row 1 — receipt (money in) is column I
    for col, head in zip("ABCDEFGHI", ["Date", "Cheque/Journal No.",
                                       "Invoice / Reference No.", "Payee Name",
                                       "Description", "Line (MYR)",
                                       "Payment (MYR)", "Balance (MYR)",
                                       "Receipt (MYR)"]):
        ws[f"{col}4"] = head
    # balance b/f, then money in
    ws["E5"] = "Balance b/f"; ws["H5"] = 7.90
    ws["A6"] = "2026-07-07"; ws["E6"] = "Fund received"
    ws["I6"] = 37195.98; ws["H6"] = 37203.88
    # one grouped payment: two invoices, two line amounts, one payment
    ws["A8"] = "2026-07-23"; ws["B8"] = "PV0726/01"; ws["C8"] = "580261111513"
    ws["D8"] = "PricewaterhouseCoopers Taxation Services Sdn. Bhd."
    ws["E8"] = "April 2026 payroll fees"; ws["F8"] = 5572.80
    ws["C9"] = "580271100169"; ws["E9"] = "Accounting fee"; ws["F9"] = 10044.00
    ws["G8"] = 15616.80; ws["H9"] = 21587.08
    # single-line payment
    ws["A10"] = "2026-07-23"; ws["B10"] = "PV0726/02"; ws["C10"] = "13561"
    ws["D10"] = "Good News Resources Sdn Bhd"; ws["E10"] = "Name cards"
    ws["G10"] = 195.00; ws["H10"] = 21392.08
    # unpairable group: two invoices, FOUR line amounts
    ws["A12"] = "2026-07-23"; ws["B12"] = "PV0726/07"; ws["C12"] = "245DHNQL-0015"
    ws["D12"] = "Lim Shea-Fee"; ws["F12"] = 8000.00
    ws["C13"] = "CA50CBEE-0015"; ws["F13"] = 1044.95
    ws["F14"] = 343.70; ws["F15"] = 127.60
    ws["G12"] = 9516.25; ws["H15"] = 11875.83
    # summary block (column B — outside the money columns, so no span)
    ws["A18"] = "Opening balance to utilise"; ws["B18"] = 7.90
    ws["A19"] = "Net Payment"; ws["B19"] = 37193.88
    return wb, ws


GOOD_COLUMNS = ColumnRoles(date="A", voucher_no="B", invoice_no="C",
                           payee="D", description="E", line_amount="F",
                           payment="G", balance="H", receipt="I")

GOOD_READING = SheetReading(
    is_payment_sheet=True,
    columns=GOOD_COLUMNS,
    entries=[
        EntrySpan(first_row=5, last_row=5, kind="other"),
        EntrySpan(first_row=6, last_row=6, kind="other"),
        EntrySpan(first_row=8, last_row=9, kind="payment"),
        EntrySpan(first_row=10, last_row=10, kind="payment"),
        EntrySpan(first_row=12, last_row=15, kind="payment"),
    ],
    why="headers row 4, grouped entries",
)


def test_correct_reading_verifies_and_flattens():
    wb, ws = _icmr_sheet()
    assert verify_reading(ws, GOOD_READING) == []

    rows = flatten_reading(ws, GOOD_READING)
    by_number = {r["invoice_number"]: r for r in rows}
    # grouped entry: invoice numbers paired to line amounts one-to-one
    assert by_number["580261111513"]["amount"] == pytest.approx(5572.80)
    assert by_number["580271100169"]["amount"] == pytest.approx(10044.00)
    assert by_number["580261111513"]["vendor"].startswith("PricewaterhouseCoopers")
    # single-invoice entry takes the payment total
    assert by_number["13561"]["amount"] == pytest.approx(195.00)
    # 2 invoices among 4 line amounts: paired BY ROW (the sheet puts each
    # number beside its amount), never by list position and never guessed
    assert by_number["245DHNQL-0015"]["amount"] == pytest.approx(8000.00)
    assert by_number["CA50CBEE-0015"]["amount"] == pytest.approx(1044.95)
    # recorded payments are Paid — the duplicate-payment guard depends on it
    assert all(r["status"] == "Paid" for r in rows)
    assert by_number["13561"]["date"] == "2026-07-23"
    assert by_number["13561"]["no"] == "PV0726/02"
    # every row knows where it came from, so a match can be pointed at:
    # the row the number sits in, and where its payment entry starts
    assert by_number["13561"]["sheet"] == "Jul'26"
    assert by_number["13561"]["row"] == 10
    assert by_number["CA50CBEE-0015"]["row"] == 13
    assert by_number["CA50CBEE-0015"]["entry_row"] == 12


def test_invoice_paired_by_row_not_by_position():
    """The invoice numbers sit on rows 2 and 3 of a 4-line entry, so the
    amount on the entry's FIRST row belongs to nobody — position-based
    zipping would hand it to the first invoice."""
    wb = Workbook(); ws = wb.active; ws.title = "T"
    for col, head in zip("ABCDEF", ["Date", "PV", "Invoice", "Payee",
                                    "Line", "Payment"]):
        ws[f"{col}1"] = head
    ws["A2"] = "2026-07-23"; ws["B2"] = "PV1"; ws["D2"] = "Lim"; ws["E2"] = 8000.00
    ws["C3"] = "GEN-1"; ws["E3"] = 1044.95
    ws["C4"] = "GPT-1"; ws["E4"] = 343.70
    ws["E5"] = 127.60; ws["F2"] = 9516.25
    reading = SheetReading(
        is_payment_sheet=True,
        columns=ColumnRoles(date="A", voucher_no="B", invoice_no="C",
                            payee="D", line_amount="E", payment="F"),
        entries=[EntrySpan(first_row=2, last_row=5, kind="payment")], why="t")
    assert verify_reading(ws, reading) == []
    by_number = {r["invoice_number"]: r for r in flatten_reading(ws, reading)}
    assert by_number["GEN-1"]["amount"] == pytest.approx(1044.95)
    assert by_number["GPT-1"]["amount"] == pytest.approx(343.70)


def test_two_lists_are_not_paired_by_position():
    """Two invoice numbers on rows 2-3 and two amounts on rows 4-5: the
    sheet has not paired them, so neither do we — amounts stay unknown."""
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws["A2"] = "2026-07-23"; ws["B2"] = "PV1"; ws["D2"] = "Vendor"
    ws["C2"] = "INV-A"; ws["C3"] = "INV-B"
    ws["E4"] = 100.0; ws["E5"] = 50.0; ws["F2"] = 150.0
    reading = SheetReading(
        is_payment_sheet=True,
        columns=ColumnRoles(date="A", voucher_no="B", invoice_no="C",
                            payee="D", line_amount="E", payment="F"),
        entries=[EntrySpan(first_row=2, last_row=5, kind="payment")], why="t")
    assert verify_reading(ws, reading) == []
    rows = flatten_reading(ws, reading)
    assert {r["invoice_number"]: r["amount"] for r in rows} == {"INV-A": None, "INV-B": None}


def test_parenthesised_reference_is_still_a_reference():
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws["A2"] = "2026-07-23"; ws["B2"] = "PV1"; ws["C2"] = "(INV-123)"
    ws["D2"] = "Vendor"; ws["F2"] = 10.0
    reading = SheetReading(
        is_payment_sheet=True,
        columns=ColumnRoles(date="A", voucher_no="B", invoice_no="C",
                            payee="D", payment="F"),
        entries=[EntrySpan(first_row=2, last_row=2, kind="payment")], why="t")
    rows = flatten_reading(ws, reading)
    assert [r["invoice_number"] for r in rows] == ["(INV-123)"] and rows[0]["note"] == ""


def test_truncated_entry_leaving_invoice_rows_uncovered_is_structural():
    """Payment on the entry's first row, balance on its last, invoices in
    between: spanning only the first and last rows leaves invoice rows
    outside every span. Money-column coverage cannot see that; invoice
    and line-amount coverage must, as a STRUCTURE problem — never as an
    arithmetic nit that could be accepted."""
    from app.pipeline.listing_agent import STRUCTURE, audit_reading
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws["A1"] = "Date"; ws["C1"] = "Invoice"; ws["D1"] = "Payee"; ws["G1"] = "Payment"; ws["H1"] = "Balance"
    ws["H2"] = 10000.0                                             # balance b/f
    ws["A3"] = "2026-07-23"; ws["B3"] = "PV1"; ws["D3"] = "Lim"; ws["E3"] = 8000.0; ws["G3"] = 9516.25
    ws["C4"] = "GEN-1"; ws["E4"] = 1044.95
    ws["C5"] = "GPT-1"; ws["E5"] = 343.70
    ws["E6"] = 127.60; ws["H6"] = 483.75
    cols = ColumnRoles(date="A", voucher_no="B", invoice_no="C", payee="D",
                       line_amount="E", payment="G", balance="H")
    cut = SheetReading(is_payment_sheet=True, columns=cols, why="t", entries=[
        EntrySpan(first_row=2, last_row=2, kind="other"),
        EntrySpan(first_row=3, last_row=3, kind="payment"),
        EntrySpan(first_row=6, last_row=6, kind="other")])
    problems = audit_reading(ws, cut)
    kinds = {k for k, _ in problems}
    assert STRUCTURE in kinds
    assert any("C4" in t and "belongs to no span" in t for _, t in problems)
    assert any("line amounts sum" in t and k == STRUCTURE for k, t in problems)


@pytest.mark.asyncio
async def test_truncated_entry_is_rejected_after_all_rounds(monkeypatch):
    """The end-to-end guarantee: a reading that keeps dropping invoice rows
    is REJECTED after MAX_ROUNDS, not accepted with a warning."""
    wb, ws = _icmr_sheet()
    # PwC's second invoice (row 9) left outside every span; the balance
    # cell H9 is blanked so money-column coverage alone would not notice.
    ws["H9"] = None
    cut = GOOD_READING.model_copy(update={"entries": [
        e if e.first_row != 8 else EntrySpan(first_row=8, last_row=8, kind="payment")
        for e in GOOD_READING.entries]})
    agent = _ScriptedAgent([cut] * listing_agent.MAX_ROUNDS)
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)
    with pytest.raises(ListingUnreadable):
        await listing_agent.read_sheet(ws)


def test_summary_rows_in_invoice_column_need_a_span():
    """The ICMR summary block puts its figures in the invoice column below
    the entries. They must be inside an 'other' span, or the reading is
    incomplete; rows ABOVE the first span (title, header) are exempt."""
    from app.pipeline.listing_agent import audit_reading
    wb, ws = _icmr_sheet()
    ws["C18"] = 7.90; ws["C19"] = 37193.88   # summary figures under 'Invoice'
    problems = audit_reading(ws, GOOD_READING)
    assert any("C18" in t or "C19" in t for _, t in problems)
    with_summary = GOOD_READING.model_copy(update={
        "entries": GOOD_READING.entries + [EntrySpan(first_row=18, last_row=19, kind="other")]})
    assert audit_reading(ws, with_summary) == []


def test_note_text_in_invoice_column_is_a_note_not_a_number():
    """'(Revised invoice)' under a reference number is a remark. Treating it
    as a second invoice would both invent a number to match against and
    strip the real invoice of its amount."""
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws["A2"] = "2026-07-23"; ws["B2"] = "PV4"; ws["C2"] = "EC26700197"
    ws["D2"] = "Enfrasys"; ws["F2"] = 1096.41
    ws["C3"] = "(Revised invoice)"
    reading = SheetReading(
        is_payment_sheet=True,
        columns=ColumnRoles(date="A", voucher_no="B", invoice_no="C",
                            payee="D", payment="F"),
        entries=[EntrySpan(first_row=2, last_row=3, kind="payment")], why="t")
    rows = flatten_reading(ws, reading)
    assert [r["invoice_number"] for r in rows] == ["EC26700197"]
    assert rows[0]["amount"] == pytest.approx(1096.41)
    assert rows[0]["note"] == "(Revised invoice)"


def test_status_column_is_used_when_the_sheet_has_one():
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws.append(["No.", "Date", "Vendor", "Invoice No.", "Amount (RM)", "Status"])
    ws.append(["0701", "2026-07-03", "Maxis Bhd", "MX-7101", 1240.00, "Planned"])
    ws.append(["0605", "2026-06-10", "Maxis Bhd", "MX-6580", 1240.00, "Paid"])
    reading = SheetReading(
        is_payment_sheet=True,
        columns=ColumnRoles(voucher_no="A", date="B", payee="C", invoice_no="D",
                            payment="E", status="F"),
        entries=[EntrySpan(first_row=2, last_row=2, kind="payment"),
                 EntrySpan(first_row=3, last_row=3, kind="payment")], why="t")
    assert verify_reading(ws, reading) == []
    by_number = {r["invoice_number"]: r for r in flatten_reading(ws, reading)}
    assert by_number["MX-7101"]["status"] == "Planned"
    assert by_number["MX-6580"]["status"] == "Paid"


def test_multi_invoice_entry_without_line_column_is_questioned():
    """A missed (often header-less) line-amount column would silently strip
    every grouped invoice of its amount — the audit must ask about it."""
    wb, ws = _icmr_sheet()
    bare = GOOD_READING.model_copy(update={
        "columns": GOOD_COLUMNS.model_copy(update={"line_amount": None})})
    problems = verify_reading(ws, bare)
    assert any("no line_amount column" in p for p in problems)


def test_wrong_payment_column_is_caught():
    wb, ws = _icmr_sheet()
    bad = GOOD_READING.model_copy(update={
        "columns": GOOD_COLUMNS.model_copy(update={"payment": "F", "line_amount": None})})
    problems = verify_reading(ws, bad)
    # the balance moves by column G, not F — the walk must object somewhere
    assert any("running balance" in p for p in problems)


def test_wrongly_cut_span_is_caught():
    from app.pipeline.listing_agent import STRUCTURE, audit_reading
    wb, ws = _icmr_sheet()
    bad = GOOD_READING.model_copy(update={
        "entries": [e for e in GOOD_READING.entries
                    if not (e.first_row == 8 and e.last_row == 9)]
        + [EntrySpan(first_row=8, last_row=8, kind="payment")]})
    problems = audit_reading(ws, bad)
    # a wrong cut is a STRUCTURE problem — it can never be waved through
    assert any("line amounts sum" in t and k == STRUCTURE for k, t in problems)
    assert any("C9" in t and "belongs to no span" in t for _, t in problems)


def test_merged_spans_are_caught():
    """Two payments glued into one span pass pure arithmetic (sums are
    additive) — identity constraints are what must catch them."""
    wb, ws = _icmr_sheet()
    merged = GOOD_READING.model_copy(update={
        "entries": [
            EntrySpan(first_row=5, last_row=5, kind="other"),
            EntrySpan(first_row=6, last_row=6, kind="other"),
            EntrySpan(first_row=8, last_row=10, kind="payment"),  # 01 + 02 merged
            EntrySpan(first_row=12, last_row=15, kind="payment"),
        ]})
    problems = verify_reading(ws, merged)
    assert any("exactly one payment total" in p for p in problems)
    assert any("different payees" in p for p in problems)


def test_omitted_payment_is_caught():
    """Showing only part of the sheet must fail: every number in a money
    column has to be inside some span."""
    wb, ws = _icmr_sheet()
    partial = GOOD_READING.model_copy(update={
        "entries": [e for e in GOOD_READING.entries if e.first_row != 12]})
    problems = verify_reading(ws, partial)
    assert any("belongs to no span" in p and "G12" in p for p in problems)


def test_missing_required_columns_is_caught():
    wb, ws = _icmr_sheet()
    bare = GOOD_READING.model_copy(update={
        "columns": GOOD_COLUMNS.model_copy(update={"payment": None, "payee": None})})
    problems = verify_reading(ws, bare)
    assert any("payment column" in p for p in problems)
    assert any("payee column" in p for p in problems)


def test_overlapping_spans_are_caught():
    wb, ws = _icmr_sheet()
    bad = GOOD_READING.model_copy(update={
        "entries": [EntrySpan(first_row=8, last_row=10, kind="payment"),
                    EntrySpan(first_row=10, last_row=15, kind="payment")]})
    assert any("overlap" in p for p in verify_reading(ws, bad))


def test_header_text_as_column_letter_is_rejected_at_schema():
    with pytest.raises(ValidationError):
        ColumnRoles(payment="Payment")


def test_overlong_sheet_fails_loudly():
    wb = Workbook()
    ws = wb.active
    for r in range(1, listing_agent.MAX_SHEET_ROWS + 2):
        ws[f"A{r}"] = "x"
    with pytest.raises(ListingUnreadable):
        listing_agent.grid_text(ws)


def test_too_many_columns_fails_loudly():
    wb = Workbook(); ws = wb.active
    from openpyxl.utils import get_column_letter
    ws[f"{get_column_letter(listing_agent.MAX_SHEET_COLUMNS + 1)}1"] = "x"
    with pytest.raises(ListingUnreadable):
        listing_agent.grid_text(ws)


@pytest.mark.asyncio
async def test_too_many_sheets_fails_before_any_ai_call(monkeypatch):
    wb = Workbook()
    for i in range(listing_agent.MAX_SHEETS):
        wb.create_sheet(f"S{i}")
    agent = _ScriptedAgent([])
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)
    with pytest.raises(ListingUnreadable):
        await listing_agent.ingest_workbook(wb)
    assert agent.prompts == []


def test_long_cell_texts_are_capped_on_the_flat_rows():
    """Grid text already truncates what the AI sees; the flattened rows must
    not carry a 10,000-character 'invoice number' into flags and outputs."""
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws["A2"] = "2026-07-23"; ws["B2"] = "PV1"; ws["C2"] = "X" * 5000
    ws["D2"] = "V" * 5000; ws["F2"] = 10.0
    reading = SheetReading(
        is_payment_sheet=True,
        columns=ColumnRoles(date="A", voucher_no="B", invoice_no="C",
                            payee="D", payment="F"),
        entries=[EntrySpan(first_row=2, last_row=2, kind="payment")], why="t")
    row = flatten_reading(ws, reading)[0]
    assert len(row["invoice_number"]) == listing_agent.MAX_CELL_TEXT
    assert len(row["vendor"]) == listing_agent.MAX_CELL_TEXT


@pytest.mark.asyncio
async def test_stale_formula_caches_are_named_not_silent(monkeypatch):
    """A workbook saved without recalculating stores formulas with no
    cached value; data_only reading gives None where the balance should
    be, which silently weakens the balance walk. Name the cells."""
    import io
    from openpyxl import load_workbook
    wb, ws = _icmr_sheet()
    ws["H9"] = "=H6-G8"; ws["H10"] = "=H9-G10"      # formulas, never computed
    buf = io.BytesIO(); wb.save(buf)
    values = load_workbook(io.BytesIO(buf.getvalue()), data_only=True)
    formulas = load_workbook(io.BytesIO(buf.getvalue()))
    assert values["Jul'26"]["H9"].value is None      # the stale-cache shape
    agent = _ScriptedAgent([GOOD_READING] * listing_agent.MAX_ROUNDS)
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)
    result = await listing_agent.read_sheet(values["Jul'26"], formulas["Jul'26"])
    stale = [t for lvl, t in result.notes if lvl == "WARNING" and "no saved value" in t]
    assert len(stale) == 1 and "H9" in stale[0] and "H10" in stale[0]
    assert "recalculate" in stale[0]


@pytest.mark.asyncio
async def test_hidden_tab_is_read_and_said_to_be_hidden(monkeypatch):
    wb, ws = _icmr_sheet()
    ws.sheet_state = "hidden"
    agent = _ScriptedAgent([GOOD_READING])
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)
    result = await listing_agent.read_sheet(ws)
    assert result.rows                       # hidden is not skipped
    assert "(hidden tab)" in result.notes[0][1]


def test_formatted_but_empty_rows_do_not_count():
    """A grid formatted far below the data is not a long sheet. Excel's
    max_row says 500; the content says 1."""
    from openpyxl.styles import PatternFill
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    fill = PatternFill("solid", fgColor="C6EFCE")
    for r in range(2, listing_agent.MAX_SHEET_ROWS + 200):
        ws[f"A{r}"].fill = fill
    assert ws.max_row > listing_agent.MAX_SHEET_ROWS
    assert listing_agent.content_rows(ws) == [1]
    listing_agent.grid_text(ws)  # does not raise


class _ScriptedAgent:
    """Stands in for the AI: answers from a fixed script, records prompts."""

    def __init__(self, outputs: list[SheetReading]) -> None:
        self._outputs = list(outputs)
        self.prompts: list[str] = []

    async def run(self, prompt, **kwargs):
        self.prompts.append(prompt)

        class R:
            output = self._outputs.pop(0)
        return R()


@pytest.mark.asyncio
async def test_loop_feeds_problems_back_and_accepts_correction(monkeypatch):
    wb, ws = _icmr_sheet()
    wrong = GOOD_READING.model_copy(update={
        "columns": GOOD_COLUMNS.model_copy(update={"payment": "F", "line_amount": None})})
    agent = _ScriptedAgent([wrong, GOOD_READING])
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)

    result = await listing_agent.read_sheet(ws)

    assert len(agent.prompts) == 2
    # the second prompt must carry the verifier's objection — that feedback
    # IS the "reason through and act" loop
    assert "failed verification" in agent.prompts[1]
    assert {r["invoice_number"] for r in result.rows} >= {"580261111513", "13561"}
    # and the Activity tab is told how the tab was read, and what round 1 got wrong
    info = [t for lvl, t in result.notes if lvl == "INFO"]
    assert info and "Jul'26" in info[0] and "round 2" in info[0]
    assert "round 1 was corrected" in info[0]


@pytest.mark.asyncio
async def test_ai_observations_reach_the_activity_tab(monkeypatch):
    """Quirks the AI noticed while mapping the sheet ('column F has no
    header') are printed, so a silent oddity becomes a visible one."""
    wb, ws = _icmr_sheet()
    seen = GOOD_READING.model_copy(update={"observations": [
        "column F has no header", "rows 26/28 are recurring payments"]})
    agent = _ScriptedAgent([seen])
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)
    result = await listing_agent.read_sheet(ws)
    obs = [t for lvl, t in result.notes if "AI observations" in t]
    assert len(obs) == 1 and lvl_ok(result.notes)
    assert "column F has no header" in obs[0] and "recurring payments" in obs[0]


def lvl_ok(notes) -> bool:
    return all(lvl in ("INFO", "WARNING") for lvl, _ in notes)


@pytest.mark.asyncio
async def test_arithmetic_only_leftover_is_accepted_with_a_warning(monkeypatch):
    """The listing exists to answer 'was this invoice paid before?'. Once
    the reading's STRUCTURE verifies (columns named, every payment covered,
    entries not merged), a client's own typo in a balance cell must not
    block that answer — it is accepted and said out loud, rows named."""
    wb, ws = _icmr_sheet()
    ws["H10"] = 21392.18  # 10 cents off: 21587.08 - 195.00 = 21392.08
    agent = _ScriptedAgent([GOOD_READING] * listing_agent.MAX_ROUNDS)
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)

    result = await listing_agent.read_sheet(ws)

    assert len(agent.prompts) == listing_agent.MAX_ROUNDS  # it did try
    assert {r["invoice_number"] for r in result.rows} >= {"580261111513", "13561"}
    warnings = [t for lvl, t in result.notes if lvl == "WARNING"]
    assert len(warnings) == 1
    assert "rows 10-10" in warnings[0] and "21392.08" in warnings[0]


@pytest.mark.asyncio
async def test_structural_leftover_still_fails(monkeypatch):
    """A reading that never names the payment column is not a reading;
    accepting it would silently drop the duplicate-payment guard."""
    wb, ws = _icmr_sheet()
    bare = GOOD_READING.model_copy(update={
        "columns": GOOD_COLUMNS.model_copy(update={"payment": None})})
    agent = _ScriptedAgent([bare] * listing_agent.MAX_ROUNDS)
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)
    with pytest.raises(ListingUnreadable):
        await listing_agent.read_sheet(ws)


@pytest.mark.asyncio
async def test_unverifiable_reading_raises_not_guesses(monkeypatch):
    wb, ws = _icmr_sheet()
    wrong = GOOD_READING.model_copy(update={
        "columns": GOOD_COLUMNS.model_copy(update={"payment": "F", "line_amount": None})})
    agent = _ScriptedAgent([wrong] * listing_agent.MAX_ROUNDS)
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)

    with pytest.raises(ListingUnreadable):
        await listing_agent.read_sheet(ws)


@pytest.mark.asyncio
async def test_not_payment_sheet_is_challenged_once(monkeypatch):
    """The miniature carries payment-style headers, so a 'not a payment
    sheet' answer gets one push-back; a repeated 'no' is accepted."""
    wb, ws = _icmr_sheet()
    no = SheetReading(is_payment_sheet=False, why="cover page")
    agent = _ScriptedAgent([no, no])
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)

    result = await listing_agent.read_sheet(ws)
    assert result.rows == []
    assert len(agent.prompts) == 2
    assert "payment-style column headers" in agent.prompts[1]
    # accepted, but as a WARNING the reviewer sees — this is the shape of
    # a silently skipped month
    assert result.notes[0][0] == "WARNING"
    assert "SKIPPED although" in result.notes[0][1] and "cover page" in result.notes[0][1]


@pytest.mark.asyncio
async def test_plain_cover_sheet_is_skipped_without_challenge(monkeypatch):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "ICMR - FY2026 Payment Listing"; ws["A2"] = "Prepared by: WC"
    agent = _ScriptedAgent([SheetReading(is_payment_sheet=False, why="cover")])
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)
    assert (await listing_agent.read_sheet(ws)).rows == []
    assert len(agent.prompts) == 1


@pytest.mark.asyncio
async def test_crashing_reading_becomes_feedback_not_exception(monkeypatch):
    """A structurally valid but nonsensical reading (column beyond the
    sheet) must turn into loop feedback, not escape as a crash."""
    wb, ws = _icmr_sheet()
    weird = GOOD_READING.model_copy(update={
        "columns": GOOD_COLUMNS.model_copy(update={"payment": "ZZZ"})})
    agent = _ScriptedAgent([weird, GOOD_READING])
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: agent)

    result = await listing_agent.read_sheet(ws)
    assert len(agent.prompts) == 2
    assert result.rows  # recovered on the corrected reading


def test_number_parsing_tolerates_text_amounts():
    assert listing_agent._num("RM 1,200.00") == pytest.approx(1200.00)
    assert listing_agent._num("(500.00)") == pytest.approx(-500.00)
    assert listing_agent._num("MYR 37,195.98") == pytest.approx(37195.98)
    assert listing_agent._num("two hundred") is None
    assert listing_agent._num(None) is None
    assert listing_agent._num(True) is None


def test_grid_text_labels_cells_with_coordinates():
    wb, ws = _icmr_sheet()
    grid = listing_agent.grid_text(ws)
    assert "Sheet name: \"Jul'26\"" in grid
    assert "B8: PV0726/01" in grid
    assert "G8: 15616.8" in grid
