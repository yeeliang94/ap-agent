"""Draft next month's payment-listing entries in the client's own layout.

A DRAFT, never the live listing. The listing is the record of EXECUTED
payments and the reader treats every entry with a payment total as paid;
writing an approved-but-unpaid month into it would make a later run call
those invoices ALREADY_PAID. So this module produces a copy of the client
workbook with one new tab appended (saved in the run's own folder, offered
for download) that a person finalises after the bank run. Nothing is
written to SharePoint or to any working file.

Deterministic writer, AI-drafted text only. Every number is computed here
in Decimal (2 dp, half-up); the only AI text is each invoice's one-line
description, read off the invoice by the extraction stage. The layout —
which column is what, where the headers are, the closing balance, the
last voucher, how the client spells each payee — comes from ListingLayout,
learned by the reader from the latest tab (listing_layout).

Business rules, as recommended in docs/LISTING-HARDENING.md and awaiting
the stakeholder's confirmation (each is one place to change):

  - Grouping: one payment block per vendor. Vendor identity is the
    listing's own spelling when it has paid them before (tolerant name
    match, the same one the checks use), else the invoice's text.
  - Voucher numbers continue the sequence on the latest tab: a month code
    in the number (PV0726/03 -> PV0826/01) rolls to the draft month and
    restarts; otherwise the trailing integer just increments. A generated
    number that already exists anywhere in the listing is a refusal.
  - Money: Balance b/f = the latest tab's closing balance; Net payment =
    sum of the new payments; Estimated bank charges = per-payment charge
    (Settings) x number of payments; Fund received = Total fund to request
    = Net payment + charges, so the balance returns to the same residual.
  - Balance cells and grouped totals are FORMULAS (as the client keeps
    them); everything else is a value. Dates are left blank: the payment
    has not happened.

Round trip: before the draft is delivered, the same plan is written a
second time as plain values and read back through the reader's own audit
(audit_reading) and flattening (flatten_reading). The result must audit
clean and flatten to exactly the approved invoices, or the draft is
refused. No AI is involved in that check.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import ROUND_HALF_UP, Decimal

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from .listing_agent import EntrySpan, SheetReading, audit_reading, flatten_reading
from .listing_layout import ListingLayout, _norm

CENT = Decimal("0.01")

# Month formats a tab title may use; the draft title copies the pattern.
_TITLE_FORMATS = ("%b'%y", "%b %y", "%b %Y", "%B %Y", "%B'%y", "%b-%y",
                  "%b-%Y", "%B-%Y", "%Y-%m", "%m-%Y", "%b%y", "%B", "%b")


class DraftError(Exception):
    """The draft was not written; the message says why."""


@dataclass
class Draft:
    data: bytes                    # the copy of the workbook with the new tab
    suffix: str                    # ".xlsx" or ".xlsm"
    tab: str
    summary: dict                  # what was written, JSON-ready (Decimals as strings)
    round_trip: dict = field(default_factory=dict)


def money(value) -> Decimal:
    return Decimal(str(value)).quantize(CENT, rounding=ROUND_HALF_UP)


def _safe(text) -> str:
    """Cell-safe text: no control characters, and never a leading formula
    character (openpyxl writes '=...' strings as formulas)."""
    s = re.sub(r"[\x00-\x1f]", " ", str(text or "")).strip()
    return s.lstrip("=+@") if s[:1] in "=+@" else s


def _add_month(d: date) -> date:
    return date(d.year + (d.month == 12), 1 if d.month == 12 else d.month + 1, 1)


# ---- voucher numbers ------------------------------------------------------------

def next_vouchers(last: str | None, last_date: date | None, draft_month: date,
                  count: int) -> list[str]:
    """Continue the client's voucher sequence for `count` new entries.

    'PV0726/03' with a July latest tab and an August draft -> PV0826/01..:
    the month code (MMYY, YYMM, MMYYYY or YYYYMM of the latest payment
    date) rolls to the draft month and the sequence restarts. Same month,
    or no month code -> the trailing integer continues, width preserved.
    No voucher to continue from -> blanks (the person numbers them).
    """
    if not last:
        return [""] * count
    m = re.match(r"^(?P<prefix>.*?)(?P<seq>\d+)$", last.strip())
    if not m:
        return [""] * count
    prefix, seq = m.group("prefix"), m.group("seq")
    width, start = len(seq), int(seq) + 1
    if last_date is not None and (last_date.year, last_date.month) != \
            (draft_month.year, draft_month.month):
        codes = {
            last_date.strftime("%m%y"): draft_month.strftime("%m%y"),
            last_date.strftime("%y%m"): draft_month.strftime("%y%m"),
            last_date.strftime("%m%Y"): draft_month.strftime("%m%Y"),
            last_date.strftime("%Y%m"): draft_month.strftime("%Y%m"),
        }
        for old, new in codes.items():
            if old in prefix:
                prefix, start = prefix.replace(old, new, 1), 1
                break
    return [f"{prefix}{n:0{width}d}" for n in range(start, start + count)]


def draft_title(latest_title: str, draft_month: date, existing: list[str]) -> str:
    """'Jul'26' -> "Aug'26 (DRAFT)": the client's own month pattern, marked
    as a draft, made unique among the workbook's tabs."""
    from datetime import datetime
    base = None
    for fmt in _TITLE_FORMATS:
        try:
            datetime.strptime(latest_title.strip(), fmt)
            base = draft_month.strftime(fmt)
            break
        except ValueError:
            continue
    if base is None:
        base = f"{latest_title.strip()} {draft_month.strftime('%b')}'{draft_month.strftime('%y')}"
    title, n = f"{base} (DRAFT)", 2
    while title in existing:
        title, n = f"{base} (DRAFT {n})", n + 1
    return title[:31]  # Excel's tab-name limit


# ---- the plan ---------------------------------------------------------------------

@dataclass
class _Line:
    number: str
    description: str
    amount: Decimal


@dataclass
class _Entry:
    payee: str
    lines: list[_Line]
    voucher: str = ""

    @property
    def total(self) -> Decimal:
        return sum((ln.amount for ln in self.lines), Decimal("0")).quantize(CENT)


def _plan_entries(docs: list, layout: ListingLayout) -> tuple[list[_Entry], int]:
    """Approved MYR invoices -> one entry per vendor, sorted by payee then
    invoice number. Returns (entries, how many non-MYR were left out)."""
    grouped: dict[str, _Entry] = {}
    non_myr = 0
    for d in docs:
        f = d.fields
        if str(f.get("currency", "")).upper() != "MYR":
            non_myr += 1
            continue
        vendor = str(f.get("vendor", "")).strip()
        payee = layout.payee_spelling(vendor) or vendor
        key = _norm(payee)
        number = _safe(f.get("invoice_number", ""))
        entry = grouped.setdefault(key, _Entry(payee=_safe(payee), lines=[]))
        entry.lines.append(_Line(
            number=number,
            description=_safe(f.get("description") or f"Invoice {number}")[:120],
            amount=money(f["amount"]),
        ))
    entries = sorted(grouped.values(), key=lambda e: e.payee.lower())
    for e in entries:
        e.lines.sort(key=lambda ln: ln.number)
    return entries, non_myr


# ---- writing one tab -------------------------------------------------------------

@dataclass
class _Written:
    """Where things landed, for the summary and the round-trip reading."""
    spans: list[EntrySpan]
    entry_cells: list[dict]
    fund_row: int | None = None
    charges_row: int | None = None
    total_row: int | None = None
    summary_first_row: int | None = None
    last_row: int = 0


def _write_tab(ws, layout: ListingLayout, entries: list[_Entry], month_name: str,
               figures: dict, settings: dict, formulas: bool) -> _Written:
    """Write the whole tab. formulas=False writes the computed values in
    the cells that would otherwise hold formulas — the twin the round trip
    reads, since openpyxl computes nothing."""
    c = layout.columns
    col = lambda role: getattr(c, role)  # noqa: E731
    text_col = c.description or c.payee   # where labels like 'Balance b/f' go
    has_bank = bool(c.balance and c.receipt)
    written = _Written(spans=[], entry_cells=[])

    def put(role_or_letter: str, row: int, value, number=False):
        letter = role_or_letter if len(role_or_letter) <= 3 and role_or_letter.isupper() \
            else col(role_or_letter)
        cell = ws[f"{letter}{row}"]
        cell.value = value if not isinstance(value, Decimal) else float(value)
        if number or isinstance(value, Decimal):
            cell.number_format = "#,##0.00"
        return cell.coordinate

    def fx(formula: str, value: Decimal):
        return formula if formulas else value

    for coord, value in layout.title_cells:
        ws[coord] = value
    for letter, width in layout.column_widths.items():
        ws.column_dimensions[letter].width = width
    for role, letter in c.model_dump().items():
        if letter and (src := figures["header_cells"].get(role)):
            ws[f"{letter}{layout.header_row}"] = src
    widest = max(column_index_from_string(l) for l in c.model_dump().values() if l)
    ws.cell(row=1, column=widest + 2, value=(
        "DRAFT — payment dates, voucher numbers and fund figures to be "
        "confirmed after the bank run. Generated by the AP checker."))

    row = layout.header_row + 1
    prev_bal: str | None = None
    if has_bank:
        put(text_col, row, "Balance b/f")
        prev_bal = put("balance", row, figures["opening"])
        written.spans.append(EntrySpan(first_row=row, last_row=row, kind="other"))
        row += 1
        put(text_col, row, f"Fund received for {month_name} payment")
        rc = put("receipt", row, figures["fund"])
        bal = money(figures["opening"] + figures["fund"])
        prev_bal = put("balance", row, fx(f"={prev_bal}+{rc}", bal), number=True)
        written.fund_row = row
        written.spans.append(EntrySpan(first_row=row, last_row=row, kind="other"))
        row += 2
        running = bal
    else:
        running = Decimal("0")

    first_entry_row = row
    for entry in entries:
        n = len(entry.lines)
        first, last = row, row + n - 1
        put("voucher_no", first, entry.voucher)
        put("payee", first, entry.payee)
        cells = {"row": first, "invoices": []}
        for i, ln in enumerate(entry.lines):
            put("invoice_no", first + i, ln.number)
            if c.description:
                put("description", first + i, ln.description)
            if n > 1 and c.line_amount:
                put("line_amount", first + i, ln.amount)
            cells["invoices"].append(first + i)
        if n > 1 and c.line_amount:
            la = c.line_amount
            cells["total"] = put("payment", first,
                                 fx(f"=SUM({la}{first}:{la}{last})", entry.total), number=True)
        else:
            cells["total"] = put("payment", first, entry.total)
        if has_bank:
            running = money(running - entry.total)
            prev_bal = put("balance", last, fx(f"={prev_bal}-{c.payment}{first}", running),
                           number=True)
        written.spans.append(EntrySpan(first_row=first, last_row=last, kind="payment"))
        written.entry_cells.append(cells)
        row = last + 2

    if has_bank:
        put(text_col, row, "Bank charges")
        put("payment", row, figures["charges"])
        running = money(running - figures["charges"])
        put("balance", row, fx(f"={prev_bal}-{c.payment}{row}", running), number=True)
        written.charges_row = row
        row += 1
        put(text_col, row, "Total")
        put("payment", row, fx(f"=SUM({c.payment}{first_entry_row}:{c.payment}{row - 1})",
                               money(figures["net"] + figures["charges"])), number=True)
        put("receipt", row, fx(f"=SUM({c.receipt}{layout.header_row + 1}:{c.receipt}{row - 1})",
                               figures["fund"]), number=True)
        written.total_row = row
        written.spans.append(EntrySpan(first_row=row - 1, last_row=row, kind="other"))
        row += 2
        written.summary_first_row = row
        label_col, value_col = c.date, c.invoice_no
        put(label_col, row, "Opening balance to utilise"); put(value_col, row, figures["opening"])
        put(label_col, row + 1, "Net payment"); net_cell = put(value_col, row + 1, figures["net"])
        put(label_col, row + 2, "Estimated bank charges")
        ch_cell = put(value_col, row + 2, figures["charges"])
        put(label_col, row + 3, "Total fund to request")
        put(value_col, row + 3, fx(f"={net_cell}+{ch_cell}", figures["fund"]), number=True)
        written.spans.append(EntrySpan(first_row=row, last_row=row + 3, kind="other"))
        row += 5

    put("date", row, "Prepared by:")
    put("voucher_no", row, _safe(settings.get("prepared_by", "")))
    put(text_col, row, "Reviewed by:")
    after = get_column_letter(column_index_from_string(text_col) + 1)
    put(after, row, _safe(settings.get("reviewed_by", "")))
    written.last_row = row
    return written


# ---- the whole draft ---------------------------------------------------------------

def build_draft(workbook: bytes, layout: ListingLayout, listing_rows: list[dict],
                docs: list, settings: dict, suffix: str = ".xlsx",
                today: date | None = None) -> Draft:
    """Append next month's draft tab to a copy of the workbook.

    workbook: the run's snapshot of the listing (bytes). listing_rows: every
    flattened past-payment row (for the voucher-collision check). docs:
    approved invoice documents. settings: prepared_by, reviewed_by,
    bank_charge (Decimal, per payment). Raises DraftError with the reason
    when nothing can be written.
    """
    entries, non_myr = _plan_entries(docs, layout)
    if not entries:
        raise DraftError("no approved MYR invoices to draft"
                         + (f" ({non_myr} non-MYR left out)" if non_myr else ""))

    anchor = layout.last_payment_date or (today or date.today())
    draft_month = _add_month(anchor) if layout.last_payment_date else anchor.replace(day=1)
    vouchers = next_vouchers(layout.last_voucher, layout.last_payment_date, draft_month,
                             len(entries))
    used = {str(r.get("no", "")).strip() for r in listing_rows} - {""}
    clash = [v for v in vouchers if v in used]
    if clash:
        raise DraftError(
            f"voucher number(s) {', '.join(clash)} already exist in the listing — "
            "the sequence cannot be continued without a duplicate; number by hand")
    for entry, voucher in zip(entries, vouchers):
        entry.voucher = voucher

    net = sum((e.total for e in entries), Decimal("0")).quantize(CENT)
    charges = money(money(settings.get("bank_charge", "0")) * len(entries))
    opening = layout.closing_balance if layout.closing_balance is not None else Decimal("0")
    fund = money(net + charges)
    figures = {"opening": opening, "net": net, "charges": charges, "fund": fund,
               "header_cells": {}}

    keep_vba = suffix.lower() == ".xlsm"
    wb = load_workbook(io.BytesIO(workbook), keep_vba=keep_vba)
    src = wb[layout.sheet]
    figures["header_cells"] = {
        role: src[f"{letter}{layout.header_row}"].value
        for role, letter in layout.columns.model_dump().items() if letter}
    title = draft_title(layout.sheet, draft_month, wb.sheetnames)
    month_name = draft_month.strftime("%B")
    ws = wb.create_sheet(title)
    written = _write_tab(ws, layout, entries, month_name, figures, settings, formulas=True)
    # header styles, so the tab looks like the client's
    for role, letter in layout.columns.model_dump().items():
        if letter:
            ws[f"{letter}{layout.header_row}"]._style = src[f"{letter}{layout.header_row}"]._style

    # Round trip through the reader's own audit and flattening (no AI).
    twin_wb = Workbook()
    twin = twin_wb.active
    twin.title = title
    _write_tab(twin, layout, entries, month_name, figures, settings, formulas=False)
    reading = SheetReading(is_payment_sheet=True, columns=layout.columns,
                           entries=written.spans, header_row=layout.header_row,
                           summary_first_row=written.summary_first_row, why="draft")
    problems = [t for _, t in audit_reading(twin, reading)]
    flat = flatten_reading(twin, reading)
    got = sorted((r["invoice_number"], f"{money(r['amount']):.2f}" if r["amount"] is not None else "?")
                 for r in flat)
    want = sorted((ln.number, f"{ln.amount:.2f}") for e in entries for ln in e.lines)
    if problems or got != want:
        raise DraftError(
            "the drafted tab did not read back as written: "
            + ("; ".join(problems) if problems else f"invoices {got} != {want}"))

    out = io.BytesIO()
    wb.save(out)
    summary = {
        "tab": title, "source_tab": layout.sheet, "month": draft_month.strftime("%Y-%m"),
        "entries": [
            {"voucher": e.voucher, "payee": e.payee, "total": f"{e.total:.2f}",
             "invoices": [{"number": ln.number, "description": ln.description,
                           "amount": f"{ln.amount:.2f}"} for ln in e.lines],
             "cells": {"row": cells["row"], "total": cells["total"]}}
            for e, cells in zip(entries, written.entry_cells)],
        "invoice_count": sum(len(e.lines) for e in entries),
        "excluded_non_myr": non_myr,
        "opening_balance": f"{opening:.2f}", "net_payment": f"{net:.2f}",
        "bank_charges": f"{charges:.2f}", "fund_to_request": f"{fund:.2f}",
        "closing_balance": f"{opening:.2f}",  # fund = payments + charges, by construction
        "prepared_by": _safe(settings.get("prepared_by", "")),
        "reviewed_by": _safe(settings.get("reviewed_by", "")),
        "rows_written": written.last_row,
        "has_bank_block": bool(layout.columns.balance and layout.columns.receipt),
    }
    return Draft(data=out.getvalue(), suffix=".xlsm" if keep_vba else ".xlsx", tab=title,
                 summary=summary,
                 round_trip={"audit_problems": problems, "invoices": got})
