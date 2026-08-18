"""Generate the synthetic demo batch.

Creates, under samples/generated/:
  batch/            the files a client would send (invoices, claims, receipts)
  demo_batch.zip    the same files zipped — what gets uploaded to the app
  reference/        payment listing, policy sheet, Maybank template (.xlsx)
  ground_truth.json what was planted where, so checks can be scored

Anomalies are planted deliberately (see GROUND_TRUTH below). Everything
else is clean. Run:  python samples/generate_samples.py
"""
from __future__ import annotations

import json
import shutil
import zipfile
from pathlib import Path

from PIL import Image, ImageDraw, ImageFilter, ImageFont
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent / "generated"
BATCH = OUT / "batch"
REF = OUT / "reference"

# ---------------------------------------------------------------- fixtures
# (vendor, invoice number, date, amount RM, paid_before, style)
# paid_before: the invoice already appears in a PAST tab of the payment
# listing, so the pipeline must raise ALREADY_PAID and point at the row.
# Every other invoice is new — the normal case, which raises no flag.
INVOICES = [
    # Anomaly: already paid on 2026-07-10 inside a grouped Maxis entry in
    # tab Jul'26 (two invoices, two line amounts, one payment).
    ("Maxis Bhd",              "MX-7101", "2026-07-03",  1240.00, True,  "pdf"),
    ("Tenaga Nasional Berhad", "TNB-5520", "2026-07-05", 3480.50, False, "pdf"),
    ("Syabas Water",           "SYB-2210", "2026-07-08",  412.30, False, "pdf"),
    ("KL Office Supplies",     "KLO-0091", "2026-07-11",  867.20, False, "photo"),
    ("Securemax Guards Sdn Bhd", "SMG-4415", "2026-07-15", 5200.00, False, "pdf"),
    ("CleanPro Services",      "CP-3302", "2026-07-18",  980.00, False, "photo_blurry"),
    # Anomaly: 7 months old (OLD_DATED) AND already paid in tab Jun'26.
    ("Maxis Bhd",              "MX-2214", "2026-01-12",  1240.00, True,  "pdf"),
    ("Apex Renovation Works",  "ARW-0808", "2026-07-22", 2750.00, False, "pdf"),
]

# The client's payment listing, in the client's own layout (modelled on
# the ICMR file): one tab per month of PAST payments. Each entry is
# (payee, [(invoice number, description, line amount), ...]); a single-
# invoice entry writes its amount in the payment column only, a grouped
# entry writes per-line amounts (column F, which — as in the real file —
# has no header) and one payment total.
LISTING_TABS = [
    ("Jun'26", "2026-06-23", "June", [
        ("Maxis Bhd", [("MX-2214", "Mobile lines - Jan 2026", 1240.00)]),
        ("Tenaga Nasional Berhad", [("TNB-5100", "Electricity - May 2026", 3390.10)]),
        ("CleanPro Services", [("CP-3210", "Office cleaning - May 2026", 980.00),
                               ("CP-3250", "Carpet shampoo", 150.00)]),
        ("Syabas Water", [("SYB-2105", "Water - May 2026", 388.40)]),
    ]),
    ("Jul'26", "2026-07-10", "July", [
        ("Maxis Bhd", [("MX-7101", "Mobile lines - Jun 2026", 1240.00),
                       ("MX-7050", "Broadband - Jun 2026", 310.00)]),
        ("Tenaga Nasional Berhad", [("TNB-5310", "Electricity - Jun 2026", 3412.75)]),
        ("CleanPro Services", [("CP-3260", "Office cleaning - Jun 2026", 980.00)]),
    ]),
]
LISTING_OPENING_BALANCE = 7.90     # the small residual the account carries
LISTING_BANK_CHARGE = 0.10         # per payment, as the client estimates it

# (claimant, expense description, amount, currency, receipt vendor)
CLAIMS = [
    ("Tan W.L.",  "Home wi-fi subscription — July",  95.00, "USD", "TIME Internet"),   # over USD 80 cap
    ("A. Rahman", "Grab to client office (2 trips)", 64.00, "MYR", "Grab Malaysia"),   # clean
    ("S. Priya",  "Team lunch after stocktake",     186.00, "MYR", "Restoran Selera"), # ambiguous category
    ("J. Lim",    "Stationery for client files",     38.50, "MYR", "MPH Bookstores"),  # clean
]

GROUND_TRUTH = {
    "documents": {},   # filename -> {"kind": ..., "fields": {...}, "paid_before": bool}
    "expected_flags": [
        {"code": "OLD_DATED",        "match": "MX-2214",  "why": "invoice dated 2026-01-12, 7 months old"},
        {"code": "ALREADY_PAID",     "match": "MX-2214",  "why": "paid in tab Jun'26 (single-invoice entry)"},
        {"code": "ALREADY_PAID",     "match": "MX-7101",  "why": "paid in tab Jul'26 inside a grouped Maxis entry"},
        {"code": "OVER_CAP",         "match": "Tan W.L.", "why": "wi-fi USD 95 exceeds USD 80 cap (policy 4.2)"},
        {"code": "AMBIGUOUS_CATEGORY", "match": "S. Priya", "why": "team lunch: staff welfare vs client entertainment (policy 5.1)"},
        {"code": "LOW_CONFIDENCE",   "match": "CP-3302",  "why": "deliberately blurry scan"},
        {"code": "UNCLASSIFIED",     "match": "memo",     "why": "an office memo is not an AP document; must be surfaced, not dropped"},
    ],
}

FONT_BIG = ImageFont.load_default(size=34)
FONT = ImageFont.load_default(size=22)
FONT_SMALL = ImageFont.load_default(size=16)


# ---------------------------------------------------------------- drawing
def _invoice_image(vendor: str, number: str, date: str, amount: float) -> Image.Image:
    """Draw a simple but realistic one-page invoice."""
    img = Image.new("RGB", (900, 1200), "white")
    d = ImageDraw.Draw(img)
    d.text((60, 60), vendor, font=FONT_BIG, fill="black")
    d.text((60, 120), "TAX INVOICE", font=FONT, fill="black")
    d.line((60, 160, 840, 160), fill="black", width=2)
    d.text((60, 200), f"Invoice No : {number}", font=FONT, fill="black")
    d.text((60, 240), f"Date       : {date}", font=FONT, fill="black")
    d.text((60, 280), "Bill To    : Client ABC Sdn Bhd", font=FONT, fill="black")
    d.line((60, 340, 840, 340), fill="black", width=1)
    d.text((60, 380), "Description                              Amount (RM)", font=FONT_SMALL, fill="black")
    d.text((60, 420), f"Monthly services - July 2026             {amount:,.2f}", font=FONT, fill="black")
    d.line((60, 900, 840, 900), fill="black", width=1)
    d.text((520, 940), f"TOTAL (RM): {amount:,.2f}", font=FONT_BIG, fill="black")
    d.text((60, 1100), "Payment due within 30 days.", font=FONT_SMALL, fill="black")
    return img


def _photo_effect(img: Image.Image, blurry: bool) -> Image.Image:
    """Make a clean render look like a phone photo (slight rotate, grey bg)."""
    img = img.rotate(2, expand=True, fillcolor="#d8d4cc")
    if blurry:
        # Strong enough that some digits are genuinely uncertain — the point
        # is to exercise the low-confidence flag, not to be decorative.
        img = img.filter(ImageFilter.GaussianBlur(radius=3.2))
    return img


def _claim_form(claimant: str, desc: str, amount: float, currency: str) -> Image.Image:
    img = Image.new("RGB", (900, 700), "white")
    d = ImageDraw.Draw(img)
    d.text((60, 50), "STAFF EXPENSE CLAIM FORM", font=FONT_BIG, fill="black")
    d.line((60, 110, 840, 110), fill="black", width=2)
    d.text((60, 150), f"Employee   : {claimant}", font=FONT, fill="black")
    d.text((60, 190), f"Period     : July 2026", font=FONT, fill="black")
    # The default font has no glyph for the em-dash, so strip it out of
    # anything we draw — otherwise the AI reads a missing-character box.
    d.text((60, 230), f"Description: {desc.replace(chr(0x2014), '-')}", font=FONT, fill="black")
    d.text((60, 270), f"Amount     : {currency} {amount:,.2f}", font=FONT, fill="black")
    d.text((60, 310), "Receipts   : attached", font=FONT, fill="black")
    d.text((60, 420), "Employee signature: ______________", font=FONT_SMALL, fill="black")
    return img


def _receipt(vendor: str, amount: float, currency: str) -> Image.Image:
    img = Image.new("RGB", (500, 700), "#fbfaf6")
    d = ImageDraw.Draw(img)
    d.text((40, 40), vendor, font=FONT, fill="black")
    d.text((40, 90), "*** RECEIPT ***", font=FONT_SMALL, fill="black")
    d.text((40, 140), "Date: July 2026", font=FONT_SMALL, fill="black")
    d.text((40, 200), f"TOTAL: {currency} {amount:,.2f}", font=FONT, fill="black")
    d.text((40, 260), "Thank you!", font=FONT_SMALL, fill="black")
    return _photo_effect(img, blurry=False)


# ---------------------------------------------------------------- workbooks
def _listing_tab(ws, title: str, pay_date: str, month: str, entries: list,
                 opening: float) -> float:
    """Write one monthly tab in the client's layout; return its closing
    balance. Numbers are written as VALUES (openpyxl computes nothing, and
    a formula with no saved result is exactly the stale-cache case the
    reader warns about); the client's real file carries Excel's cached
    values, so the reader sees the same thing either way."""
    ws.title = title
    ws["A1"] = "Name:"; ws["B1"] = "Client ABC Sdn Bhd"
    ws["A2"] = "A/C No:"; ws["B2"] = "514712417644"
    for col, head in zip("ABCDEFGHI", ["Date", "Cheque/Journal No.",
                                       "Invoice / Reference No.", "Payee Name",
                                       "Description", None,  # F: line amounts, unlabelled
                                       "Payment (MYR)", "Balance (MYR)",
                                       "Receipt (MYR)"]):
        if head:
            ws[f"{col}4"] = head
    net = round(sum(a for _, lines in entries for _, _, a in lines), 2)
    charges = round(LISTING_BANK_CHARGE * len(entries), 2)
    fund = round(net + charges, 2)          # so the balance returns to the residual
    balance = opening
    ws["E5"] = "Balance b/f"; ws["H5"] = balance
    balance = round(balance + fund, 2)
    ws["A6"] = pay_date; ws["E6"] = f"Fund received for {month} payment"
    ws["I6"] = fund; ws["H6"] = balance
    row = 8
    mm_yy = f"{pay_date[5:7]}{pay_date[2:4]}"
    for n, (payee, lines) in enumerate(entries, 1):
        total = round(sum(a for _, _, a in lines), 2)
        ws[f"A{row}"] = pay_date; ws[f"B{row}"] = f"PV{mm_yy}/{n:02d}"
        ws[f"D{row}"] = payee; ws[f"G{row}"] = total
        for i, (number, desc, amount) in enumerate(lines):
            ws[f"C{row + i}"] = number; ws[f"E{row + i}"] = desc
            if len(lines) > 1:
                ws[f"F{row + i}"] = amount
        balance = round(balance - total, 2)
        ws[f"H{row + len(lines) - 1}"] = balance
        row += len(lines) + 1                # one blank row between entries
    ws[f"E{row}"] = "Bank charges"; ws[f"G{row}"] = charges
    balance = round(balance - charges, 2); ws[f"H{row}"] = balance
    row += 1
    ws[f"E{row}"] = "Total"; ws[f"G{row}"] = round(net + charges, 2); ws[f"I{row}"] = fund
    row += 2
    for label, value in [("Opening balance to utilise", opening),
                         ("Net payment", net),
                         ("Estimated bank charges", charges),
                         ("Total fund to request", fund)]:
        ws[f"A{row}"] = label; ws[f"C{row}"] = value
        row += 1
    row += 1
    ws[f"A{row}"] = "Prepared by:"; ws[f"B{row}"] = "W. Chen"
    ws[f"E{row}"] = "Reviewed by:"; ws[f"F{row}"] = "A. Rahman"
    return balance


def _payment_listing() -> None:
    """The client's listing of PAST payments — a cover tab and one tab per
    month, in the client's own layout (title block, headers on row 4,
    grouped entries with an unlabelled line-amount column, balance b/f,
    fund received, bank charges, totals, summary block, signatures)."""
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Client ABC Sdn Bhd — FY2026 Payment Listing"
    cover["A3"] = "One tab per month. Prepared by the AP team."
    balance = LISTING_OPENING_BALANCE
    for title, pay_date, month, entries in LISTING_TABS:
        balance = _listing_tab(wb.create_sheet(), title, pay_date, month, entries, balance)
    wb.save(REF / "payment_listing.xlsx")


def _policy_sheet() -> None:
    """Client ABC's expense policy: numeric caps AND wordy clauses, because
    real SOPs mix both — that mix is what the AI judgment stage must handle."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Client ABC Policy"
    ws.append(["Clause", "Category", "Cap", "Currency", "Policy text"])
    ws.append(["4.1", "Transport", 200, "MYR",
               "Work-related taxi/e-hailing reimbursable up to RM200 per month per staff."])
    ws.append(["4.2", "Telecoms", 80, "USD",
               "Internet subsidy: home wi-fi reimbursable up to USD 80 per month."])
    ws.append(["5.1", "Meals", 50, "MYR",
               "Staff welfare meals up to RM50 per head. Client entertainment "
               "is NOT staff welfare and requires prior partner approval."])
    ws.append(["5.2", "Office", 100, "MYR",
               "Reasonable stationery and small office items for client work."])
    wb.save(REF / "policy_sheet.xlsx")


def _maybank_template() -> None:
    """Empty bank-upload template. The output stage must learn this column
    layout and produce rows that paste straight into it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "IBG Upload"
    ws.append(["Payment Type", "Beneficiary Name", "Beneficiary Account",
               "Bank Code", "Amount (RM)", "Payment Reference", "IG Code"])
    wb.save(REF / "maybank_template.xlsx")


# ---------------------------------------------------------------- main
def main() -> None:
    if OUT.exists():
        shutil.rmtree(OUT)  # regenerate from scratch every time
    BATCH.mkdir(parents=True)
    REF.mkdir(parents=True)

    for vendor, number, date, amount, paid_before, style in INVOICES:
        img = _invoice_image(vendor, number, date, amount)
        if style == "pdf":
            name = f"invoice_{number}.pdf"
            img.save(BATCH / name)
        else:
            name = f"invoice_{number}.png"
            img = _photo_effect(img, blurry=(style == "photo_blurry"))
            img.save(BATCH / name)
        GROUND_TRUTH["documents"][name] = {
            "kind": "invoice",
            "fields": {"vendor": vendor, "invoice_number": number,
                       "date": date, "amount": amount, "currency": "MYR"},
            "paid_before": paid_before,
        }

    for i, (claimant, desc, amount, currency, shop) in enumerate(CLAIMS, 1):
        cname = f"claim_{i:02d}.png"
        rname = f"claim_{i:02d}_receipt.png"
        _claim_form(claimant, desc, amount, currency).save(BATCH / cname)
        _receipt(shop, amount, currency).save(BATCH / rname)
        GROUND_TRUTH["documents"][cname] = {
            "kind": "claim",
            "fields": {"claimant": claimant, "description": desc,
                       "amount": amount, "currency": currency},
        }
        GROUND_TRUTH["documents"][rname] = {"kind": "receipt", "fields": {"vendor": shop}}

    # One document that is none of the three kinds — the pipeline must
    # surface it as UNCLASSIFIED, never silently drop it.
    memo = Image.new("RGB", (900, 500), "white")
    d = ImageDraw.Draw(memo)
    d.text((60, 60), "INTERNAL MEMO", font=FONT_BIG, fill="black")
    d.text((60, 140), "Reminder: office closed this Friday", font=FONT, fill="black")
    d.text((60, 180), "for the building maintenance day.", font=FONT, fill="black")
    memo.save(BATCH / "memo.png")
    GROUND_TRUTH["documents"]["memo.png"] = {"kind": "unknown", "fields": {}}

    _payment_listing()
    _policy_sheet()
    _maybank_template()

    with zipfile.ZipFile(OUT / "demo_batch.zip", "w") as z:
        for f in sorted(BATCH.iterdir()):
            z.write(f, f.name)

    (OUT / "ground_truth.json").write_text(json.dumps(GROUND_TRUTH, indent=2))
    n_docs = len(GROUND_TRUTH["documents"])
    print(f"Generated {n_docs} batch documents, 3 reference workbooks, "
          f"{len(GROUND_TRUTH['expected_flags'])} planted anomalies → {OUT}")


if __name__ == "__main__":
    main()
