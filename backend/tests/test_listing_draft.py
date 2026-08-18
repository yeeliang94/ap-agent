"""Drafting next month's payment-listing entries in the client's own layout.

The reader learns a ListingLayout from the latest tab (header row, column
map, closing balance, last voucher, the client's payee spellings); the
writer appends ONE new tab to a COPY of the workbook — never the live
listing — with the approved invoices grouped one payment per vendor, and
round-trips the result through the reader's own audit before delivering
it. Everything here is deterministic: the only AI text is the per-invoice
description read off the invoice itself.
"""
from __future__ import annotations

import io
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from app.pipeline import listing_agent, listing_draft, listing_layout
from app.pipeline.listing_agent import (
    ColumnRoles, EntrySpan, SheetReading, audit_reading, flatten_reading,
)


# ---- a two-tab workbook shaped like the client's (from the sample builder) ---

def _workbook():
    from samples import generate_samples as gen
    wb = Workbook()
    balance = gen.LISTING_OPENING_BALANCE
    for i, (title, pay_date, month, entries) in enumerate(gen.LISTING_TABS):
        ws = wb.active if i == 0 else wb.create_sheet()
        balance = gen._listing_tab(ws, title, pay_date, month, entries, balance)
    return wb


COLS = ColumnRoles(date="A", voucher_no="B", invoice_no="C", payee="D",
                   description="E", line_amount="F", payment="G",
                   balance="H", receipt="I")


def _reading(entries) -> SheetReading:
    """The correct structural answer for a tab built by the sample builder."""
    spans = [EntrySpan(first_row=5, last_row=5, kind="other"),
             EntrySpan(first_row=6, last_row=6, kind="other")]
    row = 8
    for _, lines in entries:
        spans.append(EntrySpan(first_row=row, last_row=row + len(lines) - 1, kind="payment"))
        row += len(lines) + 1
    spans.append(EntrySpan(first_row=row, last_row=row + 1, kind="other"))
    spans.append(EntrySpan(first_row=row + 3, last_row=row + 6, kind="other"))
    return SheetReading(is_payment_sheet=True, columns=COLS, entries=spans,
                        header_row=4, summary_first_row=row + 3, why="t")


def _jul():
    from samples import generate_samples as gen
    wb = _workbook()
    ws = wb["Jul'26"]
    entries = next(e for t, _, _, e in gen.LISTING_TABS if t == "Jul'26")
    return wb, ws, _reading(entries)


class _Doc:
    def __init__(self, doc_id, vendor, number, amount, description="", currency="MYR"):
        self.id, self.kind, self.status = doc_id, "invoice", "checked"
        self.filename = f"{doc_id}.pdf"
        self.fields = {"vendor": vendor, "invoice_number": number, "amount": amount,
                       "currency": currency, "date": "2026-07-20",
                       "description": description}
        self.confidence, self.error, self.parent_id = {}, "", None


SETTINGS = {"prepared_by": "W. Chen", "reviewed_by": "A. Rahman",
            "bank_charge": Decimal("0.10")}


# ---- the layout the reader learns --------------------------------------------

def test_layout_is_learned_from_the_latest_tab():
    wb, ws, reading = _jul()
    assert audit_reading(ws, reading) == []
    layout = listing_layout.learn_layout(ws, reading)
    assert layout.sheet == "Jul'26" and layout.header_row == 4
    assert layout.columns.payment == "G" and layout.columns.line_amount == "F"
    assert layout.first_entry_row == 5
    assert layout.summary_first_row == 18
    assert layout.closing_balance == Decimal("7.90")      # after bank charges
    assert layout.last_voucher == "PV0726/03"
    assert str(layout.last_payment_date) == "2026-07-10"
    # the client's own spelling of each payee, keyed tolerantly
    assert layout.payee_spelling("maxis bhd") == "Maxis Bhd"
    assert layout.payee_spelling("Tenaga Nasional") == "Tenaga Nasional Berhad"
    assert layout.payee_spelling("Nobody") is None
    # the title block above the header, verbatim
    assert ("B1", "Client ABC Sdn Bhd") in layout.title_cells
    # and it survives a round trip through JSON (it is cached with the rows)
    again = listing_layout.ListingLayout.from_dict(layout.to_dict())
    assert again == layout


def test_layout_needs_a_header_row_and_the_writing_columns():
    wb, ws, reading = _jul()
    no_header = reading.model_copy(update={"header_row": None})
    with pytest.raises(listing_layout.LayoutIncomplete, match="header row"):
        listing_layout.learn_layout(ws, no_header)
    no_voucher = reading.model_copy(update={"columns": COLS.model_copy(update={"voucher_no": None})})
    with pytest.raises(listing_layout.LayoutIncomplete, match="voucher"):
        listing_layout.learn_layout(ws, no_voucher)


def test_impossible_header_or_summary_rows_are_audit_problems():
    """A header row inside the entries, or a summary block that starts
    before the last payment, is a wrong reading — fed back to the AI like
    any other structural problem, not quietly accepted."""
    wb, ws, reading = _jul()
    bad = reading.model_copy(update={"header_row": 9})
    assert any("header_row" in t for _, t in audit_reading(ws, bad))
    bad = reading.model_copy(update={"summary_first_row": 10})
    assert any("summary_first_row" in t for _, t in audit_reading(ws, bad))


@pytest.mark.asyncio
async def test_ingest_picks_the_latest_tab_for_the_layout(monkeypatch):
    from samples import generate_samples as gen
    wb = _workbook()
    readings = {t: _reading(e) for t, _, _, e in gen.LISTING_TABS}

    class _Agent:
        async def run(self, prompt, **kw):
            title = prompt.split("\n", 1)[0]          # "Sheet name: 'Jun'26'"
            out = next(r for t, r in readings.items() if repr(t) in title)

            class R:
                output = out
            return R()
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: _Agent())
    result = await listing_agent.ingest_workbook(wb)
    assert result.layout is not None and result.layout.sheet == "Jul'26"
    assert result.layout.last_voucher == "PV0726/03"


# ---- voucher numbering ----------------------------------------------------------

def test_voucher_numbers_roll_the_month_code_and_restart():
    from datetime import date
    nxt = listing_draft.next_vouchers("PV0726/03", date(2026, 7, 10), date(2026, 8, 1), 3)
    assert nxt == ["PV0826/01", "PV0826/02", "PV0826/03"]
    # same month as the latest tab: continue the sequence
    assert listing_draft.next_vouchers("PV0726/03", date(2026, 7, 10), date(2026, 7, 1), 2) \
        == ["PV0726/04", "PV0726/05"]
    # no month code in the number: just continue, width kept
    assert listing_draft.next_vouchers("CHQ-000118", date(2026, 7, 10), date(2026, 8, 1), 2) \
        == ["CHQ-000119", "CHQ-000120"]
    # nothing to continue from
    assert listing_draft.next_vouchers(None, None, date(2026, 8, 1), 2) == ["", ""]


def test_draft_tab_title_follows_the_client_pattern():
    from datetime import date
    assert listing_draft.draft_title("Jul'26", date(2026, 8, 1), []) == "Aug'26 (DRAFT)"
    assert listing_draft.draft_title("July 2026", date(2026, 8, 1), []) == "August 2026 (DRAFT)"
    assert listing_draft.draft_title("Payments", date(2026, 8, 1), []) == "Payments Aug'26 (DRAFT)"
    assert listing_draft.draft_title("Jul'26", date(2026, 8, 1), ["Aug'26 (DRAFT)"]) \
        == "Aug'26 (DRAFT 2)"


# ---- the writer ----------------------------------------------------------------

def _draft(docs, wb=None, settings=SETTINGS):
    wb, ws, reading = _jul() if wb is None else (wb, wb["Jul'26"], _jul()[2])
    layout = listing_layout.learn_layout(ws, reading)
    buf = io.BytesIO(); wb.save(buf)
    used = _vouchers(ws, reading)
    return listing_draft.build_draft(buf.getvalue(), layout, used, docs,
                                     settings, suffix=".xlsx")


def _vouchers(ws, reading):
    from app.pipeline.listing_agent import _texts
    return [v for s in reading.entries if s.kind == "payment"
            for v in _texts(ws, reading.columns.voucher_no, s)]


def test_draft_appends_one_tab_in_the_client_layout_and_round_trips():
    docs = [
        _Doc("a", "Tenaga Nasional Berhad", "TNB-5520", 3480.50, "Electricity - Jul 2026"),
        _Doc("b", "Maxis Bhd", "MX-7201", 1240.00, "Mobile lines - Jul 2026"),
        _Doc("c", "Maxis Bhd", "MX-7150", 310.00, "Broadband - Jul 2026"),
        _Doc("d", "Apex Renovation Works", "ARW-0808", 2750.00, ""),
    ]
    draft = _draft(docs)
    wb = load_workbook(io.BytesIO(draft.data))
    # the original tabs are untouched; exactly one new tab is appended last
    assert wb.sheetnames == ["Jun'26", "Jul'26", "Aug'26 (DRAFT)"]
    ws = wb["Aug'26 (DRAFT)"]
    # title block and headers in the client's own columns
    assert ws["A1"].value == "Name:" and ws["B1"].value == "Client ABC Sdn Bhd"
    assert ws["A4"].value == "Date" and ws["G4"].value == "Payment (MYR)"
    # one payment per vendor, sorted by payee, PV numbers continuing Jul'26
    s = draft.summary
    assert [e["payee"] for e in s["entries"]] == ["Apex Renovation Works", "Maxis Bhd",
                                                  "Tenaga Nasional Berhad"]
    assert [e["voucher"] for e in s["entries"]] == ["PV0826/01", "PV0826/02", "PV0826/03"]
    maxis = s["entries"][1]
    assert [i["number"] for i in maxis["invoices"]] == ["MX-7150", "MX-7201"]
    assert maxis["total"] == "1550.00"
    # the money, in Decimal, as strings for JSON: fund = payments + charges,
    # so the balance returns to the same residual
    assert s["opening_balance"] == "7.90" and s["net_payment"] == "7780.50"
    assert s["bank_charges"] == "0.30" and s["fund_to_request"] == "7780.80"
    assert s["closing_balance"] == "7.90"
    # AI-drafted text only where it belongs; a missing description falls back
    apex = s["entries"][0]["invoices"][0]
    assert apex["description"] == "Invoice ARW-0808"
    # balances and grouped totals are formulas in the delivered file
    assert str(ws[maxis["cells"]["total"]].value).startswith("=SUM(")
    assert str(ws["H6"].value).startswith("=")
    # the round trip: the values twin passes the reader's own audit and
    # flattens to exactly the approved invoices
    assert draft.round_trip["audit_problems"] == []
    assert draft.round_trip["invoices"] == sorted(
        [("ARW-0808", "2750.00"), ("MX-7150", "310.00"), ("MX-7201", "1240.00"),
         ("TNB-5520", "3480.50")])
    assert draft.summary["prepared_by"] == "W. Chen"


def test_draft_uses_the_listing_spelling_of_a_known_payee():
    docs = [_Doc("a", "TENAGA NASIONAL BHD", "TNB-5520", 100.0)]
    # not tolerant enough to match "Tenaga Nasional Berhad" -> keeps the invoice's text
    d1 = _draft(docs)
    assert d1.summary["entries"][0]["payee"] == "TENAGA NASIONAL BHD"
    d2 = _draft([_Doc("a", "Tenaga Nasional", "TNB-5520", 100.0)])
    assert d2.summary["entries"][0]["payee"] == "Tenaga Nasional Berhad"


def test_draft_excludes_non_myr_and_says_so():
    docs = [_Doc("a", "Acme", "A-1", 10.0), _Doc("b", "Acme", "A-2", 5.0, currency="USD")]
    d = _draft(docs)
    assert d.summary["excluded_non_myr"] == 1
    assert [i["number"] for i in d.summary["entries"][0]["invoices"]] == ["A-1"]


def test_draft_refuses_a_voucher_collision():
    """A generated PV number that already exists anywhere in the listing is
    a duplicate voucher waiting to happen: refuse, do not renumber."""
    docs = [_Doc("a", "Acme", "A-1", 10.0)]
    wb, ws, reading = _jul()
    ws["B11"] = "PV0826/01"          # the client already used next month's first number
    with pytest.raises(listing_draft.DraftError, match="PV0826/01"):
        _draft(docs, wb)


def test_layout_without_line_amount_column_writes_one_entry_per_invoice():
    """Without a per-line amount column the sheet cannot show two invoices'
    amounts inside one payment, so each invoice becomes its own entry —
    the round trip would otherwise (rightly) refuse the draft."""
    wb, ws, reading = _jul()
    ws.delete_cols(6)  # drop F: the line-amount column (payment now G-1 ...)
    # rebuild a reading for the shifted sheet: F=payment, G=balance, H=receipt
    cols = COLS.model_copy(update={"line_amount": None, "payment": "F",
                                   "balance": "G", "receipt": "H"})
    reading = reading.model_copy(update={"columns": cols})
    layout = listing_layout.learn_layout(ws, reading)
    buf = io.BytesIO(); wb.save(buf)
    docs = [_Doc("a", "Maxis Bhd", "MX-1", 10.0), _Doc("b", "Maxis Bhd", "MX-2", 5.0)]
    d = listing_draft.build_draft(buf.getvalue(), layout, [], docs, SETTINGS)
    assert [e["payee"] for e in d.summary["entries"]] == ["Maxis Bhd", "Maxis Bhd"]
    assert d.round_trip["audit_problems"] == []


def test_batch_spellings_of_one_vendor_are_one_payment():
    docs = [_Doc("a", "Apex Renovation Works", "A-1", 10.0),
            _Doc("b", "APEX RENOVATION WORKS SDN BHD", "A-2", 5.0)]
    d = _draft(docs)
    assert len(d.summary["entries"]) == 1
    assert [i["number"] for i in d.summary["entries"][0]["invoices"]] == ["A-1", "A-2"]


def test_client_text_starting_with_equals_stays_text():
    docs = [_Doc("a", "=Odd Vendor", "=INV-1", 10.0, "=see note")]
    d = _draft(docs)
    ws = load_workbook(io.BytesIO(d.data))[d.tab]
    row = d.summary["entries"][0]["cells"]["row"]
    assert ws[f"C{row}"].value == "=INV-1" and ws[f"C{row}"].data_type == "s"
    assert ws[f"D{row}"].value == "=Odd Vendor" and ws[f"D{row}"].data_type == "s"
    assert ws[f"H{row}"].data_type == "f"      # our balance formula is still a formula


def test_voucher_collision_is_caught_even_on_a_reference_less_entry():
    """A recurring payment with no invoice number yields no flat row but
    still owns its voucher number; the guard must see it."""
    wb, ws, reading = _jul()
    layout = listing_layout.learn_layout(ws, reading)
    buf = io.BytesIO(); wb.save(buf)
    with pytest.raises(listing_draft.DraftError, match="PV0826/01"):
        listing_draft.build_draft(buf.getvalue(), layout, ["PV0826/01"],
                                  [_Doc("a", "Acme", "A-1", 10.0)], SETTINGS)


def test_long_tab_titles_stay_unique_within_excel_limits():
    from datetime import date
    long = "A very long payments tab name indeed"
    t1 = listing_draft.draft_title(long, date(2026, 8, 1), [])
    t2 = listing_draft.draft_title(long, date(2026, 8, 1), [t1])
    assert len(t1) <= 31 and len(t2) <= 31 and t1 != t2


def test_draft_is_deterministic():
    docs = [_Doc("a", "Acme", "A-1", 10.0), _Doc("b", "Zed", "Z-9", 20.0)]
    assert _draft(docs).summary == _draft(list(reversed(docs))).summary


def test_draft_with_nothing_approved_is_skipped():
    with pytest.raises(listing_draft.DraftError, match="no approved"):
        _draft([])


@pytest.mark.asyncio
async def test_outputs_carry_the_draft_and_write_the_file(monkeypatch, tmp_path):
    """Through build_outputs: the run's reference copy is read once (AI
    scripted), the draft is written beside it, and outputs say what it
    holds. When nothing is approved, outputs say why there is no draft."""
    import json
    from samples import generate_samples as gen
    from app.pipeline import output, reference

    wb = _workbook()
    refs = tmp_path / "run" / "reference"; refs.mkdir(parents=True)
    wb.save(refs / "payment_listing.xlsx")
    (refs / reference.MANIFEST).write_text(json.dumps(
        {"payment_listing": "payment_listing.xlsx", "policy_sheet": None, "bank_template": None}))
    readings = {t: _reading(e) for t, _, _, e in gen.LISTING_TABS}

    class _Agent:
        async def run(self, prompt, **kw):
            title = prompt.split("\n", 1)[0]
            out = next(r for t, r in readings.items() if repr(t) in title)

            class R:
                output = out
            return R()
    monkeypatch.setattr(listing_agent, "create_agent", lambda *a, **k: _Agent())
    monkeypatch.setattr("app.settings_store.draft_settings", lambda: SETTINGS)

    docs = [_Doc("a", "Maxis Bhd", "MX-7201", 1240.00, "Mobile lines - Jul 2026"),
            _Doc("b", "Nobody New", "NN-1", 10.00)]
    res = await output.build_outputs(docs, excluded_doc_ids=set(), refs=refs)
    d = res["listing_draft"]
    assert d["tab"] == "Aug'26 (DRAFT)" and d["invoice_count"] == 2
    assert d["file"].endswith(".xlsx") and (output.draft_dir(refs) / d["file"]).is_file()
    back = load_workbook(output.draft_dir(refs) / d["file"])
    assert back.sheetnames[-1] == "Aug'26 (DRAFT)"
    assert res["bank_skipped"] is True   # no template in this folder, as before

    # everything excluded -> no draft, and the stale file is gone
    res = await output.build_outputs(docs, excluded_doc_ids={"a", "b"}, refs=refs)
    assert "skipped" in res["listing_draft"] and "no approved" in res["listing_draft"]["skipped"]
    assert "file" not in res["listing_draft"]


def test_draft_settings_are_validated(monkeypatch):
    from fastapi import HTTPException
    from app import routes, settings_store as store
    saved: dict = {}
    monkeypatch.setattr(store, "set_settings", lambda values: saved.update(values))
    monkeypatch.setattr(store, "get_setting", lambda key: saved.get(key, store.DEFAULTS[key]))
    base = {"client_name": "Client ABC", "sharepoint_folder_url": "https://x/AP"}
    with pytest.raises(HTTPException):
        routes.update_settings({**base, "draft_bank_charge": "ten sen"})
    with pytest.raises(HTTPException):
        routes.update_settings({**base, "draft_prepared_by": "x" * 81})
    out = routes.update_settings({**base, "draft_bank_charge": "0.1", "draft_prepared_by": " W. Chen "})
    assert out["draft_bank_charge"] == "0.10" and out["draft_prepared_by"] == "W. Chen"


def test_draft_keeps_macros_for_xlsm():
    """A .xlsm copy is written as .xlsm with keep_vba, never silently
    downgraded to .xlsx (which would strip the client's macros)."""
    docs = [_Doc("a", "Acme", "A-1", 10.0)]
    wb, ws, reading = _jul()
    layout = listing_layout.learn_layout(ws, reading)
    buf = io.BytesIO(); wb.save(buf)
    d = listing_draft.build_draft(buf.getvalue(), layout, _vouchers(ws, reading),
                                  docs, SETTINGS, suffix=".xlsm")
    assert d.suffix == ".xlsm"
    back = load_workbook(io.BytesIO(d.data), keep_vba=True)
    assert back["Aug'26 (DRAFT)"]["C8"].value == "A-1"
