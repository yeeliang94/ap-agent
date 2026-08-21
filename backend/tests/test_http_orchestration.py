"""The HTTP / orchestration half of the five-axis review (2026-08-19).

What is pinned here is not domain logic but the plumbing every route sits
on, each item a defect the review found:

  #13 a background stage really is dispatched from a SYNC route (the
      threadpool has no running loop; the loop is remembered at startup),
      is kept referenced, and its death fails the run instead of leaving it
      in an in-progress status for ever
  #14 the revision check is a COMPARE-AND-SET: two requests carrying the
      same expected_revision, exactly one wins
  #15 a throw inside `worker.retry_employee` fails the run rather than
      leaving it `verifying`
   #9 a cancel reaches a worker already inside `_work`: it stops between
      stages and commits nothing
  #16 a malformed body is a 422 with a field path, not a 500

Feature-flag gating, run-status guards and the `/file` errors are pinned
alongside them.
"""
from __future__ import annotations

import sys
import threading

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import config, main, settings_store
from app.claims import profile, routes as claims_routes, runner, worker
from app.claims.models import ClaimCase, ClaimEmployee, ClaimEvidence, ClaimRow, ClaimSourceArtifact, ClaimsRun
from app.db import Base
from app.main import app

from . import claims_scripted as scripted
from .test_claims_baseline import client, run_client_a

needs_sample = pytest.mark.skipif(not scripted.GEN.is_dir(),
                                  reason="run samples/generate_claims_sample.py first")

# Background stages are stubbed out for the SETUP of a test and switched on
# for the one action under examination, so a test drives verification itself
# except where the dispatch is the thing being tested.
_LIVE = {"on": False}


@pytest.fixture()
def db(tmp_path, monkeypatch):
    """test_claims_baseline's fixture, but with a start_background that can
    be switched to the REAL one — the review's #13 is precisely that every
    test stubbed it, so the dispatch itself was never exercised."""
    engine = create_engine(f"sqlite:///{tmp_path / 'test.sqlite3'}",
                           connect_args={"check_same_thread": False, "timeout": 30})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    for module in (claims_routes, runner, profile, settings_store, worker):
        monkeypatch.setattr(module, "SessionLocal", Session)
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    real_start = runner.start_background

    def dispatch(coro, run_id=None):
        if _LIVE["on"]:
            return real_start(coro, run_id)
        coro.close()
    monkeypatch.setattr(runner, "start_background", dispatch)
    _LIVE["on"] = False
    yield Session
    _LIVE["on"] = False


@pytest.fixture()
def quiet_startup(monkeypatch):
    """Entering the app's lifespan must not touch the developer's real
    database: only the loop registration is wanted from it."""
    from app.pipeline import runner as pipeline_runner

    monkeypatch.setattr(main, "init_db", lambda: None)
    monkeypatch.setattr(pipeline_runner, "fail_interrupted_runs", lambda: 0)


# ---- #13 the dispatch itself ------------------------------------------------


def test_unknown_api_get_is_a_json_404_not_the_spa():
    response = client.get("/api/this-route-does-not-exist")
    assert response.status_code == 404
    assert response.headers["content-type"].startswith("application/json")
    assert response.json() == {"detail": "API route not found."}

@needs_sample
@pytest.mark.asyncio
async def test_a_sync_route_really_dispatches_its_background_stage(db, monkeypatch, quiet_startup):
    """`set_artifact_disposition` is a plain `def`, so FastAPI runs it in a
    worker thread where no event loop is running. The re-verification it
    starts must still reach the application's loop — un-stubbed, the way it
    runs in production — and the case must actually be verified again."""
    run_id = await run_client_a(db, monkeypatch)
    got = client.get(f"/api/claims-runs/{run_id}").json()
    assert got["status"] == "ready"
    s = db()
    receipt = next(a for a in s.query(ClaimSourceArtifact).filter(ClaimSourceArtifact.run_id == run_id)
                   if a.disposition == "used" and a.media_type == "pdf" and a.case_id)
    case = s.get(ClaimCase, receipt.case_id)
    emp_id = case.legacy_employee_id
    assert receipt.path in case.roles["receipt_files"]
    assert s.get(ClaimEmployee, emp_id).status == "verified"
    s.close()

    _LIVE["on"] = True
    with TestClient(app):  # the lifespan runs → runner.set_loop() on the loop thread
        assert runner._loop is not None, "startup did not register the event loop"
        r = client.post(f"/api/claims-runs/{run_id}/artifacts/{receipt.artifact_id}/disposition",
                        json={"disposition": "irrelevant", "reason": "a personal receipt",
                              "expected_revision": got["revision"]})
        assert r.status_code == 200, r.text
        assert r.json()["reverify_employee_id"] == emp_id
        assert runner.wait_background(60), "the re-verification never ran"
    s = db()
    emp = s.get(ClaimEmployee, emp_id)
    assert emp.status == "verified", emp.error       # re-verified, not left pending
    assert receipt.path not in emp.roles["receipt_files"]
    assert s.get(ClaimsRun, run_id).status == "ready"
    s.close()


@pytest.mark.asyncio
async def test_a_background_stage_that_dies_fails_its_run(db, monkeypatch):
    """The loop keeps only a weak reference to a task and nobody reads its
    exception: a stage that throws used to vanish, leaving the run in an
    in-progress status until the next restart."""
    monkeypatch.setattr(runner, "_loop", None)  # restored when the test ends
    s = db()
    s.add(ClaimsRun(id="rb", client="c", status="verifying"))
    s.commit()
    s.close()

    async def boom(run_id: str) -> None:
        raise RuntimeError("the stage exploded")

    runner.set_loop()
    _LIVE["on"] = True
    runner.start_background(boom("rb"))
    assert runner.pending_background() == 1
    for _ in range(500):
        if runner.pending_background() == 0:
            break
        import asyncio

        await asyncio.sleep(0.01)
    s = db()
    run = s.get(ClaimsRun, "rb")
    assert run.status == "failed" and "the stage exploded" in run.error
    s.close()


def test_a_stage_with_nowhere_to_run_is_a_loud_error(monkeypatch):
    """Dropping the stage silently would leave the run stuck with no sign
    of why: no loop anywhere is a programming error and says so. (No `db`
    fixture here, so start_background is the real one.)"""
    monkeypatch.setattr(runner, "_loop", None)

    async def stage(run_id: str) -> None:
        pass

    with pytest.raises(RuntimeError, match="set_loop"):
        runner.start_background(stage("x"))


# ---- #14 compare-and-set ------------------------------------------------------

def test_two_requests_with_the_same_revision_only_one_wins(db, monkeypatch):
    """Read-compare-write is not a control: both requests read revision 0,
    both find it current, both commit — and the second silently overwrites
    the first. The claim has to be made by the database in one statement."""
    s = db()
    s.add(ClaimsRun(id="rc", client="c", status="ready", revision=0))
    s.add(ClaimEmployee(id="ec", run_id="rc", folder="A", name="A", status="verified", roles={}))
    s.commit()
    s.close()

    # Both requests are held until both have loaded the run at revision 0 —
    # otherwise the race is a matter of scheduling luck.
    route_module = sys.modules[claims_routes.set_employee_category.__module__]
    real_check = route_module._revision_check
    gate = threading.Barrier(2, timeout=20)

    def gated(*a, **k):
        gate.wait()
        return real_check(*a, **k)
    monkeypatch.setattr(route_module, "_revision_check", gated)

    results: list[int] = []
    lock = threading.Lock()

    def go(category: str) -> None:
        c = TestClient(app)
        r = c.put("/api/claims-runs/rc/employees/ec/category",
                  json={"category": category, "gl": "1", "reason": "x", "expected_revision": 0})
        with lock:
            results.append(r.status_code)

    threads = [threading.Thread(target=go, args=(name,)) for name in ("Taxi", "Meals")]
    for t in threads:
        t.start()
    for t in threads:
        t.join(30)
    assert sorted(results) == [200, 409], results
    s = db()
    assert s.get(ClaimsRun, "rc").revision == 1
    s.close()


# ---- #15 a retry that throws ----------------------------------------------------

@pytest.mark.asyncio
async def test_a_retry_that_throws_fails_the_run(db, monkeypatch):
    s = db()
    s.add(ClaimsRun(id="rr", client="c", status="ready"))
    s.add(ClaimEmployee(id="er", run_id="rr", folder="A", name="A", status="failed", roles={}))
    s.commit()
    s.close()

    async def explode(run_id, employee_id):
        raise RuntimeError("the retry exploded")
    monkeypatch.setattr(worker, "verify_employee", explode)

    await worker.retry_employee("rr", "er")     # must not raise
    s = db()
    run = s.get(ClaimsRun, "rr")
    assert run.status == "failed" and "the retry exploded" in run.error, run.error
    s.close()


# ---- #9 a cancel reaches a worker inside _work ----------------------------------

@needs_sample
@pytest.mark.asyncio
async def test_cancelling_mid_verification_stops_the_workers_and_writes_nothing(db, monkeypatch):
    """The only status check used to be at the worker's door: everything
    after it — up to two hundred page reads, the tie-break and the category
    judge — ran on regardless and then committed rows onto a failed run."""
    run_id = await run_client_a(db, monkeypatch, verify=False)
    s = db()
    assert s.get(ClaimsRun, run_id).status == "verifying"
    s.close()

    from app.claims import evidence as evidence_mod

    real_bundle = evidence_mod.read_bundle
    stopped = threading.Event()

    async def cancel_then_read(path, rel, usage, sem, context=""):
        if not stopped.is_set():
            stopped.set()
            s = db()
            run = s.get(ClaimsRun, run_id)
            run.status, run.error = "failed", "cancelled by the reviewer"
            s.commit()
            s.close()
        return await real_bundle(path, rel, usage, sem, context=context)
    monkeypatch.setattr(evidence_mod, "read_bundle", cancel_then_read)

    await runner.start_verification(run_id)

    s = db()
    run = s.get(ClaimsRun, run_id)
    assert run.status == "failed", "a cancelled run must never be resurrected as ready"
    employees = s.query(ClaimEmployee).filter(ClaimEmployee.run_id == run_id).all()
    assert not any(e.status == "verifying" for e in employees), "a worker was left mid-flight"
    assert any(e.status == "failed" and "stopped" in (e.error or "") for e in employees), \
        [(e.name, e.status, e.error) for e in employees]
    # Nothing a stopped worker had read was committed.
    for e in employees:
        if e.status != "verified":
            assert s.query(ClaimRow).filter(ClaimRow.employee_id == e.id).count() == 0
            assert s.query(ClaimEvidence).filter(ClaimEvidence.employee_id == e.id).count() == 0
    s.close()


# ---- #16 a malformed body is a 422, not a 500 --------------------------------

def _map_ready_run(db, **kw) -> str:
    s = db()
    run = ClaimsRun(id=kw.pop("id", "rm"), client="c", status=kw.pop("status", "map_ready"),
                    survey={"files": [], "folders": []}, map={"employees": []},
                    listing_headers={"state": "missing"}, snapshot={}, **kw)
    s.add(run)
    s.commit()
    rid = run.id
    s.close()
    return rid


def test_a_malformed_confirm_map_body_is_a_422_not_a_500(db):
    """`body["map"]["employees"][0]["folder"]` used to be reached into
    directly: the wrong shape came back as the server's own error."""
    run_id = _map_ready_run(db)
    r = client.post(f"/api/claims-runs/{run_id}/confirm-map", json={"map": "not an object"})
    assert r.status_code == 422, r.text
    r = client.post(f"/api/claims-runs/{run_id}/confirm-map", json={})
    assert r.status_code == 422 and any(d["loc"][-1] == "map" for d in r.json()["detail"])
    # an employee with no folder: the field is named, and nothing 500s
    r = client.post(f"/api/claims-runs/{run_id}/confirm-map",
                    json={"map": {"employees": [{"name": "Aegene Ong"}]}})
    assert r.status_code == 422, r.text
    assert any("folder" in d["loc"] for d in r.json()["detail"]), r.json()


def test_malformed_review_bodies_are_422(db):
    run_id = _map_ready_run(db, id="rv", status="ready")
    s = db()
    s.add(ClaimEmployee(id="ev", run_id="rv", folder="A", name="A", status="verified", roles={}))
    s.commit()
    s.close()
    # fields must be an object of field -> value, not a list
    r = client.post("/api/claims-runs/rv/rows/x/correct",
                    json={"fields": ["amount"], "reason": "x", "expected_revision": 0})
    assert r.status_code == 422, r.text
    # artifact_ids must be a list of ids
    r = client.post("/api/claims-runs/rv/cases", json={"label": "x", "artifact_ids": "a1",
                                                      "expected_revision": 0})
    assert r.status_code in (400, 404, 422), r.text
    # a category that is not text
    r = client.put("/api/claims-runs/rv/employees/ev/category",
                   json={"category": {"a": 1}, "expected_revision": 0})
    assert r.status_code == 422, r.text


# ---- feature-flag gating and run-status guards -------------------------------

def test_case_routes_are_absent_not_refused_when_the_case_model_is_off(db, monkeypatch):
    run_id = _map_ready_run(db, id="rf")
    monkeypatch.setattr(config, "CLAIMS_CASE_MODEL", False)
    monkeypatch.setattr(config, "CLAIMS_FULL_DUMP_GROUPING", True)
    r = client.post(f"/api/claims-runs/{run_id}/cases",
                    json={"label": "x", "artifact_ids": [], "expected_revision": 0})
    assert r.status_code == 404, r.text
    # with the case model on but regrouping off, the route EXISTS and refuses
    monkeypatch.setattr(config, "CLAIMS_CASE_MODEL", True)
    monkeypatch.setattr(config, "CLAIMS_FULL_DUMP_GROUPING", False)
    r = client.post(f"/api/claims-runs/{run_id}/cases",
                    json={"label": "x", "artifact_ids": [], "expected_revision": 0})
    assert r.status_code == 400 and "switched off" in r.text


def test_review_actions_need_a_ready_run(db):
    """A worker deletes and rewrites an employee's flags wholesale, so a
    decision recorded while it runs lands on rows that are about to go."""
    from app.claims.models import ClaimFlag

    run_id = _map_ready_run(db, id="rg", status="verifying")
    s = db()
    s.add(ClaimEmployee(id="eg", run_id="rg", folder="A", name="A", status="verifying", roles={}))
    s.add(ClaimFlag(id="fg", run_id="rg", employee_id="eg", code="NO_RECEIPT", reason="r", basis="b"))
    s.commit()
    s.close()
    r = client.post("/api/claims-runs/rg/flags/fg/decide",
                    json={"decision": "dismissed", "note": "x", "expected_revision": 0})
    assert r.status_code == 400 and "ready" in r.text, r.text
    r = client.put("/api/claims-runs/rg/employees/eg/category",
                   json={"category": "Taxi", "expected_revision": 0})
    assert r.status_code == 400 and "ready" in r.text, r.text
    s = db()
    assert s.get(ClaimsRun, "rg").revision == 0, "a refused action must not consume a revision"
    s.close()


def test_a_skipped_employee_is_not_re_verified(db):
    run_id = _map_ready_run(db, id="rs", status="ready")
    s = db()
    s.add(ClaimEmployee(id="es", run_id="rs", folder="A", name="A", status="skipped", roles={}))
    s.commit()
    s.close()
    r = client.post("/api/claims-runs/rs/employees/es/retry", json={"expected_revision": 0})
    assert r.status_code == 400 and "skipped" in r.text, r.text


# ---- /file ------------------------------------------------------------------

def test_the_file_route_refuses_escapes_and_names_unshowable_types(db, monkeypatch):
    run_id = _map_ready_run(db, id="rp", status="ready")
    files = runner.files_dir(run_id)
    files.mkdir(parents=True, exist_ok=True)
    (files / "page.png").write_bytes(scripted._ONE_PIXEL_PNG)
    (files / "notes.txt").write_text("a note nobody can rasterise")

    assert client.get(f"/api/claims-runs/{run_id}/file?path=page.png").status_code == 200
    # a real file of a kind that has no page image: 415, not the server's
    # own error
    r = client.get(f"/api/claims-runs/{run_id}/file?path=notes.txt")
    assert r.status_code == 415, r.text
    # out of the workspace, three ways
    for path in ("../../../../etc/passwd", "/etc/passwd", "sub/../../outside.pdf"):
        r = client.get(f"/api/claims-runs/{run_id}/file", params={"path": path})
        assert r.status_code == 404, (path, r.status_code)
    assert client.get(f"/api/claims-runs/{run_id}/file?path=page.png&page=0").status_code == 404
    # a box that is not four finite numbers is a 400, never a 500
    for bad in ("1,2,3", "a,b,c,d", "inf,0,50,50", "nan,0,50,50"):
        assert client.get(f"/api/claims-runs/{run_id}/file", params={"path": "page.png", "box": bad}).status_code == 400, bad
    assert client.get(f"/api/claims-runs/{run_id}/file", params={"path": "page.png", "box": "10,10,60,60"}).status_code == 200


# ---- /sheet -----------------------------------------------------------------

def test_the_sheet_route_returns_the_header_and_the_rows_around_the_cited_line(db):
    import datetime
    import openpyxl

    run_id = _map_ready_run(db, id="rs", status="ready")
    files = runner.files_dir(run_id)
    files.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    ws.append(["Name:", "Aegene Ong"])
    ws.append([])
    ws.append(["Date", "Expense Item", "Itemized Receipt Included (Y/N)", "Total (MYR)"])
    for n in range(1, 30):
        ws.append([datetime.datetime(2026, 7, n % 28 + 1), f"Taxi {n}", "Y", 10.0 * n + 0.5])
    wb.create_sheet("KM")
    wb.save(files / "report.xlsx")
    (files / "notes.txt").write_text("not a workbook")

    r = client.get(f"/api/claims-runs/{run_id}/sheet", params={"path": "report.xlsx", "sheet": "Expense Report", "row": 20})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["sheet"] == "Expense Report" and body["sheets"] == ["Expense Report", "KM"]
    assert body["focus"] == 20
    # the all-words header row is found even though it is outside the window
    assert body["header"]["n"] == 3 and body["header"]["cells"][0] == "Date"
    numbers = [row["n"] for row in body["rows"]]
    assert numbers[0] == 14 and numbers[-1] == 24 and 20 in numbers
    focus = next(row for row in body["rows"] if row["n"] == 20)
    # dates read as dates, money without a trailing .0 artefact, no formulas, no guesses
    assert focus["cells"][0] == "2026-07-18" and focus["cells"][3] == "170.5"
    assert body["columns"] == ["A", "B", "C", "D"]

    # a row number a sheet could never have is refused, not walked to
    assert client.get(f"/api/claims-runs/{run_id}/sheet", params={"path": "report.xlsx", "sheet": "Expense Report", "row": 10 ** 9}).status_code == 400
    # a text-only data row between the header and the cited line does not
    # displace the header (rows are judged by their cells' types)
    ws2 = openpyxl.load_workbook(files / "report.xlsx")["Expense Report"]
    ws2.insert_rows(6)
    for col, text in enumerate(["note", "carried over from June", "see e-mail", "n/a"], 1):
        ws2.cell(row=6, column=col, value=text)
    ws2.parent.save(files / "report.xlsx")
    r2 = client.get(f"/api/claims-runs/{run_id}/sheet", params={"path": "report.xlsx", "sheet": "Expense Report", "row": 21})
    assert r2.status_code == 200 and r2.json()["header"]["n"] == 3, r2.text
    assert client.get(f"/api/claims-runs/{run_id}/sheet", params={"path": "report.xlsx", "sheet": "Nope", "row": 2}).status_code == 404
    assert client.get(f"/api/claims-runs/{run_id}/sheet", params={"path": "notes.txt"}).status_code == 415
    for path in ("../../../../etc/passwd", "/etc/passwd", "sub/../../outside.xlsx"):
        assert client.get(f"/api/claims-runs/{run_id}/sheet", params={"path": path}).status_code == 404, path


# ---- local-mode ingestion ----------------------------------------------------

def test_a_folder_path_is_refused_unless_an_operator_named_a_root(db, monkeypatch, tmp_path):
    """DOC_SOURCE=local used to accept ANY folder on the machine as the
    batch: an arbitrary-file-read dressed as a convenience."""
    batch = tmp_path / "allowed" / "july"
    batch.mkdir(parents=True)
    monkeypatch.setenv("DOC_SOURCE", "local")
    monkeypatch.setattr(config, "CLAIMS_LOCAL_ROOT", "")
    r = client.post("/api/claims-runs", data={"received_date": "2026-08-03", "folder_url": str(batch)})
    assert r.status_code == 400 and "https://" in r.text
    # named root: inside is accepted, outside is still refused
    monkeypatch.setattr(config, "CLAIMS_LOCAL_ROOT", str(tmp_path / "allowed"))
    r = client.post("/api/claims-runs", data={"received_date": "2026-08-03", "folder_url": str(batch)})
    assert r.status_code == 200, r.text
    outside = tmp_path / "elsewhere"
    outside.mkdir()
    r = client.post("/api/claims-runs", data={"received_date": "2026-08-03", "folder_url": str(outside)})
    assert r.status_code == 400, r.text
